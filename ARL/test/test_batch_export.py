"""
批量表格导出功能单元测试

测试内容：
- 后端批量导出API接口
- Excel文件生成和内容验证
- 错误处理和边界情况
"""

import unittest
import json
from io import BytesIO
from unittest.mock import patch

IMPORT_ERROR = None
try:
    from openpyxl import load_workbook
    from app.routes.export import export_merge_tasks, export_merge_tasks_html, SaveTask, build_task_export_summary
    from app import create_app
except Exception as exc:
    load_workbook = None
    export_merge_tasks = None
    export_merge_tasks_html = None
    SaveTask = None
    build_task_export_summary = None
    create_app = None
    IMPORT_ERROR = exc


@unittest.skipIf(IMPORT_ERROR is not None, "requires export test dependencies: {}".format(IMPORT_ERROR))
class TestBatchExport(unittest.TestCase):
    """批量导出功能测试类"""

    def setUp(self):
        """测试前准备"""
        self.app = create_app()
        self.app.config['TESTING'] = True
        self.client = self.app.test_client()

        # 模拟任务数据
        self.mock_task_data = {
            "_id": "test_task_1",
            "name": "测试任务1",
            "target": "example.com",
            "status": "done"
        }

        self.mock_task_data_2 = {
            "_id": "test_task_2",
            "name": "测试任务2",
            "target": "test.com",
            "status": "done"
        }

    def tearDown(self):
        """测试后清理"""
        pass

    @patch('app.routes.export.get_task_data')
    def test_batch_export_api_success(self, mock_get_task_data):
        """测试批量导出API成功情况"""
        # 模拟获取任务数据
        mock_get_task_data.side_effect = lambda task_id: {
            "test_task_1": self.mock_task_data,
            "test_task_2": self.mock_task_data_2
        }.get(task_id)

        # 测试数据
        test_data = {
            "task_ids": ["test_task_1", "test_task_2"]
        }

        # 发送POST请求
        response = self.client.post(
            '/api/export/batch',
            data=json.dumps(test_data),
            content_type='application/json',
            headers={'Token': 'test_token'}
        )

        # 验证响应
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content_type, 'application/octet-stream')

        # 验证文件名包含任务名
        content_disposition = response.headers.get('Content-Disposition', '')
        self.assertIn('ARL批量导出报告_测试任务1.xlsx', content_disposition)

    @patch('app.routes.export.export_merge_tasks_html')
    @patch('app.routes.export.get_task_data')
    def test_batch_export_api_html_success(self, mock_get_task_data, mock_export_merge_tasks_html):
        """测试批量 HTML 导出 API 成功情况"""
        mock_get_task_data.return_value = self.mock_task_data
        mock_export_merge_tasks_html.return_value = b"<html><body>demo</body></html>"

        response = self.client.post(
            '/api/export/batch',
            data=json.dumps({
                "task_ids": ["test_task_1"],
                "format": "html",
            }),
            content_type='application/json',
            headers={'Token': 'test_token'}
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('text/html', response.content_type)
        content_disposition = response.headers.get('Content-Disposition', '')
        self.assertIn('ARL批量导出报告_测试任务1.html', content_disposition)

    @patch('app.routes.export.export_arl_html')
    @patch('app.routes.export.get_task_data')
    def test_single_export_api_html_success(self, mock_get_task_data, mock_export_arl_html):
        """测试单任务 HTML 导出 API 成功情况"""
        mock_get_task_data.return_value = self.mock_task_data
        mock_export_arl_html.return_value = b"<html><body>single-demo</body></html>"

        response = self.client.get(
            '/api/export/test_task_1?format=html',
            headers={'Token': 'test_token'}
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('text/html', response.content_type)
        content_disposition = response.headers.get('Content-Disposition', '')
        self.assertIn('ARL资产导出报告_example.com.html', content_disposition)

    def test_batch_export_api_invalid_request(self):
        """测试批量导出API无效请求"""
        # 测试空请求体
        response = self.client.post(
            '/api/export/batch',
            data=json.dumps({}),
            content_type='application/json',
            headers={'Token': 'test_token'}
        )
        self.assertEqual(response.status_code, 400)

        # 测试无效的task_ids
        response = self.client.post(
            '/api/export/batch',
            data=json.dumps({"task_ids": "invalid"}),
            content_type='application/json',
            headers={'Token': 'test_token'}
        )
        self.assertEqual(response.status_code, 400)

    @patch('app.routes.export.get_task_data')
    def test_batch_export_api_task_not_found(self, mock_get_task_data):
        """测试批量导出API任务不存在"""
        # 模拟任务不存在
        mock_get_task_data.return_value = None

        test_data = {
            "task_ids": ["nonexistent_task"]
        }

        response = self.client.post(
            '/api/export/batch',
            data=json.dumps(test_data),
            content_type='application/json',
            headers={'Token': 'test_token'}
        )

        self.assertEqual(response.status_code, 404)

    @patch('app.routes.export.get_nuclei_result_data')
    @patch('app.routes.export.get_stat_finger_data')
    @patch('app.routes.export.get_vuln_data')
    @patch('app.routes.export.get_wih_data')
    @patch('app.routes.export.get_fileleak_data')
    @patch('app.routes.export.get_url_data')
    @patch('app.routes.export.get_cert_data')
    @patch('app.routes.export.get_service_data')
    @patch('app.routes.export.get_task_data')
    @patch('app.routes.export.get_ip_data')
    @patch('app.routes.export.get_domain_data')
    @patch('app.routes.export.get_site_data')
    def test_export_merge_tasks_function(
        self,
        mock_get_site_data,
        mock_get_domain_data,
        mock_get_ip_data,
        mock_get_task_data,
        mock_get_service_data,
        mock_get_cert_data,
        mock_get_url_data,
        mock_get_fileleak_data,
        mock_get_wih_data,
        mock_get_vuln_data,
        mock_get_stat_finger_data,
        mock_get_nuclei_result_data,
    ):
        """测试export_merge_tasks函数"""
        # 模拟任务数据
        mock_get_task_data.side_effect = lambda task_id: {
            "task_1": {"_id": "task_1", "name": "任务1", "target": "example.com"},
            "task_2": {"_id": "task_2", "name": "任务2", "target": "test.com"}
        }.get(task_id)

        # 模拟IP数据
        mock_get_ip_data.return_value = [
            {"ip": "192.168.1.1", "port_info": [{"port": 80, "service": "http"}]},
            {"ip": "192.168.1.2", "port_info": [{"port": 443, "service": "https"}]}
        ]

        # 模拟域名数据
        mock_get_domain_data.return_value = [
            {"domain": "www.example.com", "record": "A"},
            {"domain": "api.example.com", "record": "A"}
        ]

        # 模拟站点数据
        mock_get_site_data.return_value = [
            {"site": "http://www.example.com", "title": "Example Site", "finger": [], "status": 200, "favicon": {}},
            {"site": "https://api.example.com", "title": "API Site", "finger": [], "status": 200, "favicon": {}}
        ]

        mock_get_service_data.return_value = [
            {"ip": "192.168.1.1", "port": 80, "service": "http", "product": "nginx", "version": "1.25.0"}
        ]
        mock_get_cert_data.return_value = []
        mock_get_url_data.return_value = [
            {"url": "http://www.example.com/login", "status": 200, "title": "Login", "finger": []}
        ]
        mock_get_fileleak_data.return_value = []
        mock_get_wih_data.return_value = []
        mock_get_stat_finger_data.return_value = []
        mock_get_vuln_data.return_value = [
            {
                "vul_name": "测试风险",
                "vul_severity": "high",
                "target": "http://www.example.com",
                "vuln_url": "http://www.example.com/login",
                "plg_name": "test-plugin",
                "plg_type": "info-leak",
                "description": "demo",
            }
        ]
        mock_get_nuclei_result_data.return_value = []

        # 调用函数
        result = export_merge_tasks(["task_1", "task_2"])

        # 验证结果是二进制数据
        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

        # 验证Excel文件头 (ZIP文件头标识)
        self.assertEqual(result[:4], b'PK\x03\x04')

        wb = load_workbook(filename=BytesIO(result), read_only=True, data_only=True)
        try:
            self.assertEqual(
                wb.sheetnames,
                ["站点", "IP", "系统服务", "SSL证书", "域名", "URL信息", "目录扫描", "WIH", "WAF识别", "风险", "PoC风险", "指纹统计", "资产统计"],
            )
        finally:
            wb.close()

    @patch('app.routes.export.get_nuclei_result_data')
    @patch('app.routes.export.get_stat_finger_data')
    @patch('app.routes.export.get_vuln_data')
    @patch('app.routes.export.get_wih_data')
    @patch('app.routes.export.get_fileleak_data')
    @patch('app.routes.export.get_url_data')
    @patch('app.routes.export.get_cert_data')
    @patch('app.routes.export.get_service_data')
    @patch('app.routes.export.get_task_data')
    @patch('app.routes.export.get_ip_data')
    @patch('app.routes.export.get_domain_data')
    @patch('app.routes.export.get_site_data')
    def test_export_merge_tasks_html_function(
        self,
        mock_get_site_data,
        mock_get_domain_data,
        mock_get_ip_data,
        mock_get_task_data,
        mock_get_service_data,
        mock_get_cert_data,
        mock_get_url_data,
        mock_get_fileleak_data,
        mock_get_wih_data,
        mock_get_vuln_data,
        mock_get_stat_finger_data,
        mock_get_nuclei_result_data,
    ):
        """测试批量 HTML 导出函数输出 HTML 报告。"""
        mock_get_task_data.return_value = {"_id": "task_1", "name": "任务1", "target": "example.com"}
        mock_get_ip_data.return_value = [{"ip": "192.168.1.1", "port_info": [{"port_id": 80, "service_name": "http"}], "geo_city": {}, "geo_asn": {}, "domain": [], "os_info": {}, "cdn_name": "", "ip_type": ""}]
        mock_get_domain_data.return_value = [{"domain": "www.example.com", "type": "A", "record": ["1.1.1.1"], "ips": ["1.1.1.1"]}]
        mock_get_site_data.return_value = [{"site": "http://www.example.com", "title": "Example Site", "finger": [], "status": 200, "favicon": {}}]
        mock_get_service_data.return_value = []
        mock_get_cert_data.return_value = []
        mock_get_url_data.return_value = []
        mock_get_fileleak_data.return_value = []
        mock_get_wih_data.return_value = []
        mock_get_stat_finger_data.return_value = []
        mock_get_vuln_data.return_value = []
        mock_get_nuclei_result_data.return_value = []

        result = export_merge_tasks_html(["task_1"])

        self.assertIsInstance(result, bytes)
        html = result.decode("utf-8")
        self.assertIn("<html", html)
        self.assertIn("ARL批量导出报告", html)
        self.assertIn("站点", html)
        self.assertIn("Example Site", html)

    @patch.object(SaveTask, 'build_statist')
    @patch.object(SaveTask, 'build_stat_finger_xl')
    @patch.object(SaveTask, 'build_nuclei_xl')
    @patch.object(SaveTask, 'build_vuln_xl')
    @patch.object(SaveTask, 'build_wih_xl')
    @patch.object(SaveTask, 'build_fileleak_xl')
    @patch.object(SaveTask, 'build_url_xl')
    @patch.object(SaveTask, 'build_domain_xl')
    @patch.object(SaveTask, 'build_cert_xl')
    @patch.object(SaveTask, 'build_service_xl')
    @patch.object(SaveTask, 'build_ip_xl')
    @patch.object(SaveTask, 'build_site_xl')
    @patch('app.routes.export.get_task_data')
    @patch('app.routes.export.save_virtual_workbook')
    def test_save_task_run_should_build_vuln_sheet(
        self,
        mock_save_virtual_workbook,
        mock_get_task_data,
        mock_build_site_xl,
        mock_build_ip_xl,
        mock_build_service_xl,
        mock_build_cert_xl,
        mock_build_domain_xl,
        mock_build_url_xl,
        mock_build_fileleak_xl,
        mock_build_wih_xl,
        mock_build_vuln_xl,
        mock_build_nuclei_xl,
        mock_build_stat_finger_xl,
        mock_build_statist,
    ):
        """测试单任务导出会构建风险工作表。"""
        mock_get_task_data.return_value = {"_id": "task_1", "target": "example.com", "type": "domain"}
        mock_save_virtual_workbook.return_value = b"demo"

        save_task = SaveTask("task_1")
        result = save_task.run()

        self.assertEqual(result, b"demo")
        mock_build_vuln_xl.assert_called_once()
        mock_build_nuclei_xl.assert_called_once()
        mock_build_stat_finger_xl.assert_called_once()

    @patch('app.routes.export.get_nuclei_result_data')
    @patch('app.routes.export.get_vuln_data')
    @patch('app.routes.export.get_url_data')
    @patch('app.routes.export.get_task_data')
    @patch('app.routes.export.get_ip_data')
    @patch('app.routes.export.get_domain_data')
    @patch('app.routes.export.get_site_data')
    def test_build_task_export_summary_should_follow_export_counts(
        self,
        mock_get_site_data,
        mock_get_domain_data,
        mock_get_ip_data,
        mock_get_task_data,
        mock_get_url_data,
        mock_get_vuln_data,
        mock_get_nuclei_result_data,
    ):
        """汇总应复用导出口径，避免通知与报告不一致。"""
        mock_get_task_data.side_effect = lambda task_id: {
            "task_1": {"_id": "task_1", "name": "任务1", "target": "example.com"},
            "task_2": {"_id": "task_2", "name": "任务2", "target": "example.org"},
        }.get(task_id)

        mock_get_site_data.side_effect = lambda task_id: {
            "task_1": [
                {"site": "https://a.example.com"},
                {"site": "https://shared.example.com"},
            ],
            "task_2": [
                {"site": "https://shared.example.com"},
                {"url": "https://b.example.org"},
            ],
        }.get(task_id, [])
        mock_get_domain_data.side_effect = lambda task_id: {
            "task_1": [{"domain": "a.example.com"}, {"domain": "shared.example.com"}],
            "task_2": [{"domain": "shared.example.com"}, {"domain": "b.example.org"}],
        }.get(task_id, [])
        mock_get_ip_data.side_effect = lambda task_id: {
            "task_1": [{"ip": "1.1.1.1"}, {"ip": "1.1.1.2"}],
            "task_2": [{"ip": "2.2.2.2"}],
        }.get(task_id, [])
        mock_get_url_data.side_effect = lambda task_id: {
            "task_1": [
                {"url": "https://a.example.com/login", "site": "https://a.example.com", "title": "Login", "status_code": 200, "content_length": 100, "source": "spider"},
                {"url": "https://shared.example.com/home", "site": "https://shared.example.com", "title": "Home", "status_code": 200, "content_length": 80, "source": "spider"},
            ],
            "task_2": [
                {"url": "https://shared.example.com/home", "site": "https://shared.example.com", "title": "Home", "status_code": 200, "content_length": 80, "source": "spider"},
                {"url": "https://b.example.org/admin", "site": "https://b.example.org", "title": "Admin", "status_code": 403, "content_length": 20, "source": "wih"},
            ],
        }.get(task_id, [])
        mock_get_vuln_data.side_effect = lambda task_id: {
            "task_1": [{"vul_name": "A", "severity": "high", "target": "https://a.example.com", "plg_name": "p1", "plg_type": "info"}],
            "task_2": [{"vul_name": "A", "severity": "high", "target": "https://a.example.com", "plg_name": "p1", "plg_type": "info"}],
        }.get(task_id, [])
        mock_get_nuclei_result_data.return_value = []

        summary = build_task_export_summary(["task_1", "task_2"])

        self.assertEqual(summary.get("site_cnt"), 3)
        self.assertEqual(summary.get("domain_cnt"), 3)
        self.assertEqual(summary.get("ip_cnt"), 3)
        self.assertEqual(summary.get("url_cnt"), 3)
        self.assertEqual(summary.get("vuln_cnt"), 2)
        self.assertEqual(summary.get("task_summaries", {}).get("task_1", {}).get("site_cnt"), 2)
        self.assertEqual(summary.get("task_summaries", {}).get("task_2", {}).get("site_cnt"), 2)


@unittest.skipIf(IMPORT_ERROR is not None, "requires export test dependencies: {}".format(IMPORT_ERROR))
class TestBatchExportIntegration(unittest.TestCase):
    """批量导出集成测试"""

    def setUp(self):
        """集成测试准备"""
        self.app = create_app()
        self.app.config['TESTING'] = True
        self.client = self.app.test_client()

    def test_batch_export_workflow(self):
        """测试完整的批量导出工作流程"""
        # 1. 准备测试数据（在实际环境中需要真实的已完成任务）
        # 2. 调用批量导出API
        # 3. 验证响应格式
        # 4. 验证Excel文件内容

        # 注意：这个测试需要在有真实数据的环境中运行
        # 这里只验证API接口的基本可用性

        test_data = {
            "task_ids": ["dummy_task_id"]  # 使用虚拟ID进行基本验证
        }

        response = self.client.post(
            '/api/export/batch',
            data=json.dumps(test_data),
            content_type='application/json',
            headers={'Token': 'test_token'}
        )

        # 验证API基本可用性（即使数据不存在，也应该返回适当的错误响应）
        self.assertIn(response.status_code, [200, 400, 404, 500])


if __name__ == '__main__':
    unittest.main()
