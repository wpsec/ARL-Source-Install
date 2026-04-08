#coding: utf-8
"""
任务报告导出模块

功能说明：
- 导出任务扫描结果为Excel报告
- 包含完整的统计分析和数据汇总
- 提供可视化的资产信息展示

报告内容：
1. 任务概览：任务名称、目标、时间、配置等
2. IP统计：IP总数、端口分布、服务分布
3. 域名统计：域名总数、类型分布
4. 站点统计：站点总数、状态码分布、指纹分布
5. 详细数据：完整的IP、域名、站点、服务列表

导出格式：
- Excel (.xlsx) 文件
- 多个工作表分类展示数据
- 包含样式和格式化
"""

from flask import make_response, request, send_file
from flask_restx import Resource, Namespace
from openpyxl import Workbook
from bson import ObjectId
import re
import json
import os
from pathlib import Path
from datetime import datetime, timedelta
from collections import Counter
from html import escape
import ipaddress
import yaml
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from app.utils import get_logger, auth
from app.utils.tls_policy import get_ssl_security_compliance
from app import utils
from app.config import Config
from app.modules import CeleryAction
from urllib.parse import quote, urlparse

ns = Namespace('export', description="任务报告导出接口")

logger = get_logger()

MONGO_EXPORT_BATCH_SIZE = 500
EXPORT_JOB_STATUS_QUEUED = "queued"
EXPORT_JOB_STATUS_RUNNING = "running"
EXPORT_JOB_STATUS_DONE = "done"
EXPORT_JOB_STATUS_ERROR = "error"
TASK_EXPORT_PROJECTION = {
    "_id": 1,
    "target": 1,
    "name": 1,
    "type": 1,
    "start_time": 1,
    "end_time": 1,
    "waf_skip_summary": 1,
}
IP_EXPORT_PROJECTION = {
    "task_id": 1,
    "ip": 1,
    "port_info": 1,
    "geo_city": 1,
    "geo_asn": 1,
    "domain": 1,
    "os_info": 1,
    "cdn_name": 1,
    "ip_type": 1,
}
SITE_EXPORT_PROJECTION = {
    "task_id": 1,
    "site": 1,
    "url": 1,
    "title": 1,
    "headers": 1,
    "finger": 1,
    "screenshot": 1,
    "status": 1,
    "favicon": 1,
}
DOMAIN_EXPORT_PROJECTION = {
    "task_id": 1,
    "domain": 1,
    "type": 1,
    "record": 1,
    "ips": 1,
    "source": 1,
}
URL_EXPORT_PROJECTION = {
    "url": 1,
    "site": 1,
    "title": 1,
    "status_code": 1,
    "content_length": 1,
    "source": 1,
}
FILELEAK_EXPORT_PROJECTION = {
    "url": 1,
    "site": 1,
    "title": 1,
    "status_code": 1,
    "content_length": 1,
    "source": 1,
}
WIH_EXPORT_PROJECTION = {
    "record_type": 1,
    "content": 1,
    "source": 1,
    "site": 1,
}
WIH_ENDPOINT_EXPORT_PROJECTION = {
    "task_id": 1,
    "target": 1,
    "site": 1,
    "page_url": 1,
    "url": 1,
    "request_url": 1,
    "method": 1,
    "status_code": 1,
    "response_status": 1,
    "response_size": 1,
    "content_length": 1,
    "request_packet": 1,
    "request_template": 1,
}
SERVICE_EXPORT_PROJECTION = {
    "task_id": 1,
    "service_name": 1,
    "service_info": 1,
}
VULN_EXPORT_PROJECTION = {
    "task_id": 1,
    "vul_name": 1,
    "severity": 1,
    "target": 1,
    "plg_name": 1,
    "plg_type": 1,
    "description": 1,
    "detail": 1,
    "verify_data": 1,
}
NUCLEI_RESULT_EXPORT_PROJECTION = {
    "task_id": 1,
    "scanner_type": 1,
    "rule_id": 1,
    "vuln_name": 1,
    "vuln_severity": 1,
    "target": 1,
    "vuln_url": 1,
    "save_date": 1,
    "verify_data": 1,
    "template_id": 1,
    "template_url": 1,
}
STAT_FINGER_EXPORT_PROJECTION = {
    "task_id": 1,
    "name": 1,
    "cnt": 1,
}
CERT_EXPORT_PROJECTION = {
    "task_id": 1,
    "ip": 1,
    "port": 1,
    "host": 1,
    "scan_mode": 1,
    "sni_domain": 1,
    "domain": 1,
    "domains": 1,
    "cert": 1,
}
AI_PEN_TEST_EXPORT_PROJECTION = {
    "task_id": 1,
    "source_collection": 1,
    "source_module": 1,
    "risk_type": 1,
    "risk_name": 1,
    "target": 1,
    "vuln_url": 1,
    "decision": 1,
    "confidence": 1,
    "status": 1,
    "verification_step": 1,
    "payload_type": 1,
    "payload_variant": 1,
    "payload_expected_signal": 1,
    "payload": 1,
    "request_method": 1,
    "request_url": 1,
    "request_path": 1,
    "request_packet": 1,
    "request_template_mode": 1,
    "request_template_content_type": 1,
    "request_template_summary": 1,
    "evidence_snippet": 1,
    "http_status": 1,
    "response_hash_diff": 1,
    "proof_family": 1,
    "proof_type": 1,
    "proof_strength": 1,
    "unauth_access_hit": 1,
    "unauth_access_type": 1,
    "unauth_access_reason": 1,
    "unauth_probe_summary": 1,
    "unauth_negative_type": 1,
    "decision_guard_action": 1,
    "decision_guard_reason": 1,
    "proof_signals": 1,
    "proof_summary": 1,
    "reason": 1,
    "high_value_summary": 1,
    "high_value_family": 1,
    "high_value_family_rank": 1,
    "high_value_keywords": 1,
    "tool_trace": 1,
    "agent_trace": 1,
    "tool_calls": 1,
    "tool_results": 1,
    "stop_reason": 1,
    "budget_used": 1,
    "runtime_version": 1,
    "session_summary": 1,
    "tool_plan_source": 1,
    "ai_status": 1,
    "ai_plan_decision": 1,
    "ai_plan_confidence": 1,
    "ai_plan_reason": 1,
    "ai_plan_actions": 1,
    "ai_plan_tool_plan": 1,
    "ai_plan_request": 1,
    "ai_plan_reply": 1,
    "external_tool_hit": 1,
    "external_tool_runs": 1,
    "save_date": 1,
    "update_date": 1,
}
AI_DENOISE_RESULT_EXPORT_PROJECTION = {
    "task_id": 1,
    "module_id": 1,
    "row_key": 1,
    "data_id": 1,
    "result_level": 1,
    "risk_level": 1,
    "trust": 1,
    "display_text": 1,
    "summary": 1,
    "evidence": 1,
    "suggestions": 1,
    "source": 1,
    "prompt_id": 1,
    "prompt_name": 1,
    "note": 1,
    "analyzed_at": 1,
    "updated_at": 1,
}
AI_DENOISE_RESULT_LEVEL_ORDER = {
    "disabled": 0,
    "safe": 1,
    "suspicious": 2,
    "danger": 3,
}
AI_DENOISE_RESULT_LEVEL_LABEL = {
    "disabled": "未分析",
    "safe": "正常",
    "suspicious": "可疑",
    "danger": "危险",
}
AI_DENOISE_SOURCE_ORDER = {
    "disabled": 0,
    "rule": 1,
    "ai": 2,
}
AI_DENOISE_SOURCE_LABEL = {
    "disabled": "未分析",
    "rule": "规则",
    "ai": "AI模型",
}
AI_DENOISE_MODULE_LABEL_MAP = {
    "site": "站点",
    "fileleak": "目录扫描",
    "cert": "SSL证书",
    "url": "URL信息",
    "vuln": "风险",
    "nuclei_result": "PoC风险",
}
AI_DENOISE_MODULE_COLLECTION_MAP = {
    "site": "site",
    "fileleak": "fileleak",
    "cert": "cert",
    "url": "url",
    "vuln": "vuln",
    "nuclei_result": "nuclei_result",
}
AI_DENOISE_MODULE_TARGET_PROJECTION = {
    "site": {"_id": 1, "site": 1, "url": 1, "title": 1},
    "fileleak": {"_id": 1, "url": 1, "site": 1, "title": 1},
    "cert": {"_id": 1, "host": 1, "domain": 1, "ip": 1, "port": 1},
    "url": {"_id": 1, "url": 1, "site": 1, "title": 1},
    "vuln": {"_id": 1, "target": 1, "vul_name": 1},
    "nuclei_result": {"_id": 1, "target": 1, "vuln_url": 1, "vuln_name": 1},
}


def normalize_export_format(value):
    """
    规范化导出格式，兼容 table/excel/xlsx/html/ai_markdown 等输入。
    """
    export_format = sanitize_excel_value(value).strip().lower()
    if export_format in ["html", "htm"]:
        return "html"
    if export_format in ["ai", "ai_markdown", "ai-markdown", "markdown", "md"]:
        return "ai_markdown"
    if export_format in ["table", "excel", "xlsx"]:
        return "excel"
    return "excel"


def build_export_response(file_content, filename, content_type):
    """
    构建统一的文件下载响应。
    """
    response = make_response(file_content)
    response.headers['Content-Type'] = content_type
    response.headers["Content-Disposition"] = "attachment; filename={}".format(quote(filename))
    return response


def _resolve_export_job_dir() -> Path:
    export_dir = Path(getattr(Config, "EXPORT_REPORT_DIR", "") or "")
    if not export_dir:
        export_dir = Path(__file__).resolve().parents[1] / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def _get_export_job_collection():
    return utils.conn_db("export_job")


def _ensure_export_job_indexes():
    try:
        collection = _get_export_job_collection()
        collection.create_index("created_at", background=True)
        collection.create_index("status", background=True)
        collection.create_index("expire_at", expireAfterSeconds=0, background=True)
    except Exception as exc:
        logger.warning("ensure export_job indexes failed: %s", exc)


def _cleanup_stale_export_files():
    export_dir = _resolve_export_job_dir()
    keep_days = max(1, int(getattr(Config, "EXPORT_REPORT_KEEP_DAYS", 3) or 3))
    expire_before = datetime.utcnow() - timedelta(days=keep_days)
    try:
        for item in export_dir.iterdir():
            if not item.is_file():
                continue
            try:
                modified_at = datetime.utcfromtimestamp(item.stat().st_mtime)
            except Exception:
                modified_at = datetime.utcnow()
            if modified_at < expire_before:
                try:
                    item.unlink()
                except Exception:
                    continue
    except Exception as exc:
        logger.warning("cleanup stale export files failed: %s", exc)


def _build_export_content(task_ids, export_format):
    fmt = normalize_export_format(export_format)
    task_id_list = _normalize_task_id_list(task_ids)
    if not task_id_list:
        raise ValueError("task_ids is empty")

    if len(task_id_list) == 1:
        task_id = task_id_list[0]
        task_data = get_task_data(task_id)
        if not task_data:
            raise ValueError("task not found")
        target = sanitize_excel_value(task_data.get("target", "")).strip() or task_id
        target_name = target.replace("/", "_")[:20]
        if fmt == "html":
            return export_arl_html(task_id), "ARL资产导出报告_{}.html".format(target_name), "text/html; charset=utf-8"
        if fmt == "ai_markdown":
            return export_arl_ai_markdown(task_id), "ARL_AI分析报告_{}.md".format(target_name), "text/markdown; charset=utf-8"
        return build_single_task_workbook(task_id, apply_style=True), "ARL资产导出报告_{}.xlsx".format(target_name), "application/octet-stream"

    first_task = get_task_data(task_id_list[0])
    if not first_task:
        raise ValueError("task not found")
    task_name = sanitize_excel_value(first_task.get("name", "未知")).strip()[:20] or "未知"
    if fmt == "html":
        return export_merge_tasks_html(task_id_list), "ARL批量导出报告_{}.html".format(task_name), "text/html; charset=utf-8"
    if fmt == "ai_markdown":
        return export_merge_tasks_ai_markdown(task_id_list), "ARL_AI分析报告_{}.md".format(task_name), "text/markdown; charset=utf-8"
    return build_merge_tasks_workbook(task_id_list, apply_style=True), "ARL批量导出报告_{}.xlsx".format(task_name), "application/octet-stream"


def _write_export_job_file(job_id: str, filename: str, file_content):
    export_dir = _resolve_export_job_dir()
    safe_name = re.sub(r'[^A-Za-z0-9._-]+', "_", str(filename or "arl-export.bin")).strip("._") or "arl-export.bin"
    suffix = Path(safe_name).suffix or ".bin"
    file_name = "{}_{}{}".format(str(job_id), Path(safe_name).stem[:48] or "report", suffix)
    file_path = export_dir / file_name
    tmp_path = export_dir / "{}.part".format(file_name)

    if hasattr(file_content, "save") and callable(getattr(file_content, "save", None)):
        file_content.save(str(tmp_path))
        tmp_path.replace(file_path)
        return file_path, int(file_path.stat().st_size or 0)

    if isinstance(file_content, str):
        file_bytes = file_content.encode("utf-8")
    else:
        file_bytes = bytes(file_content or b"")

    with tmp_path.open("wb") as file_obj:
        file_obj.write(file_bytes)
    tmp_path.replace(file_path)
    return file_path, len(file_bytes)


def run_export_report_job(job_id: str):
    job_id_text = str(job_id or "").strip()
    if not job_id_text:
        raise ValueError("job_id missing")

    _ensure_export_job_indexes()
    collection = _get_export_job_collection()
    now_text = utils.curr_date()
    collection.update_one(
        {"_id": ObjectId(job_id_text)},
        {"$set": {"status": EXPORT_JOB_STATUS_RUNNING, "started_at": now_text, "updated_at": now_text}},
    )

    job_doc = collection.find_one({"_id": ObjectId(job_id_text)})
    if not job_doc:
        raise ValueError("export job not found")

    task_ids = _normalize_task_id_list(job_doc.get("task_ids", []))
    export_format = normalize_export_format(job_doc.get("format", "excel"))
    file_content, filename, content_type = _build_export_content(task_ids, export_format)
    file_path, file_size = _write_export_job_file(job_id_text, filename, file_content)

    completed_text = utils.curr_date()
    expire_at = datetime.utcnow() + timedelta(days=max(1, int(getattr(Config, "EXPORT_REPORT_KEEP_DAYS", 3) or 3)))
    collection.update_one(
        {"_id": ObjectId(job_id_text)},
        {
            "$set": {
                "status": EXPORT_JOB_STATUS_DONE,
                "filename": filename,
                "content_type": content_type,
                "file_path": str(file_path),
                "file_size": int(file_size or 0),
                "completed_at": completed_text,
                "updated_at": completed_text,
                "expire_at": expire_at,
            }
        },
    )
    return {
        "job_id": job_id_text,
        "filename": filename,
        "content_type": content_type,
        "file_size": int(file_size or 0),
    }


def enqueue_export_report_job(task_ids, export_format="excel"):
    normalized_task_ids = _normalize_task_id_list(task_ids)
    if not normalized_task_ids:
        raise ValueError("task_ids is empty")

    _ensure_export_job_indexes()
    _cleanup_stale_export_files()

    first_task = get_task_data(normalized_task_ids[0])
    if not first_task:
        raise ValueError("task not found")

    now_text = utils.curr_date()
    expire_at = datetime.utcnow() + timedelta(days=max(1, int(getattr(Config, "EXPORT_REPORT_KEEP_DAYS", 3) or 3)))
    doc = {
        "task_ids": normalized_task_ids,
        "format": normalize_export_format(export_format),
        "status": EXPORT_JOB_STATUS_QUEUED,
        "created_at": now_text,
        "updated_at": now_text,
        "expire_at": expire_at,
        "task_name": sanitize_excel_value(first_task.get("name", "未知")).strip(),
        "task_target": sanitize_excel_value(first_task.get("target", "")).strip(),
        "task_count": len(normalized_task_ids),
    }
    insert_ret = _get_export_job_collection().insert_one(doc)
    job_id_text = str(insert_ret.inserted_id)

    from app import celerytask

    celery_id = str(
        celerytask.arl_task_web.delay(
            options={
                "celery_action": CeleryAction.EXPORT_REPORT_TASK,
                "data": {"job_id": job_id_text},
            }
        )
    )
    _get_export_job_collection().update_one(
        {"_id": insert_ret.inserted_id},
        {"$set": {"celery_id": celery_id}},
    )
    return {
        "job_id": job_id_text,
        "status": EXPORT_JOB_STATUS_QUEUED,
        "celery_id": celery_id,
        "task_count": len(normalized_task_ids),
    }


def _resolve_export_config_path() -> Path:
    """
    解析导出模块读取配置的路径，优先使用运行时挂载配置。
    """
    custom_path = str(os.environ.get("ARL_CONFIG_EDIT_PATH", "") or "").strip()
    candidates = [
        Path(custom_path) if custom_path else None,
        Path("/code/app/config.yaml"),
        Path(__file__).resolve().parents[2] / "docker" / "config-docker.yaml",
    ]
    for item in candidates:
        if not item:
            continue
        if item.exists() and item.is_file():
            return item
    return Path(__file__).resolve().parents[2] / "docker" / "config-docker.yaml"


def _load_ai_export_config():
    """
    从配置文件读取 AI 导出相关配置。
    """
    config_path = _resolve_export_config_path()
    if not config_path.exists():
        return {}

    try:
        with config_path.open("r", encoding="utf-8") as file_obj:
            loaded = yaml.safe_load(file_obj) or {}
    except Exception:
        loaded = {}

    if not isinstance(loaded, dict):
        return {}

    ai_conf = loaded.get("AI", {})
    if not isinstance(ai_conf, dict):
        ai_conf = {}
    return ai_conf


def _get_ai_export_settings():
    """
    获取 AI 报告导出配置。
    兼容多模型配置，未配置完整凭据时也允许导出模板报告（不抛错）。
    """
    ai_conf = _load_ai_export_config()
    provider_alias = {
        "tongyi": "qwen",
        "qianwen": "qwen",
        "moonshot": "kimi",
        "openai_compatible": "custom_compatible",
        "compatible": "custom_compatible",
    }
    provider_presets = {
        "qwen": {"base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1", "default_model": "qwen-plus"},
        "kimi": {"base_url": "https://api.moonshot.cn/v1", "default_model": "moonshot-v1-8k"},
        "openai": {"base_url": "https://api.openai.com/v1", "default_model": "gpt-4o-mini"},
        "glm": {"base_url": "https://open.bigmodel.cn/api/paas/v4", "default_model": "glm-4-flash"},
        "deepseek": {"base_url": "https://api.deepseek.com/v1", "default_model": "deepseek-chat"},
        "custom_compatible": {"base_url": "", "default_model": ""},
    }

    def normalize_provider(raw_provider):
        provider = str(raw_provider or "").strip().lower()
        provider = provider_alias.get(provider, provider)
        if provider not in provider_presets:
            return "openai"
        return provider

    def normalize_model_profiles(raw_profiles):
        profiles = []
        seen = set()
        if isinstance(raw_profiles, list):
            for index, item in enumerate(raw_profiles):
                if not isinstance(item, dict):
                    continue
                profile_id = str(item.get("id") or "").strip() or "model_{}".format(index + 1)
                if profile_id in seen:
                    continue
                seen.add(profile_id)
                provider = normalize_provider(item.get("provider"))
                preset = provider_presets.get(provider, {})
                profiles.append(
                    {
                        "id": profile_id,
                        "name": str(item.get("name") or profile_id).strip(),
                        "provider": provider,
                        "base_url": str(item.get("base_url") or "").strip() or str(preset.get("base_url") or ""),
                        "api_key": str(item.get("api_key") or "").strip(),
                        "model": str(item.get("model") or "").strip() or str(preset.get("default_model") or ""),
                    }
                )

        if profiles:
            return profiles

        legacy_provider = normalize_provider(ai_conf.get("PROVIDER"))
        legacy_preset = provider_presets.get(legacy_provider, {})
        return [
            {
                "id": "default_model",
                "name": "默认模型",
                "provider": legacy_provider,
                "base_url": str(ai_conf.get("BASE_URL") or "").strip() or str(legacy_preset.get("base_url") or ""),
                "api_key": str(ai_conf.get("API_KEY") or "").strip(),
                "model": str(ai_conf.get("MODEL") or "").strip() or str(legacy_preset.get("default_model") or ""),
            }
        ]

    model_profiles = normalize_model_profiles(ai_conf.get("MODEL_PROFILES"))
    active_model_profile_id = str(ai_conf.get("ACTIVE_MODEL_PROFILE_ID") or "").strip()
    active_profile = {}
    for profile in model_profiles:
        if str(profile.get("id") or "").strip() == active_model_profile_id:
            active_profile = profile
            break
    if not active_profile and model_profiles:
        active_profile = model_profiles[0]
        active_model_profile_id = str(active_profile.get("id") or "").strip()

    provider = str(active_profile.get("provider") or ai_conf.get("PROVIDER") or "").strip() or "openai"
    model = str(active_profile.get("model") or ai_conf.get("MODEL") or "").strip()
    base_url = str(active_profile.get("base_url") or ai_conf.get("BASE_URL") or "").strip()
    api_key = str(active_profile.get("api_key") or "").strip() or str(os.environ.get("ARL_AI_API_KEY", "") or "").strip()
    active_prompt_id = str(ai_conf.get("ACTIVE_PROMPT_ID") or "").strip()
    configured = bool(api_key and base_url and model)

    missing_fields = []
    if not api_key:
        missing_fields.append("api_key")
    if not base_url:
        missing_fields.append("base_url")
    if not model:
        missing_fields.append("model")

    return {
        "enable": bool(ai_conf.get("ENABLE", True)),
        "configured": configured,
        "missing_fields": missing_fields,
        "provider": provider,
        "model": model,
        "base_url": base_url,
        "active_model_profile_id": active_model_profile_id,
        "active_model_profile_name": str(active_profile.get("name") or "").strip(),
        "active_prompt_id": active_prompt_id,
    }


def _normalize_html_cell_value(value):
    """
    将工作表单元格值转换为适合 HTML 展示的安全文本。
    """
    text = sanitize_excel_value(value).replace("\r\n", "\n").replace("\r", "\n")
    return escape(text)


def _format_html_report_time(value):
    """
    格式化报告时间字段，空值返回占位符。
    """
    text = sanitize_excel_value(value).strip()
    return text or "-"


def _render_html_meta_list(items, empty_text="-"):
    """
    将任务名、目标等多值字段渲染为逐行展示的 HTML 列表。
    """
    values = [
        sanitize_excel_value(item).strip()
        for item in as_list(items)
        if sanitize_excel_value(item).strip()
    ]
    if not values:
        values = [sanitize_excel_value(empty_text).strip() or "-"]

    return "".join(
        '<div class="meta-list-item">{}</div>'.format(escape(value))
        for value in values
    )


def build_html_report_metadata(task_items):
    """
    基于任务数据生成 HTML 报告元信息。
    """
    valid_tasks = []
    for item in as_list(task_items):
        if isinstance(item, dict):
            valid_tasks.append(item)

    start_times = []
    end_times = []
    task_names = []
    targets = []

    for item in valid_tasks:
        start_text = sanitize_excel_value(item.get("start_time", "")).strip()
        if start_text:
            parsed = _parse_datetime_safe(start_text)
            sort_key = parsed.strftime("%Y-%m-%d %H:%M:%S.%f") if parsed else start_text
            start_times.append((sort_key, start_text))

        end_text = sanitize_excel_value(item.get("end_time", "")).strip()
        if end_text:
            parsed = _parse_datetime_safe(end_text)
            sort_key = parsed.strftime("%Y-%m-%d %H:%M:%S.%f") if parsed else end_text
            end_times.append((sort_key, end_text))

        task_name = sanitize_excel_value(item.get("name", "")).strip()
        if task_name and task_name not in task_names:
            task_names.append(task_name)

        target = sanitize_excel_value(item.get("target", "")).strip()
        if target and target not in targets:
            targets.append(target)

    start_value = "-"
    if start_times:
        start_times.sort(key=lambda x: x[0])
        start_value = start_times[0][1]

    end_value = "-"
    if end_times:
        end_times.sort(key=lambda x: x[0])
        end_value = end_times[-1][1]

    return {
        "task_count": len(valid_tasks),
        "task_names": task_names,
        "targets": targets,
        "scan_start_time": start_value,
        "scan_end_time": end_value,
    }


def render_workbook_html(wb, title, metadata=None):
    """
    将导出工作簿渲染为 HTML 报告，保持与 Excel 工作表内容一致。
    """
    report_title = escape(sanitize_excel_value(title) or "ARL任务报告")
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    metadata = metadata if isinstance(metadata, dict) else {}

    meta_cards = [
        ("扫描开始时间", _format_html_report_time(metadata.get("scan_start_time", ""))),
        ("截止时间", _format_html_report_time(metadata.get("scan_end_time", ""))),
        ("任务数", str(int(metadata.get("task_count", 0) or 0) or "-")),
        ("生成时间", generated_at),
    ]

    task_name_html = _render_html_meta_list(metadata.get("task_names", []))
    target_html = _render_html_meta_list(metadata.get("targets", []))

    toc_items = []
    sections = []
    for index, ws in enumerate(wb.worksheets, start=1):
        row_iter = ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True)
        first_row = next(row_iter, None)
        if first_row is None:
            continue

        section_id = "sheet-{}".format(index)
        header = [_normalize_html_cell_value(cell) for cell in first_row]
        header_html = "".join(
            "<th><div class='cell-header'>{}</div></th>".format(cell or "&nbsp;")
            for cell in header
        )

        body_count = 0
        body_rows = []
        for row in row_iter:
            body_count += 1
            body_rows.append(
                "<tr>{}</tr>".format(
                    "".join(
                        "<td><div class='cell-content'>{}</div></td>".format(
                            _normalize_html_cell_value(cell) or "&nbsp;"
                        )
                        for cell in row
                    )
                )
            )

        toc_items.append(
            '<a href="#{section_id}" class="toc-link"><span>{title}</span><span>{count} 行</span></a>'.format(
                section_id=section_id,
                title=escape(sanitize_excel_value(ws.title)),
                count=body_count,
            )
        )

        if body_rows:
            body_html = "".join(body_rows)
        else:
            body_html = "<tr><td colspan='{}' class='empty-cell'>暂无数据</td></tr>".format(max(len(header), 1))

        sections.append(
            """
            <section class="sheet-card" id="{section_id}">
              <div class="sheet-header">
                <h2>{title}</h2>
                <span>{count} 行</span>
              </div>
              <div class="table-wrapper">
                <table>
                  <thead>
                    <tr>{header}</tr>
                  </thead>
                  <tbody>
                    {body}
                  </tbody>
                </table>
              </div>
            </section>
            """.format(
                section_id=section_id,
                title=escape(sanitize_excel_value(ws.title)),
                count=body_count,
                header=header_html,
                body=body_html,
            )
        )

    html = """
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
      <meta charset="utf-8" />
      <meta name="viewport" content="width=device-width, initial-scale=1" />
      <title>{title}</title>
      <style>
        :root {{
          color-scheme: light;
          --bg: #f4f7fb;
          --card: #ffffff;
          --border: #d8e1ee;
          --text: #1f2937;
          --muted: #6b7280;
          --accent: #1958a6;
          --accent-soft: #edf4ff;
        }}
        * {{
          box-sizing: border-box;
        }}
        body {{
          margin: 0;
          font-family: "PingFang SC", "Microsoft YaHei", "Helvetica Neue", Arial, sans-serif;
          background: linear-gradient(180deg, #f8fbff 0%, var(--bg) 100%);
          color: var(--text);
        }}
        .page {{
          max-width: 1500px;
          margin: 0 auto;
          padding: 32px 20px 48px;
        }}
        .layout {{
          display: grid;
          grid-template-columns: 280px minmax(0, 1fr);
          gap: 20px;
          align-items: start;
        }}
        .hero {{
          background: linear-gradient(135deg, #ffffff 0%, #eef5ff 100%);
          border: 1px solid var(--border);
          border-radius: 24px;
          padding: 28px 30px;
          box-shadow: 0 18px 48px rgba(15, 23, 42, 0.08);
          margin-bottom: 24px;
        }}
        .hero h1 {{
          margin: 0 0 10px;
          font-size: 30px;
          line-height: 1.2;
        }}
        .hero p {{
          margin: 0;
          color: var(--muted);
          font-size: 14px;
        }}
        .meta-grid {{
          display: grid;
          grid-template-columns: repeat(4, minmax(0, 1fr));
          gap: 12px;
          margin-top: 18px;
        }}
        .meta-card {{
          border: 1px solid var(--border);
          border-radius: 16px;
          background: rgba(255, 255, 255, 0.86);
          padding: 14px 16px;
        }}
        .meta-label {{
          color: var(--muted);
          font-size: 12px;
          margin-bottom: 8px;
        }}
        .meta-value {{
          font-size: 15px;
          font-weight: 700;
          line-height: 1.5;
          word-break: break-word;
        }}
        .meta-list {{
          display: grid;
          gap: 8px;
        }}
        .meta-list-item {{
          padding: 10px 12px;
          border-radius: 12px;
          background: var(--accent-soft);
          border: 1px solid rgba(25, 88, 166, 0.08);
          font-size: 14px;
          font-weight: 600;
          line-height: 1.65;
          word-break: break-word;
        }}
        .meta-wide-grid {{
          display: grid;
          grid-template-columns: repeat(2, minmax(0, 1fr));
          gap: 12px;
          margin-top: 12px;
        }}
        .sidebar {{
          position: sticky;
          top: 20px;
          display: grid;
          gap: 16px;
        }}
        .nav-card {{
          background: var(--card);
          border: 1px solid var(--border);
          border-radius: 20px;
          box-shadow: 0 14px 34px rgba(15, 23, 42, 0.06);
          overflow: hidden;
        }}
        .nav-card h2 {{
          margin: 0;
          padding: 18px 18px 12px;
          font-size: 18px;
        }}
        .toc-list {{
          display: grid;
          gap: 8px;
          padding: 0 12px 14px;
        }}
        .toc-link {{
          display: flex;
          align-items: center;
          justify-content: space-between;
          gap: 12px;
          padding: 10px 12px;
          border-radius: 14px;
          color: var(--text);
          text-decoration: none;
          background: var(--accent-soft);
          border: 1px solid transparent;
          font-size: 13px;
          line-height: 1.5;
        }}
        .toc-link:hover {{
          border-color: var(--border);
          background: #e5efff;
        }}
        .content {{
          min-width: 0;
        }}
        .sheet-list {{
          display: grid;
          gap: 20px;
        }}
        .sheet-card {{
          background: var(--card);
          border: 1px solid var(--border);
          border-radius: 20px;
          box-shadow: 0 14px 34px rgba(15, 23, 42, 0.06);
          overflow: hidden;
        }}
        .sheet-header {{
          display: flex;
          align-items: center;
          justify-content: space-between;
          gap: 12px;
          padding: 18px 20px;
          border-bottom: 1px solid var(--border);
          background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
        }}
        .sheet-header h2 {{
          margin: 0;
          font-size: 18px;
        }}
        .sheet-header span {{
          color: var(--muted);
          font-size: 13px;
        }}
        .table-wrapper {{
          overflow-x: auto;
          padding: 0 0 6px;
        }}
        table {{
          width: max-content;
          min-width: 100%;
          border-collapse: collapse;
          table-layout: auto;
        }}
        th, td {{
          padding: 14px 16px;
          border-bottom: 1px solid var(--border);
          border-right: 1px solid var(--border);
          text-align: left;
          vertical-align: top;
          font-size: 13px;
          line-height: 1.7;
        }}
        th:last-child, td:last-child {{
          border-right: 0;
        }}
        th {{
          min-width: 120px;
        }}
        thead th {{
          position: sticky;
          top: 0;
          background: var(--accent);
          color: #ffffff;
          font-weight: 700;
          z-index: 1;
        }}
        .cell-header {{
          min-width: 120px;
          white-space: normal;
          word-break: break-word;
          overflow-wrap: anywhere;
          line-height: 1.5;
        }}
        .cell-content {{
          min-width: 120px;
          max-width: 420px;
          white-space: pre-wrap;
          word-break: break-word;
          overflow-wrap: anywhere;
          line-height: 1.75;
        }}
        tbody tr:nth-child(even) {{
          background: var(--accent-soft);
        }}
        .empty-cell {{
          text-align: center;
          color: var(--muted);
        }}
        @media (max-width: 768px) {{
          .page {{
            padding: 20px 12px 36px;
          }}
          .layout {{
            grid-template-columns: 1fr;
          }}
          .sidebar {{
            position: static;
          }}
          .hero {{
            padding: 22px 18px;
          }}
          .hero h1 {{
            font-size: 24px;
          }}
          .meta-grid {{
            grid-template-columns: repeat(2, minmax(0, 1fr));
          }}
          .meta-wide-grid {{
            grid-template-columns: 1fr;
          }}
          th, td {{
            padding: 10px 12px;
            font-size: 12px;
          }}
          .cell-content {{
            max-width: 320px;
            min-width: 96px;
          }}
        }}
      </style>
    </head>
    <body>
      <div class="page">
        <section class="hero">
          <h1>{title}</h1>
          <p>报告内容与表格导出保持一致，支持按目录快速跳转查看。</p>
          <div class="meta-grid">
            {meta_cards}
          </div>
          <div class="meta-wide-grid">
            <div class="meta-card">
              <div class="meta-label">任务名</div>
              <div class="meta-list">{task_names}</div>
            </div>
            <div class="meta-card">
              <div class="meta-label">目标</div>
              <div class="meta-list">{targets}</div>
            </div>
          </div>
        </section>
        <div class="layout">
          <aside class="sidebar">
            <section class="nav-card">
              <h2>目录</h2>
              <div class="toc-list">
                {toc_items}
              </div>
            </section>
          </aside>
          <div class="content">
            <div class="sheet-list">
              {sections}
            </div>
          </div>
        </div>
      </div>
    </body>
    </html>
    """.format(
        title=report_title,
        meta_cards="".join(
            """
            <div class="meta-card">
              <div class="meta-label">{label}</div>
              <div class="meta-value">{value}</div>
            </div>
            """.format(label=escape(label), value=escape(value))
            for label, value in meta_cards
        ),
        task_names=task_name_html,
        targets=target_html,
        toc_items="".join(toc_items) or '<div class="meta-card"><div class="meta-value">暂无目录</div></div>',
        sections="".join(sections),
    )
    return html.encode("utf-8")


def sanitize_excel_value(value):
    """
    清洗Excel单元格值，避免非法字符导致导出失败

    说明：
    - 处理 None/bytes/复杂对象类型，统一转换为字符串
    - 过滤 openpyxl 不支持的控制字符
    - 截断超长内容（Excel单元格上限 32767）
    """
    if value is None:
        return ""

    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="ignore")

    if not isinstance(value, str):
        value = str(value)

    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    return value[:32767]


def extract_finger_names(finger_data):
    """
    提取指纹名称列表，兼容 dict/list/str/None 等多种数据格式
    """
    if not finger_data:
        return ""

    if not isinstance(finger_data, list):
        return sanitize_excel_value(finger_data)

    names = []
    for item in finger_data:
        if isinstance(item, dict):
            names.append(sanitize_excel_value(item.get("name", "")))
        else:
            names.append(sanitize_excel_value(item))
    return ",".join([name for name in names if name])


def _beautify_sheet(ws, center_cols=None):
    """
    统一增强工作表可读性：
    - 冻结首行
    - 表头高亮
    - 自动筛选
    - 斑马纹 + 边框 + 自动换行
    """
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row <= 0 or max_col <= 0:
        return

    center_cols = set(center_cols or [])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(max_col), max_row)

    header_fill = PatternFill(fill_type="solid", fgColor="2F75B5")
    zebra_fill = PatternFill(fill_type="solid", fgColor="F6FAFF")
    thin_side = Side(style="thin", color="D9E2F3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_font = Font(name="Consolas", color="FFFFFF", bold=True)
    body_font = Font(name="Consolas", color="111111")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.row_dimensions[1].height = 24
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in range(2, max_row + 1):
        use_zebra = row % 2 == 0
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            if use_zebra:
                cell.fill = zebra_fill
            cell.border = thin_border
            if col in center_cols:
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment


def set_sheet_style(ws):
    """
    通用工作表样式（默认左对齐）
    """
    _beautify_sheet(ws)


def beautify_cert_sheet(ws):
    """
    SSL 证书工作表样式（保留部分字段居中）
    """
    _beautify_sheet(ws, center_cols={5, 6, 7, 8, 9, 10})


def as_list(value):
    """
    将值标准化为列表，兼容 None/单值/列表
    """
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _extract_domain_source_list(value):
    """
    将 domain.source 统一转为列表，兼容 str/list/None。
    """
    items = []
    for raw in as_list(value):
        text = sanitize_excel_value(raw).strip()
        if not text:
            continue
        if "," in text:
            parts = [x.strip() for x in text.split(",") if x.strip()]
            items.extend(parts)
            continue
        items.append(text)

    dedup = []
    seen = set()
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        dedup.append(item)
    return dedup


def _format_domain_source_text(value):
    """
    域名来源展示文本（多来源按换行显示）。
    """
    source_list = _extract_domain_source_list(value)
    if not source_list:
        return "-"
    return " \r\n".join(source_list)


def _is_ip_address(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    try:
        ipaddress.ip_address(text)
        return True
    except Exception:
        return False


def _parse_waf_host_port(host: str, last_url: str):
    """
    解析 WAF 记录中的 IP/域名/端口，兼容 host 与 last_url 两种来源。
    """
    host_text = str(host or "").strip().lower()
    last_url_text = str(last_url or "").strip()
    parsed = urlparse(last_url_text)
    if not getattr(parsed, "hostname", None) and last_url_text and "://" not in last_url_text:
        parsed = urlparse("//{}".format(last_url_text))

    hostname = str(parsed.hostname or "").strip().lower() or host_text
    try:
        parsed_port = parsed.port
    except Exception:
        parsed_port = 0

    try:
        port = int(parsed_port or 0)
    except Exception:
        port = 0

    if port <= 0:
        if parsed.scheme == "https":
            port = 443
        elif parsed.scheme == "http":
            port = 80

    ip = ""
    domain = ""
    if hostname:
        if _is_ip_address(hostname):
            ip = hostname
        elif utils.is_valid_domain(hostname):
            domain = hostname
        else:
            domain = hostname

    return ip, domain, port


def calc_port_service_product_statist_from_ip_items(ip_items):
    """
    基于合并后的IP数据计算资产统计（与单任务统计口径保持一致）
    """
    total = 0
    port_info_list = []
    for item in ip_items:
        port_info = item.get("port_info", [])
        if not port_info:
            continue
        port_info_list.extend(port_info)
        total += len(port_info)

    counter = Counter([info.get("port_id") for info in port_info_list if info.get("port_id") is not None])
    top_20 = counter.most_common(20)
    port_percent_list = []
    for port_id, amount in top_20:
        percent = "{:.2f}%".format((amount * 100.0) / total) if total else "0.00%"
        port_percent_list.append({
            "port_id": port_id,
            "amount": amount,
            "percent": percent
        })

    service_name_list = []
    for info in port_info_list:
        if not info.get("product"):
            continue
        if info.get("product") or info.get("version"):
            service_name = info.get("service_name", "")
            if service_name == "https-alt":
                service_name = "https"
            service_name_list.append(service_name)

    service_top_20 = Counter(service_name_list).most_common(20)
    service_percent_list = []
    for service_name, amount in service_top_20:
        percent = "{:.2f}%".format((amount * 100.0) / len(service_name_list)) if service_name_list else "0.00%"
        service_percent_list.append({
            "service_name": service_name,
            "amount": amount,
            "percent": percent
        })

    product_name_list = []
    for info in port_info_list:
        product = info.get("product")
        if not product:
            continue
        product = sanitize_excel_value(product).strip()
        if product and "**" not in product:
            product_name_list.append(product)

    product_top_20 = Counter(product_name_list).most_common(20)
    product_percent_list = []
    for product, amount in product_top_20:
        percent = "{:.2f}%".format((amount * 100.0) / len(product_name_list)) if product_name_list else "0.00%"
        product_percent_list.append({
            "product": product,
            "amount": amount,
            "percent": percent
        })

    return {
        "port_total": total,
        "port_percent_list": port_percent_list,
        "service_total": len(service_name_list),
        "service_percent_list": service_percent_list,
        "product_total": len(product_name_list),
        "product_percent_list": product_percent_list
    }


@ns.route('/<string:task_id>')
class ARLExport(Resource):
    """任务报告导出接口"""
    
    @auth
    def get(self, task_id):
        """
        导出任务扫描报告为Excel文件
        
        参数：
            task_id: 任务ID
        
        返回：
            Excel文件下载
        
        说明：
        - 生成包含完整扫描结果的Excel报告
        - 文件名：ARL资产导出报告_目标.xlsx
        - 包含多个工作表：
          * 任务概览
          * IP列表及端口服务
          * 域名列表及DNS记录
          * 站点列表及指纹
          * 统计分析（端口Top20、服务Top20等）
        - 适合报告归档和资产分析
        """
        task_data = get_task_data(task_id)
        if not task_data:
            return "not found"

        export_format = normalize_export_format(request.args.get("format", "excel"))
        # 生成文件名（截取目标前20个字符）
        domain = task_data["target"].replace("/", "_")[:20]
        if export_format == "html":
            filename = "ARL资产导出报告_{}.html".format(domain)
            html_data = export_arl_html(task_id)
            return build_export_response(html_data, filename, "text/html; charset=utf-8")
        if export_format == "ai_markdown":
            try:
                markdown_data = export_arl_ai_markdown(task_id)
            except ValueError as exc:
                return {"error": str(exc)}, 400
            filename = "ARL_AI分析报告_{}.md".format(domain)
            return build_export_response(markdown_data, filename, "text/markdown; charset=utf-8")

        filename = "ARL资产导出报告_{}.xlsx".format(domain)
        excel_data = export_arl(task_id)
        return build_export_response(excel_data, filename, "application/octet-stream")



@ns.route('/batch')
class ARLBatchExcel(Resource):
    """批量合并导出接口 - 支持POST请求接收多个任务ID"""
    
    @auth
    def post(self):
        """
        批量导出多个任务并合并成一个Excel文件
        
        请求体：
            {
                "task_ids": ["任务ID1", "任务ID2", ...]
            }
        
        返回：
            合并后的Excel文件下载
        
        说明：
        - 接收多个任务ID列表
        - 合并所有任务的扫描数据（IP、域名、站点等）
        - 自动去重
        - 生成统一的整合Excel报告
        - 文件名：ARL批量导出报告_任务名.xlsx
        """
        try:
            data = request.get_json(silent=True)
            if not data:
                return {"error": "请求体为空"}, 400
                
            task_ids = data.get("task_ids", [])
            export_format = normalize_export_format(data.get("format", "excel"))
            
            if not task_ids or not isinstance(task_ids, list):
                return {"error": "task_ids 必须是非空的列表"}, 400
            
            # 获取任务名（从第一个任务）
            first_task = get_task_data(task_ids[0])
            if not first_task:
                return {"error": "任务不存在"}, 404
            
            task_name = first_task.get("name", "未知")
            if export_format == "html":
                filename = "ARL批量导出报告_{}.html".format(task_name[:20])
                html_data = export_merge_tasks_html(task_ids)
                return build_export_response(html_data, filename, "text/html; charset=utf-8")
            if export_format == "ai_markdown":
                try:
                    markdown_data = export_merge_tasks_ai_markdown(task_ids)
                except ValueError as exc:
                    return {"error": str(exc)}, 400
                filename = "ARL_AI分析报告_{}.md".format(task_name[:20])
                return build_export_response(markdown_data, filename, "text/markdown; charset=utf-8")

            filename = "ARL批量导出报告_{}.xlsx".format(task_name[:20])
            excel_data = export_merge_tasks(task_ids)
            return build_export_response(excel_data, filename, "application/octet-stream")
        except Exception as e:
            logger.exception("批量导出失败: {}".format(str(e)))
            return {"error": "导出失败: {}".format(str(e))}, 500


@ns.route('/job')
class ARLExportJobCreate(Resource):
    """异步报告导出任务创建接口"""

    @auth
    def post(self):
        try:
            data = request.get_json(silent=True) or {}
            task_ids = data.get("task_ids")
            task_id = data.get("task_id")
            export_format = normalize_export_format(data.get("format", "excel"))
            if not task_ids and task_id:
                task_ids = [task_id]
            if not isinstance(task_ids, list) or not task_ids:
                return {"error": "task_ids 必须是非空列表"}, 400

            job_info = enqueue_export_report_job(task_ids, export_format=export_format)
            return {"code": 200, "data": job_info, "message": "export job queued"}
        except ValueError as exc:
            return {"error": str(exc)}, 400
        except Exception as exc:
            logger.exception("create export job failed: %s", exc)
            return {"error": "创建导出任务失败: {}".format(str(exc))}, 500


@ns.route('/job/<string:job_id>')
class ARLExportJobStatus(Resource):
    """异步报告导出任务状态接口"""

    @auth
    def get(self, job_id):
        normalized_job_id = str(job_id or "").strip()
        if not normalized_job_id:
            return {"error": "job_id 不能为空"}, 400
        try:
            job_doc = _get_export_job_collection().find_one({"_id": ObjectId(normalized_job_id)})
        except Exception:
            return {"error": "无效的 job_id"}, 400
        if not job_doc:
            return {"error": "导出任务不存在"}, 404

        data = {
            "job_id": normalized_job_id,
            "status": str(job_doc.get("status", "") or "").strip(),
            "format": str(job_doc.get("format", "") or "").strip(),
            "filename": str(job_doc.get("filename", "") or "").strip(),
            "file_size": int(job_doc.get("file_size", 0) or 0),
            "error": str(job_doc.get("error", "") or "").strip(),
            "created_at": str(job_doc.get("created_at", "") or "").strip(),
            "updated_at": str(job_doc.get("updated_at", "") or "").strip(),
            "started_at": str(job_doc.get("started_at", "") or "").strip(),
            "completed_at": str(job_doc.get("completed_at", "") or "").strip(),
            "task_count": int(job_doc.get("task_count", 0) or 0),
        }
        return {"code": 200, "data": data}


@ns.route('/job/<string:job_id>/download')
class ARLExportJobDownload(Resource):
    """异步报告导出任务下载接口"""

    @auth
    def get(self, job_id):
        normalized_job_id = str(job_id or "").strip()
        if not normalized_job_id:
            return {"error": "job_id 不能为空"}, 400
        try:
            job_doc = _get_export_job_collection().find_one({"_id": ObjectId(normalized_job_id)})
        except Exception:
            return {"error": "无效的 job_id"}, 400
        if not job_doc:
            return {"error": "导出任务不存在"}, 404
        if str(job_doc.get("status", "") or "").strip() != EXPORT_JOB_STATUS_DONE:
            return {"error": "导出任务未完成"}, 409

        file_path = str(job_doc.get("file_path", "") or "").strip()
        filename = str(job_doc.get("filename", "") or "").strip()
        content_type = str(job_doc.get("content_type", "") or "").strip() or "application/octet-stream"
        if not file_path or not os.path.isfile(file_path):
            return {"error": "导出文件不存在或已过期"}, 410

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename or os.path.basename(file_path),
            mimetype=content_type,
        )






def get_task_data(task_id):
    """
    获取任务数据
    
    参数：
        task_id: 任务ID
    
    返回：
        任务数据字典或None
    """
    try:
        task_data = utils.conn_db('task').find_one(
            {'_id': ObjectId(task_id)},
            projection=TASK_EXPORT_PROJECTION,
        )
        return task_data
    except Exception as e:
        pass


def get_ip_data(task_id):
    """
    获取任务的IP数据
    
    参数：
        task_id: 任务ID
    
    返回：
        IP数据游标
    """
    data = utils.conn_db('ip').find(
        {'task_id': task_id},
        projection=IP_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)
    return data


def get_site_data(task_id):
    """
    获取任务的站点数据
    
    参数：
        task_id: 任务ID
    
    返回：
        站点数据游标
    """
    data = utils.conn_db('site').find(
        {'task_id': task_id},
        projection=SITE_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)
    return data


def get_domain_data(task_id):
    """
    获取任务的域名数据
    
    参数：
        task_id: 任务ID
    
    返回：
        域名数据游标
    """
    data = utils.conn_db('domain').find(
        {'task_id': task_id},
        projection=DOMAIN_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)
    return data


def get_url_data(task_id):
    """
    获取任务的 URL 信息数据。
    """
    return utils.conn_db('url').find(
        {'task_id': task_id},
        projection=URL_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_fileleak_data(task_id):
    """
    获取任务的目录扫描（文件泄露）数据。
    """
    return utils.conn_db('fileleak').find(
        {'task_id': task_id},
        projection=FILELEAK_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_wih_data(task_id):
    """
    获取任务的 WIH 数据。
    """
    return utils.conn_db('wih').find(
        {'task_id': task_id},
        projection=WIH_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_wih_endpoint_data(task_id):
    """
    获取任务的 WIH 接口提取数据。
    """
    return utils.conn_db('wih_endpoint').find(
        {'task_id': task_id},
        projection=WIH_ENDPOINT_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def _normalize_task_id_list(task_ids):
    """
    规范化任务ID列表，兼容 str/list/tuple/set 输入。
    """
    if isinstance(task_ids, str):
        raw_items = [task_ids]
    elif isinstance(task_ids, (list, tuple, set)):
        raw_items = list(task_ids)
    else:
        raw_items = []

    result = []
    for item in raw_items:
        task_id = sanitize_excel_value(item).strip()
        if task_id and task_id not in result:
            result.append(task_id)
    return result


def get_service_data(task_ids):
    """
    获取任务的系统服务数据（service 集合）。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    if not task_id_list:
        return []

    if len(task_id_list) == 1:
        query = {"task_id": task_id_list[0]}
    else:
        query = {"task_id": {"$in": task_id_list}}
    return utils.conn_db('service').find(
        query,
        projection=SERVICE_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def _build_service_rows(task_ids, fallback_ip_items=None):
    """
    生成系统服务导出行，优先与页面一致使用 service 集合；
    若 service 集合为空，则回退到 ip.port_info。
    """
    rows = []
    task_id_list = _normalize_task_id_list(task_ids)
    expected_task_ids = set(task_id_list)
    service_hit_task_ids = set()

    for service_item in get_service_data(task_id_list):
        if not isinstance(service_item, dict):
            continue
        service_task_id = sanitize_excel_value(service_item.get("task_id", "")).strip()
        if service_task_id:
            service_hit_task_ids.add(service_task_id)
        service_name = service_item.get("service_name", "")
        for service_info in as_list(service_item.get("service_info", [])):
            if not isinstance(service_info, dict):
                continue
            rows.append([
                service_info.get("ip", ""),
                service_info.get("port_id", ""),
                service_name or service_info.get("service_name", ""),
                service_info.get("product", ""),
                service_info.get("version", ""),
            ])

    missing_task_ids = set()
    if expected_task_ids:
        missing_task_ids = expected_task_ids - service_hit_task_ids
        if rows and not missing_task_ids:
            return rows
    elif rows:
        return rows

    for item in fallback_ip_items or []:
        if not isinstance(item, dict):
            continue

        item_task_id = sanitize_excel_value(item.get("task_id", "")).strip()
        if expected_task_ids:
            if item_task_id:
                if item_task_id not in missing_task_ids:
                    continue
            elif rows and missing_task_ids:
                # 无 task_id 的回退数据无法确定归属，避免与已命中的服务数据重复
                continue

        ip = item.get("ip", "")
        for port_info in as_list(item.get("port_info", [])):
            if not isinstance(port_info, dict):
                continue
            rows.append([
                ip,
                port_info.get("port_id", ""),
                port_info.get("service_name", ""),
                port_info.get("product", ""),
                port_info.get("version", ""),
            ])

    return rows


def get_vuln_data(task_id):
    """
    获取任务的漏洞数据（nPoc/风险巡航等）
    """
    return utils.conn_db('vuln').find(
        {'task_id': task_id},
        projection=VULN_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_nuclei_result_data(task_id):
    """
    获取任务的 nuclei 漏洞结果
    """
    return utils.conn_db('nuclei_result').find(
        {'task_id': task_id},
        projection=NUCLEI_RESULT_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_stat_finger_data(task_id):
    """
    获取任务的指纹统计结果。
    """
    return utils.conn_db('stat_finger').find(
        {'task_id': task_id},
        projection=STAT_FINGER_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_cert_data(task_id):
    """
    获取任务的 SSL 证书结果
    """
    return utils.conn_db('cert').find(
        {'task_id': task_id},
        projection=CERT_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_ai_pen_test_data(task_id):
    """
    获取任务的 AI 渗透测试结果。
    """
    return utils.conn_db('ai_pen_test_result').find(
        {'task_id': task_id},
        projection=AI_PEN_TEST_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def get_ai_denoise_result_data(task_ids):
    """
    获取任务范围内 AI 去噪落库结果。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    if not task_id_list:
        return []
    if len(task_id_list) == 1:
        query = {"task_id": task_id_list[0]}
    else:
        query = {"task_id": {"$in": task_id_list}}
    return utils.conn_db("ai_denoise_result").find(
        query,
        projection=AI_DENOISE_RESULT_EXPORT_PROJECTION,
    ).batch_size(MONGO_EXPORT_BATCH_SIZE)


def _normalize_ai_denoise_module_id(value):
    module_id = sanitize_excel_value(value).strip()
    if module_id in AI_DENOISE_MODULE_LABEL_MAP:
        return module_id
    return module_id or "unknown"


def _normalize_ai_denoise_result_level(value):
    level = sanitize_excel_value(value).strip().lower()
    if level in ("danger", "critical", "high", "严重"):
        return "danger"
    if level in ("suspicious", "warn", "warning", "medium", "中"):
        return "suspicious"
    if level in ("safe", "normal", "ok", "low", "info", "正常", "安全"):
        return "safe"
    if level in ("disabled", "none", "未分析"):
        return "disabled"
    return "disabled"


def _normalize_ai_denoise_source(value):
    source = sanitize_excel_value(value).strip().lower()
    if source in ("ai", "rule", "disabled"):
        return source
    return "disabled"


def _normalize_ai_denoise_risk_level(value):
    risk_text = sanitize_excel_value(value).strip()
    if not risk_text or risk_text == "-":
        return "-"

    level = risk_text.lower()
    if level in ("critical", "severe", "严重"):
        return "严重"
    if level in ("high", "高"):
        return "高"
    if level in ("medium", "moderate", "中"):
        return "中"
    if level in ("low", "info", "informational", "低", "信息"):
        return "低"
    return risk_text


def _ai_denoise_risk_rank(value):
    risk_level = _normalize_ai_denoise_risk_level(value)
    return {
        "严重": 4,
        "高": 3,
        "中": 2,
        "低": 1,
    }.get(risk_level, 0)


def _normalize_ai_denoise_trust(value):
    trust_text = sanitize_excel_value(value).strip()
    if not trust_text or trust_text == "-":
        return "-"

    low = trust_text.lower()
    if "误报" in trust_text or low in ("false_positive", "fp", "suspected_false_positive"):
        return "疑似误报"
    if "可信" in trust_text or low in ("trusted", "可信", "reliable"):
        return "可信"
    return trust_text


def _truncate_report_text(value, max_length=160):
    text = sanitize_excel_value(value).replace("\r", " ").replace("\n", " ").strip()
    if not text:
        return ""
    if len(text) > max_length:
        return "{}...".format(text[:max_length].rstrip())
    return text


def _normalize_text_list(value, max_items=3, max_item_len=180):
    if not isinstance(value, list):
        return []
    items = []
    for item in value:
        text = _truncate_report_text(item, max_item_len)
        if text:
            items.append(text)
        if len(items) >= max_items:
            break
    return items


def _resolve_ai_denoise_target_text(module_id, item):
    if not isinstance(item, dict):
        return ""

    if module_id == "site":
        return _truncate_report_text(item.get("site") or item.get("url") or item.get("title"), 180)
    if module_id == "fileleak":
        return _truncate_report_text(item.get("url") or item.get("site") or item.get("title"), 180)
    if module_id == "cert":
        host = sanitize_excel_value(item.get("host", "")).strip()
        if host:
            return _truncate_report_text(host, 180)
        domain = sanitize_excel_value(item.get("domain", "")).strip()
        if domain:
            return _truncate_report_text(domain, 180)
        ip = sanitize_excel_value(item.get("ip", "")).strip()
        port = sanitize_excel_value(item.get("port", "")).strip()
        if ip and port:
            return "{}:{}".format(ip, port)
        return _truncate_report_text(ip or port, 180)
    if module_id == "url":
        return _truncate_report_text(item.get("url") or item.get("site") or item.get("title"), 180)
    if module_id == "vuln":
        return _truncate_report_text(item.get("target") or item.get("vul_name"), 180)
    if module_id == "nuclei_result":
        return _truncate_report_text(item.get("vuln_url") or item.get("target") or item.get("vuln_name"), 180)

    return ""


def _query_docs_by_ids(collection_name, id_values, projection):
    if not collection_name or not isinstance(projection, dict):
        return {}

    text_ids = set()
    object_ids = []
    seen_object_ids = set()
    for raw_id in id_values or []:
        data_id = sanitize_excel_value(raw_id).strip()
        if not data_id:
            continue
        text_ids.add(data_id)
        if ObjectId.is_valid(data_id) and data_id not in seen_object_ids:
            seen_object_ids.add(data_id)
            object_ids.append(ObjectId(data_id))

    query_parts = []
    if object_ids:
        query_parts.append({"_id": {"$in": object_ids}})
    if text_ids:
        query_parts.append({"_id": {"$in": list(text_ids)}})
    if not query_parts:
        return {}

    query = query_parts[0] if len(query_parts) == 1 else {"$or": query_parts}
    doc_map = {}
    for item in utils.conn_db(collection_name).find(query, projection=projection).batch_size(MONGO_EXPORT_BATCH_SIZE):
        doc_id = sanitize_excel_value(item.get("_id", "")).strip()
        if doc_id:
            doc_map[doc_id] = item
    return doc_map


def _build_ai_denoise_target_map(ai_rows):
    module_id_map = {}
    for row in ai_rows:
        if not isinstance(row, dict):
            continue
        module_id = _normalize_ai_denoise_module_id(row.get("module_id"))
        if module_id not in AI_DENOISE_MODULE_COLLECTION_MAP:
            continue

        candidates = []
        for raw_id in (row.get("data_id"), row.get("row_key")):
            data_id = sanitize_excel_value(raw_id).strip()
            if data_id:
                candidates.append(data_id)
        if not candidates:
            continue

        id_set = module_id_map.setdefault(module_id, set())
        for data_id in candidates:
            id_set.add(data_id)

    target_map = {}
    for module_id, id_set in module_id_map.items():
        collection_name = AI_DENOISE_MODULE_COLLECTION_MAP.get(module_id, "")
        projection = AI_DENOISE_MODULE_TARGET_PROJECTION.get(module_id, {"_id": 1})
        doc_map = _query_docs_by_ids(collection_name, id_set, projection)
        for doc_id, item in doc_map.items():
            target_text = _resolve_ai_denoise_target_text(module_id, item)
            if target_text:
                target_map[(module_id, doc_id)] = target_text
    return target_map


def _build_default_ai_lookup_result():
    return {
        "text": "未分析",
        "result_level": "disabled",
        "source": "disabled",
        "level_rank": AI_DENOISE_RESULT_LEVEL_ORDER.get("disabled", 0),
        "source_rank": AI_DENOISE_SOURCE_ORDER.get("disabled", 0),
    }


def _normalize_ai_lookup_key(value):
    if isinstance(value, ObjectId):
        return str(value)
    return sanitize_excel_value(value).strip()


def _build_ai_display_text_from_doc(result_doc):
    display_text = _truncate_report_text(result_doc.get("display_text"), 80)
    if display_text:
        return display_text

    result_level = _normalize_ai_denoise_result_level(result_doc.get("result_level"))
    risk_level = _normalize_ai_denoise_risk_level(result_doc.get("risk_level"))
    trust = _normalize_ai_denoise_trust(result_doc.get("trust"))
    result_label = AI_DENOISE_RESULT_LEVEL_LABEL.get(result_level, "未分析")

    if risk_level != "-" and trust != "-":
        return "{}（{}/{}）".format(result_label, risk_level, trust)
    if risk_level != "-":
        return "{}（{}）".format(result_label, risk_level)
    return result_label


def _build_ai_lookup_result_from_doc(result_doc):
    result_level = _normalize_ai_denoise_result_level(result_doc.get("result_level"))
    source = _normalize_ai_denoise_source(result_doc.get("source"))
    return {
        "text": _build_ai_display_text_from_doc(result_doc),
        "result_level": result_level,
        "source": source,
        "level_rank": AI_DENOISE_RESULT_LEVEL_ORDER.get(result_level, 0),
        "source_rank": AI_DENOISE_SOURCE_ORDER.get(source, 0),
    }


def _is_ai_lookup_result_better(candidate, current):
    if not isinstance(candidate, dict):
        return False
    if not isinstance(current, dict):
        return True

    candidate_tuple = (
        int(candidate.get("level_rank", 0) or 0),
        int(candidate.get("source_rank", 0) or 0),
    )
    current_tuple = (
        int(current.get("level_rank", 0) or 0),
        int(current.get("source_rank", 0) or 0),
    )
    return candidate_tuple > current_tuple


def _build_ai_denoise_lookup(task_ids, module_id):
    """
    构建 ai_denoise_result 的 data_id/row_key 索引，供 Excel 导出补充 AI 分析列。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    module_id_text = _normalize_ai_denoise_module_id(module_id)
    if not task_id_list or not module_id_text:
        return {"by_data_id": {}, "by_row_key": {}}

    if len(task_id_list) == 1:
        query = {"task_id": task_id_list[0], "module_id": module_id_text}
    else:
        query = {"task_id": {"$in": task_id_list}, "module_id": module_id_text}

    by_data_id = {}
    by_row_key = {}
    projection = {
        "data_id": 1,
        "row_key": 1,
        "display_text": 1,
        "result_level": 1,
        "risk_level": 1,
        "trust": 1,
        "source": 1,
    }
    for item in utils.conn_db("ai_denoise_result").find(query, projection=projection).batch_size(MONGO_EXPORT_BATCH_SIZE):
        result = _build_ai_lookup_result_from_doc(item)
        data_id = _normalize_ai_lookup_key(item.get("data_id", ""))
        row_key = _normalize_ai_lookup_key(item.get("row_key", ""))

        if data_id:
            current = by_data_id.get(data_id)
            if _is_ai_lookup_result_better(result, current):
                by_data_id[data_id] = result
        if row_key:
            current = by_row_key.get(row_key)
            if _is_ai_lookup_result_better(result, current):
                by_row_key[row_key] = result

    return {"by_data_id": by_data_id, "by_row_key": by_row_key}


def _resolve_ai_lookup_result(lookup, data_id="", row_key=""):
    by_data_id = (lookup or {}).get("by_data_id", {})
    by_row_key = (lookup or {}).get("by_row_key", {})

    data_id_key = _normalize_ai_lookup_key(data_id)
    row_key_key = _normalize_ai_lookup_key(row_key)
    result = None

    if data_id_key and data_id_key in by_data_id:
        result = by_data_id.get(data_id_key)
    if row_key_key and row_key_key in by_row_key:
        candidate = by_row_key.get(row_key_key)
        if _is_ai_lookup_result_better(candidate, result):
            result = candidate
    if isinstance(result, dict):
        return result
    return _build_default_ai_lookup_result()


def _extract_ai_denoise_rows(task_ids):
    """
    汇总 AI 去噪结果，供 Markdown 报告展示。
    """
    rows = []
    dedup_keys = set()

    for item in get_ai_denoise_result_data(task_ids):
        if not isinstance(item, dict):
            continue

        task_id = sanitize_excel_value(item.get("task_id", "")).strip()
        module_id = _normalize_ai_denoise_module_id(item.get("module_id"))
        row_key = sanitize_excel_value(item.get("row_key", "")).strip()
        data_id = sanitize_excel_value(item.get("data_id", "")).strip()
        dedup_key = (task_id, module_id, row_key or data_id)
        if dedup_key in dedup_keys:
            continue
        dedup_keys.add(dedup_key)

        result_level = _normalize_ai_denoise_result_level(item.get("result_level"))
        risk_level = _normalize_ai_denoise_risk_level(item.get("risk_level"))
        trust = _normalize_ai_denoise_trust(item.get("trust"))
        source = _normalize_ai_denoise_source(item.get("source"))

        display_text = _truncate_report_text(item.get("display_text"), 120)
        if not display_text:
            display_text = AI_DENOISE_RESULT_LEVEL_LABEL.get(result_level, "未分析")

        rows.append(
            {
                "task_id": task_id,
                "module_id": module_id,
                "module_label": AI_DENOISE_MODULE_LABEL_MAP.get(module_id, module_id),
                "row_key": row_key,
                "data_id": data_id,
                "result_level": result_level,
                "result_label": AI_DENOISE_RESULT_LEVEL_LABEL.get(result_level, "未分析"),
                "risk_level": risk_level,
                "trust": trust,
                "source": source,
                "source_label": AI_DENOISE_SOURCE_LABEL.get(source, "未分析"),
                "display_text": display_text,
                "summary": _truncate_report_text(item.get("summary"), 280),
                "analyzed_at": sanitize_excel_value(item.get("analyzed_at", "")).strip()
                or sanitize_excel_value(item.get("updated_at", "")).strip(),
                "evidence": _normalize_text_list(item.get("evidence"), max_items=3, max_item_len=180),
                "suggestions": _normalize_text_list(item.get("suggestions"), max_items=3, max_item_len=180),
            }
        )

    target_map = _build_ai_denoise_target_map(rows)
    for row in rows:
        module_id = row.get("module_id", "")
        target = ""
        for data_id in (row.get("data_id", ""), row.get("row_key", "")):
            key = (module_id, sanitize_excel_value(data_id).strip())
            if key in target_map:
                target = target_map[key]
                break
        if not target:
            target = (
                _truncate_report_text(row.get("summary"), 120)
                or _truncate_report_text(row.get("display_text"), 120)
                or _truncate_report_text(row.get("row_key"), 80)
                or "-"
            )
        row["target"] = target

    return rows


def _build_ai_denoise_overview(ai_rows):
    """
    生成 AI 去噪报告概览统计。
    """
    overview = {
        "total": 0,
        "analyzed": 0,
        "high_value": 0,
        "suspected_fp": 0,
        "source_stat": {"ai": 0, "rule": 0, "disabled": 0},
        "result_stat": {"danger": 0, "suspicious": 0, "safe": 0, "disabled": 0},
        "module_rows": [],
    }
    module_stats = {}

    for row in ai_rows:
        if not isinstance(row, dict):
            continue

        overview["total"] += 1
        source = _normalize_ai_denoise_source(row.get("source"))
        level = _normalize_ai_denoise_result_level(row.get("result_level"))
        trust = _normalize_ai_denoise_trust(row.get("trust"))
        module_id = _normalize_ai_denoise_module_id(row.get("module_id"))
        module_label = AI_DENOISE_MODULE_LABEL_MAP.get(module_id, module_id)

        overview["source_stat"][source] = overview["source_stat"].get(source, 0) + 1
        overview["result_stat"][level] = overview["result_stat"].get(level, 0) + 1
        if source in ("ai", "rule") and level != "disabled":
            overview["analyzed"] += 1
        if level in ("danger", "suspicious"):
            overview["high_value"] += 1
        if trust == "疑似误报":
            overview["suspected_fp"] += 1

        module_item = module_stats.get(module_id)
        if not module_item:
            module_item = {
                "module_id": module_id,
                "module_label": module_label,
                "total": 0,
                "ai": 0,
                "rule": 0,
                "disabled": 0,
                "danger": 0,
                "suspicious": 0,
                "safe": 0,
                "suspected_fp": 0,
            }
            module_stats[module_id] = module_item

        module_item["total"] += 1
        module_item[source] += 1
        if level in ("danger", "suspicious", "safe"):
            module_item[level] += 1
        if trust == "疑似误报":
            module_item["suspected_fp"] += 1

    overview["module_rows"] = sorted(
        list(module_stats.values()),
        key=lambda item: (
            -int(item.get("danger", 0) or 0),
            -int(item.get("suspicious", 0) or 0),
            -int(item.get("total", 0) or 0),
            sanitize_excel_value(item.get("module_label", "")),
        ),
    )
    return overview


def _build_ai_high_value_rows(ai_rows, limit=20):
    """
    取 AI 去噪中危险/可疑的高价值目标。
    """
    rows = []
    for row in ai_rows:
        if not isinstance(row, dict):
            continue
        level = _normalize_ai_denoise_result_level(row.get("result_level"))
        if level not in ("danger", "suspicious"):
            continue
        rows.append(row)

    rows.sort(
        key=lambda item: (
            -AI_DENOISE_RESULT_LEVEL_ORDER.get(_normalize_ai_denoise_result_level(item.get("result_level")), 0),
            -_ai_denoise_risk_rank(item.get("risk_level")),
            -AI_DENOISE_SOURCE_ORDER.get(_normalize_ai_denoise_source(item.get("source")), 0),
            sanitize_excel_value(item.get("module_label", "")),
            sanitize_excel_value(item.get("target", "")),
        )
    )
    return rows[:limit]


def _build_ai_suspected_fp_rows(ai_rows, limit=20):
    """
    取 AI 去噪中“疑似误报”的候选项。
    """
    rows = []
    for row in ai_rows:
        if not isinstance(row, dict):
            continue
        trust = _normalize_ai_denoise_trust(row.get("trust"))
        if trust != "疑似误报":
            continue
        rows.append(row)

    rows.sort(
        key=lambda item: (
            AI_DENOISE_RESULT_LEVEL_ORDER.get(_normalize_ai_denoise_result_level(item.get("result_level")), 0),
            _ai_denoise_risk_rank(item.get("risk_level")),
            sanitize_excel_value(item.get("analyzed_at", "")),
            sanitize_excel_value(item.get("module_label", "")),
        ),
        reverse=True,
    )
    return rows[:limit]


def _extract_url_rows(task_ids):
    """
    汇总 URL 信息导出行，按关键字段去重。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    ai_lookup = _build_ai_denoise_lookup(task_id_list, "url")
    rows = []
    dedup_keys = set()
    for task_id in task_id_list:
        for item in get_url_data(task_id):
            base_row = [
                sanitize_excel_value(item.get("url", "")),
                sanitize_excel_value(item.get("site", "")),
                sanitize_excel_value(item.get("title", "")),
                sanitize_excel_value(item.get("status_code", "")),
                sanitize_excel_value(item.get("content_length", "")),
                sanitize_excel_value(item.get("source", "")),
            ]
            key = tuple(base_row)
            if key in dedup_keys:
                continue
            dedup_keys.add(key)
            item_id = _normalize_ai_lookup_key(item.get("_id", ""))
            ai_result = _resolve_ai_lookup_result(ai_lookup, data_id=item_id, row_key=item_id)
            row = base_row + [sanitize_excel_value(ai_result.get("text", "未分析"))]
            rows.append(row)
    return rows


def _extract_fileleak_rows(task_ids):
    """
    汇总目录扫描（文件泄露）导出行，按 URL 去重。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    ai_lookup = _build_ai_denoise_lookup(task_id_list, "fileleak")
    rows = []
    dedup_urls = set()
    for task_id in task_id_list:
        for item in get_fileleak_data(task_id):
            url = sanitize_excel_value(item.get("url", "")).strip()
            if not url or url in dedup_urls:
                continue
            dedup_urls.add(url)
            item_id = _normalize_ai_lookup_key(item.get("_id", ""))
            ai_result = _resolve_ai_lookup_result(ai_lookup, data_id=item_id, row_key=item_id)
            rows.append(
                [
                    url,
                    sanitize_excel_value(item.get("site", "")),
                    sanitize_excel_value(item.get("title", "")),
                    sanitize_excel_value(item.get("status_code", "")),
                    sanitize_excel_value(item.get("content_length", "")),
                    sanitize_excel_value(_format_fileleak_source(item.get("source"))),
                    sanitize_excel_value(ai_result.get("text", "未分析")),
                ]
            )
    return rows


def _format_fileleak_source(value):
    source = str(value or "").strip().lower()
    if source in {"", "dict_brute", "dictionary_brute", "brute"}:
        return "字典爆破"
    if source == "wih_url_probe":
        return "wih_url_probe"
    return value


def _extract_wih_rows(task_ids):
    """
    汇总 WIH 导出行，按 record_type+content+source+site 去重。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    rows = []
    dedup_keys = set()
    for task_id in task_id_list:
        for item in get_wih_data(task_id):
            row = [
                sanitize_excel_value(item.get("record_type", "")),
                sanitize_excel_value(item.get("content", "")),
                sanitize_excel_value(item.get("source", "")),
                sanitize_excel_value(item.get("site", "")),
            ]
            key = tuple(row)
            if key in dedup_keys:
                continue
            dedup_keys.add(key)
            rows.append(row)
    return rows


def _extract_wih_endpoint_request_packet(item):
    """
    从 WIH 接口记录中提取请求报文，兼容历史 request_template 结构。
    """
    request_packet = sanitize_excel_value(item.get("request_packet", "")).strip()
    if request_packet:
        return request_packet

    request_template = item.get("request_template", {})
    if isinstance(request_template, dict):
        return sanitize_excel_value(request_template.get("request_packet", "")).strip()
    return ""


def _format_wih_endpoint_status(item):
    status_text = sanitize_excel_value(item.get("status_code") or item.get("response_status", "")).strip()
    try:
        status_num = int(float(status_text))
    except Exception:
        return status_text if status_text else "-"
    return str(status_num) if status_num > 0 else "-"


def _format_wih_endpoint_response_size(item):
    size_value = item.get("response_size")
    if size_value is None or size_value == "":
        size_value = item.get("content_length", "")
    size_text = sanitize_excel_value(size_value).strip()
    try:
        size_num = int(float(size_text))
    except Exception:
        return size_text if size_text else "-"

    status_text = _format_wih_endpoint_status(item)
    if size_num <= 0 and status_text == "-":
        return "-"
    return str(size_num)


def _extract_wih_endpoint_rows(task_ids):
    """
    汇总 WIH 接口提取导出行，按目标+页面URL+方法+请求URL+请求报文去重。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    rows = []
    dedup_keys = set()

    for task_id in task_id_list:
        for item in get_wih_endpoint_data(task_id):
            target = sanitize_excel_value(item.get("target") or item.get("site") or "").strip()
            page_url = sanitize_excel_value(item.get("page_url", "")).strip()
            method = sanitize_excel_value(item.get("method", "")).strip().upper() or "GET"
            status_code = _format_wih_endpoint_status(item)
            response_size = _format_wih_endpoint_response_size(item)
            request_url = sanitize_excel_value(item.get("url") or item.get("request_url") or "").strip()
            request_packet = _extract_wih_endpoint_request_packet(item)

            if not request_url and not request_packet:
                continue

            dedup_key = (target, page_url, method, request_url, request_packet)
            if dedup_key in dedup_keys:
                continue
            dedup_keys.add(dedup_key)
            rows.append(
                [
                    len(rows) + 1,
                    target,
                    page_url,
                    method,
                    status_code,
                    response_size,
                    request_url,
                    request_packet,
                ]
            )

    return rows


def _extract_waf_rows(task_ids):
    """
    汇总任务的 WAF 识别结果（来源 task.waf_skip_summary.blocked_hosts）。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    rows = []
    dedup_keys = set()

    for task_id in task_id_list:
        task_data = get_task_data(task_id)
        if not isinstance(task_data, dict):
            continue

        waf_skip_summary = task_data.get("waf_skip_summary", {})
        blocked_hosts = (
            (waf_skip_summary or {}).get("blocked_hosts", [])
            if isinstance(waf_skip_summary, dict)
            else []
        )

        for host_item in blocked_hosts:
            if isinstance(host_item, dict):
                host_data = host_item
            else:
                host_data = {"host": sanitize_excel_value(host_item).strip()}

            host = sanitize_excel_value(host_data.get("host", "")).strip().lower()
            last_url = sanitize_excel_value(host_data.get("last_url", "")).strip()
            waf_name = sanitize_excel_value(host_data.get("waf_name", "")).strip() or "unknown"
            waf_confidence = sanitize_excel_value(host_data.get("waf_confidence", "")).strip()
            module = sanitize_excel_value(host_data.get("module", "")).strip()
            rule = sanitize_excel_value(host_data.get("rule", "")).strip()
            reason = sanitize_excel_value(host_data.get("reason", "")).strip()
            hit_count = sanitize_excel_value(host_data.get("hit_count", ""))
            skip_count = sanitize_excel_value(host_data.get("skip_count", ""))
            last_status = sanitize_excel_value(host_data.get("last_status", ""))
            waf_evidence = " \r\n".join(
                [
                    sanitize_excel_value(item).strip()
                    for item in as_list(host_data.get("waf_evidence", []))
                    if sanitize_excel_value(item).strip()
                ]
            )

            ip, domain, port = _parse_waf_host_port(host, last_url)
            port_text = str(port) if int(port or 0) > 0 else ""
            dedup_key = (task_id, host, port_text, waf_name)
            if dedup_key in dedup_keys:
                continue
            dedup_keys.add(dedup_key)

            rows.append(
                [
                    sanitize_excel_value(ip),
                    sanitize_excel_value(domain),
                    sanitize_excel_value(port_text),
                    sanitize_excel_value(waf_name),
                    sanitize_excel_value(waf_confidence),
                    sanitize_excel_value(module),
                    sanitize_excel_value(rule),
                    sanitize_excel_value(reason),
                    sanitize_excel_value(hit_count),
                    sanitize_excel_value(skip_count),
                    sanitize_excel_value(last_status),
                    sanitize_excel_value(last_url),
                    sanitize_excel_value(waf_evidence),
                ]
            )

    return rows


def _build_url_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增 URL 信息工作表。
    """
    ws = wb.create_sheet(title="URL信息")
    ws.column_dimensions['A'].width = 62.0
    ws.column_dimensions['B'].width = 46.0
    ws.column_dimensions['C'].width = 52.0
    ws.column_dimensions['D'].width = 10.0
    ws.column_dimensions['E'].width = 12.0
    ws.column_dimensions['F'].width = 24.0
    ws.column_dimensions['G'].width = 24.0
    ws.append(["URL", "站点", "标题", "状态码", "body长度", "来源", "AI分析"])

    for row in _extract_url_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _build_waf_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增 WAF 识别工作表。
    """
    ws = wb.create_sheet(title="WAF识别")
    ws.column_dimensions['A'].width = 18.0
    ws.column_dimensions['B'].width = 36.0
    ws.column_dimensions['C'].width = 10.0
    ws.column_dimensions['D'].width = 20.0
    ws.column_dimensions['E'].width = 12.0
    ws.column_dimensions['F'].width = 16.0
    ws.column_dimensions['G'].width = 16.0
    ws.column_dimensions['H'].width = 42.0
    ws.column_dimensions['I'].width = 10.0
    ws.column_dimensions['J'].width = 10.0
    ws.column_dimensions['K'].width = 12.0
    ws.column_dimensions['L'].width = 64.0
    ws.column_dimensions['M'].width = 42.0
    ws.append(
        [
            "IP",
            "域名",
            "端口",
            "WAF厂家",
            "置信度",
            "命中模块",
            "触发规则",
            "触发原因",
            "命中次数",
            "跳过次数",
            "最后状态码",
            "最后URL",
            "命中证据",
        ]
    )

    for row in _extract_waf_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _extract_nuclei_rows(task_ids):
    """
    汇总 PoC 风险导出行（nuclei_result），保留页面级关键字段。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    ai_lookup = _build_ai_denoise_lookup(task_id_list, "nuclei_result")
    rows = []
    dedup_keys = set()

    for task_id in task_id_list:
        for item in get_nuclei_result_data(task_id):
            scanner_type = sanitize_excel_value(item.get("scanner_type", "")).strip() or "nuclei"
            rule_id = sanitize_excel_value(
                item.get("rule_id", "")
                or item.get("template_id", "")
            ).strip()
            target = sanitize_excel_value(item.get("target", "")).strip()
            vuln_url = sanitize_excel_value(item.get("vuln_url", "")).strip()
            vuln_name = sanitize_excel_value(item.get("vuln_name", "")).strip()
            vuln_severity = sanitize_excel_value(item.get("vuln_severity", "")).strip()
            save_date = sanitize_excel_value(item.get("save_date", "")).strip()
            verify_data = sanitize_excel_value(
                item.get("verify_data", "")
                or item.get("template_url", "")
            ).strip()

            dedup_key = (
                task_id,
                scanner_type,
                rule_id,
                target,
                vuln_url,
                vuln_name,
                vuln_severity,
            )
            if dedup_key in dedup_keys:
                continue
            dedup_keys.add(dedup_key)
            item_id = _normalize_ai_lookup_key(item.get("_id", ""))
            ai_result = _resolve_ai_lookup_result(ai_lookup, data_id=item_id, row_key=item_id)

            rows.append([
                scanner_type,
                rule_id,
                target,
                vuln_url,
                vuln_name,
                vuln_severity,
                save_date,
                verify_data,
                sanitize_excel_value(ai_result.get("text", "未分析")),
            ])

    return rows


def _build_nuclei_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增 PoC 风险工作表。
    """
    ws = wb.create_sheet(title="PoC风险")
    ws.column_dimensions['A'].width = 12.0
    ws.column_dimensions['B'].width = 32.0
    ws.column_dimensions['C'].width = 36.0
    ws.column_dimensions['D'].width = 62.0
    ws.column_dimensions['E'].width = 36.0
    ws.column_dimensions['F'].width = 14.0
    ws.column_dimensions['G'].width = 21.0
    ws.column_dimensions['H'].width = 80.0
    ws.column_dimensions['I'].width = 24.0
    ws.append(["扫描器", "规则ID", "目标", "风险URL", "风险名称", "风险等级", "发现时间", "验证信息", "AI分析"])

    for row in _extract_nuclei_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _build_fileleak_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增目录扫描工作表。
    """
    ws = wb.create_sheet(title="目录扫描")
    ws.column_dimensions['A'].width = 62.0
    ws.column_dimensions['B'].width = 46.0
    ws.column_dimensions['C'].width = 52.0
    ws.column_dimensions['D'].width = 10.0
    ws.column_dimensions['E'].width = 12.0
    ws.column_dimensions['F'].width = 18.0
    ws.column_dimensions['G'].width = 24.0
    ws.append(["URL", "站点", "标题", "状态码", "body长度", "来源", "AI分析"])

    for row in _extract_fileleak_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _build_wih_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增 WIH 工作表。
    """
    ws = wb.create_sheet(title="WIH")
    ws.column_dimensions['A'].width = 22.0
    ws.column_dimensions['B'].width = 64.0
    ws.column_dimensions['C'].width = 52.0
    ws.column_dimensions['D'].width = 46.0
    ws.append(["记录类型", "内容", "来源", "站点"])

    for row in _extract_wih_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _build_wih_endpoint_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增 WIH 接口提取工作表。
    """
    ws = wb.create_sheet(title="WIH接口提取")
    ws.column_dimensions['A'].width = 8.0
    ws.column_dimensions['B'].width = 46.0
    ws.column_dimensions['C'].width = 58.0
    ws.column_dimensions['D'].width = 10.0
    ws.column_dimensions['E'].width = 10.0
    ws.column_dimensions['F'].width = 12.0
    ws.column_dimensions['G'].width = 72.0
    ws.column_dimensions['H'].width = 96.0
    ws.append(["序号", "目标", "页面URL", "方法", "状态码", "响应大小", "请求url", "请求报文"])

    for row in _extract_wih_endpoint_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _extract_stat_finger_rows(task_ids):
    """
    汇总指纹统计导出行；批量导出时按指纹名累计数量。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    finger_counter = Counter()

    for task_id in task_id_list:
        for item in get_stat_finger_data(task_id):
            name = sanitize_excel_value(item.get("name", "")).strip()
            if not name:
                continue
            try:
                cnt = int(item.get("cnt", 0) or 0)
            except Exception:
                cnt = 0
            finger_counter[name] += cnt

    rows = []
    for name, cnt in sorted(finger_counter.items(), key=lambda kv: (-int(kv[1]), kv[0])):
        rows.append([sanitize_excel_value(name), sanitize_excel_value(cnt)])
    return rows


def _build_stat_finger_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增指纹统计工作表。
    """
    ws = wb.create_sheet(title="指纹统计")
    ws.column_dimensions['A'].width = 56.0
    ws.column_dimensions['B'].width = 12.0
    ws.append(["finger", "数量"])

    for row in _extract_stat_finger_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _extract_ai_pen_rows(task_ids):
    """
    汇总 AI 渗透测试导出行。
    """
    def _safe_list(value):
        return list(value or []) if isinstance(value, (list, tuple)) else []

    def _parse_json_object(value):
        text = sanitize_excel_value(value).strip()
        if not text:
            return {}
        try:
            parsed = json.loads(text)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}

    def _normalize_text_list(items, max_items=32):
        result = []
        seen = set()
        for item in _safe_list(items):
            text = sanitize_excel_value(item).strip()
            lowered = text.lower()
            if not text or lowered in seen:
                continue
            seen.add(lowered)
            result.append(text)
            if len(result) >= max(1, int(max_items or 1)):
                break
        return result

    def _extract_ai_pen_interface_fetch_counts(item):
        summary = item.get("api_surface_summary") if isinstance(item.get("api_surface_summary"), dict) else {}
        sample_interfaces = [entry for entry in _safe_list(summary.get("sample_interfaces")) if isinstance(entry, dict)]
        runtime_calls = [entry for entry in _safe_list(item.get("runtime_api_calls")) if isinstance(entry, dict)]
        seen = set()
        get_count = 0
        post_count = 0

        source_items = sample_interfaces if sample_interfaces else runtime_calls
        for entry in source_items:
            method_text = sanitize_excel_value(entry.get("method", "GET")).strip().upper() or "GET"
            target_text = sanitize_excel_value(
                entry.get("url")
                or entry.get("url_template")
                or entry.get("path_template")
                or entry.get("path")
            ).strip()
            if not target_text:
                continue
            cache_key = "{}|{}".format(method_text, target_text)
            if cache_key in seen:
                continue
            seen.add(cache_key)
            if method_text == "GET":
                get_count += 1
            elif method_text in {"POST", "PUT", "PATCH"}:
                post_count += 1
        return {
            "get_count": get_count,
            "post_count": post_count,
            "summary": "POST：{}条\nGET：{}条".format(post_count, get_count),
        }

    def _build_ai_pen_curl_payload(item):
        method_text = sanitize_excel_value(item.get("request_method", "")).strip().upper() or "GET"
        request_url = sanitize_excel_value(item.get("request_url") or item.get("vuln_url") or item.get("target") or "").strip()
        if not request_url:
            return ""

        parts = ["curl", "-X", method_text]
        request_headers = item.get("request_headers") if isinstance(item.get("request_headers"), dict) else {}
        for header_key, header_value in request_headers.items():
            key_text = sanitize_excel_value(header_key).strip()
            value_text = sanitize_excel_value(header_value).strip()
            lowered = key_text.lower()
            if not key_text or not value_text:
                continue
            if lowered in {"host", "content-length"}:
                continue
            parts.extend(["-H", "'{}: {}'".format(key_text, value_text.replace("'", "\\'"))])

        body_text = sanitize_excel_value(item.get("request_body") or "").strip()
        if not body_text:
            request_packet = sanitize_excel_value(item.get("request_packet") or "").strip()
            if request_packet:
                split_token = "\r\n\r\n" if "\r\n\r\n" in request_packet else "\n\n"
                if split_token in request_packet:
                    body_text = request_packet.split(split_token, 1)[1].strip()
        if body_text:
            parts.extend(["--data-raw", "'{}'".format(body_text.replace("'", "\\'"))])

        parts.append("'{}'".format(request_url.replace("'", "\\'")))
        return _truncate_report_text(" ".join(parts), 1600)

    def _build_ai_pen_effective_interfaces_text(item):
        summary = item.get("api_surface_summary") if isinstance(item.get("api_surface_summary"), dict) else {}
        sample_interfaces = [entry for entry in _safe_list(summary.get("sample_interfaces")) if isinstance(entry, dict)]
        runtime_calls = [entry for entry in _safe_list(item.get("runtime_api_calls")) if isinstance(entry, dict)]
        source_items = sample_interfaces if sample_interfaces else runtime_calls
        lines = []
        seen = set()
        for entry in source_items:
            method_text = sanitize_excel_value(entry.get("method", "GET")).strip().upper() or "GET"
            target_text = sanitize_excel_value(
                entry.get("url_template")
                or entry.get("path_template")
                or entry.get("url")
                or entry.get("path")
            ).strip()
            if not target_text:
                continue
            cache_key = "{}|{}".format(method_text, target_text)
            if cache_key in seen:
                continue
            seen.add(cache_key)
            lines.append("{} {}".format(method_text, target_text))
            if len(lines) >= 16:
                break
        return _truncate_report_text("\n".join(lines), 1200)

    def _format_ai_plan_request_text(value):
        parsed = _parse_json_object(value)
        if not parsed:
            return _truncate_report_text(value, 1200)
        lines = []

        def append(label, raw):
            text = sanitize_excel_value(raw).strip()
            if not text:
                return
            lines.append("{}: {}".format(label, text))

        append("目标", parsed.get("target", ""))
        append("漏洞URL", parsed.get("vuln_url", ""))
        append("来源", parsed.get("source_collection", ""))
        append("来源模块", parsed.get("source_module", ""))
        append("风险类型", parsed.get("risk_type", ""))
        append("风险名称", parsed.get("risk_name", ""))
        append("严重级别", parsed.get("severity", ""))
        append("路由提示", parsed.get("route_hint", ""))
        append("默认探针类型", parsed.get("default_payload_type", ""))
        append("默认Payload", parsed.get("default_payload", ""))
        capability_profile = parsed.get("capability_profile", {})
        if isinstance(capability_profile, dict):
            append("能力画像", capability_profile.get("name", ""))
        surface_hints = [sanitize_excel_value(x).strip() for x in (parsed.get("surface_hints") or []) if sanitize_excel_value(x).strip()]
        if surface_hints:
            lines.append("能力线索: {}".format(", ".join(surface_hints[:8])))
        browser_surface_summary = parsed.get("browser_surface_summary", {})
        if isinstance(browser_surface_summary, dict):
            append("页面标题", browser_surface_summary.get("page_title", ""))
            append("页面URL", browser_surface_summary.get("page_url", ""))
        login_surface_summary = parsed.get("login_surface_summary", {})
        if isinstance(login_surface_summary, dict):
            if bool(login_surface_summary.get("login_page_hint")):
                lines.append("登录页提示: 是")
            append("密码表单数", login_surface_summary.get("password_form_count", ""))
            append("验证码表单数", login_surface_summary.get("captcha_form_count", ""))
        return _truncate_report_text("\n".join(lines), 1200)

    def _format_ai_plan_reply_text(value):
        parsed = _parse_json_object(value)
        if not parsed:
            return _truncate_report_text(value, 1200)
        lines = []

        def append(label, raw):
            text = sanitize_excel_value(raw).strip()
            if not text:
                return
            lines.append("{}: {}".format(label, text))

        append("结论", parsed.get("decision", ""))
        append("置信度", parsed.get("confidence", ""))
        append("原因", parsed.get("reason", ""))
        append("探针类型", parsed.get("payload_type", ""))
        append("Payload", parsed.get("payload", ""))
        evidence_list = [sanitize_excel_value(x).strip() for x in (parsed.get("evidence") or []) if sanitize_excel_value(x).strip()]
        if evidence_list:
            lines.append("关键证据:")
            for idx, item in enumerate(evidence_list[:6], 1):
                lines.append("{}. {}".format(idx, item))
        next_actions = [sanitize_excel_value(x).strip() for x in (parsed.get("next_actions") or []) if sanitize_excel_value(x).strip()]
        if next_actions:
            lines.append("下一步动作:")
            for idx, item in enumerate(next_actions[:6], 1):
                lines.append("{}. {}".format(idx, item))
        return _truncate_report_text("\n".join(lines), 1200)

    task_id_list = _normalize_task_id_list(task_ids)
    rows = []
    dedup_keys = set()

    for task_id in task_id_list:
        for item in get_ai_pen_test_data(task_id):
            source_collection = sanitize_excel_value(item.get("source_collection", "")).strip()
            risk_type = sanitize_excel_value(item.get("risk_type", "")).strip()
            risk_name = sanitize_excel_value(item.get("risk_name", "")).strip()
            target = sanitize_excel_value(item.get("target", "")).strip()
            vuln_url = sanitize_excel_value(item.get("vuln_url", "")).strip()
            status = sanitize_excel_value(item.get("status", "")).strip()
            payload = _build_ai_pen_curl_payload(item)
            interface_fetch_summary = _extract_ai_pen_interface_fetch_counts(item).get("summary", "")
            effective_interfaces = _build_ai_pen_effective_interfaces_text(item)
            request_packet = _truncate_report_text(item.get("request_packet", ""), 1200)
            reason = _truncate_report_text(item.get("reason", ""), 800)
            save_date = sanitize_excel_value(item.get("save_date") or item.get("update_date") or "").strip()

            dedup_key = (
                task_id,
                source_collection,
                risk_type,
                risk_name,
                target,
                vuln_url,
                payload,
            )
            if dedup_key in dedup_keys:
                continue
            dedup_keys.add(dedup_key)

            rows.append([
                source_collection,
                risk_type,
                risk_name,
                target,
                status,
                effective_interfaces,
                interface_fetch_summary,
                payload,
                request_packet,
                reason,
                save_date,
            ])

    return rows


def _build_ai_pen_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增 AI 渗透测试工作表。
    """
    ws = wb.create_sheet(title="AI渗透测试")
    for key, width in {
        "A": 12.0,
        "B": 14.0,
        "C": 26.0,
        "D": 34.0,
        "E": 18.0,
        "F": 46.0,
        "G": 18.0,
        "H": 96.0,
        "I": 96.0,
        "J": 88.0,
        "K": 21.0,
    }.items():
        ws.column_dimensions[key].width = width

    ws.append([
        "来源",
        "风险类型",
        "风险名称",
        "目标",
        "状态",
        "有效接口",
        "获取接口",
        "Payload",
        "Request请求包",
        "说明",
        "时间",
    ])

    for row in _extract_ai_pen_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def _cert_record_rank(item):
    """
    证书记录优先级（值越小优先级越高）：
    1) scan_mode=sni
    2) 有 sni_domain
    3) 有 domain
    """
    if not isinstance(item, dict):
        return (9, 9, 9)

    scan_mode = sanitize_excel_value(item.get("scan_mode", "")).strip().lower()
    sni_domain = _normalize_cert_domain(item.get("sni_domain", ""))
    item_domain = _normalize_cert_domain(item.get("domain", ""))

    mode_rank = 0 if scan_mode == "sni" else 1
    sni_rank = 0 if sni_domain else 1
    domain_rank = 0 if item_domain else 1
    return (mode_rank, sni_rank, domain_rank)


def _select_preferred_cert_items(cert_items):
    """
    按 task_id+ip+port 聚合证书记录，优先保留业务域名证书，抑制 default 默认证书干扰。
    """
    if not isinstance(cert_items, list):
        return []

    grouped = {}
    for item in cert_items:
        if not isinstance(item, dict):
            continue

        task_id = sanitize_excel_value(item.get("task_id", "")).strip()
        ip = sanitize_excel_value(item.get("ip", "")).strip()
        port = sanitize_excel_value(item.get("port", "")).strip()
        if not task_id or not ip or not port:
            # 结构异常记录按原样保留，避免误丢数据
            key = ("raw", str(len(grouped)))
            grouped[key] = item
            continue

        key = (task_id, ip, port)
        current = grouped.get(key)
        if not current:
            grouped[key] = item
            continue

        if _cert_record_rank(item) < _cert_record_rank(current):
            grouped[key] = item

    return list(grouped.values())


def _parse_datetime_safe(value):
    """
    兼容多种时间字符串格式，解析失败时返回 None。
    """
    text = sanitize_excel_value(value).strip()
    if not text:
        return None

    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
    ]:
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue

    return None


def _normalize_cert_domain(value):
    """
    归一化证书相关域名文本，非法值返回空字符串。
    """
    domain = utils.normalize_domain(sanitize_excel_value(value))
    if not utils.is_valid_domain(domain):
        return ""
    return domain


def _extract_cert_san_domains(cert_obj):
    """
    从证书扩展字段提取 SAN 域名列表。
    """
    if not isinstance(cert_obj, dict):
        return []

    extensions = cert_obj.get("extensions", {})
    if not isinstance(extensions, dict):
        return []

    san_text = sanitize_excel_value(extensions.get("subjectAltName", ""))
    if not san_text:
        return []

    domains = []
    seen = set()
    for part in san_text.split(","):
        item = sanitize_excel_value(part).strip()
        if not item:
            continue

        if ":" in item:
            prefix, value = item.split(":", 1)
            if prefix.strip().upper() != "DNS":
                continue
            candidate = value.strip()
        else:
            candidate = item

        domain = _normalize_cert_domain(candidate)
        if not domain or domain in seen:
            continue
        seen.add(domain)
        domains.append(domain)

    return domains


def _build_cert_domain_context(task_id):
    """
    构建任务级 IP->域名映射，供 SSL 证书导出补全域名。
    """
    ip_domain_map = {}
    task_domain_set = set()

    for item in get_domain_data(task_id):
        domain = _normalize_cert_domain(item.get("domain", ""))
        if not domain:
            continue
        task_domain_set.add(domain)

        raw_ips = item.get("ips", [])
        if isinstance(raw_ips, str):
            raw_ips = [x.strip() for x in raw_ips.split(",") if x.strip()]
        if not isinstance(raw_ips, list):
            raw_ips = []

        for raw_ip in raw_ips:
            ip = sanitize_excel_value(raw_ip).strip()
            if not ip:
                continue
            ip_domain_map.setdefault(ip, []).append(domain)

    for item in get_ip_data(task_id):
        ip = sanitize_excel_value(item.get("ip", "")).strip()
        if not ip:
            continue

        raw_domains = item.get("domain", [])
        if isinstance(raw_domains, str):
            raw_domains = [raw_domains]
        if not isinstance(raw_domains, list):
            raw_domains = []

        for raw_domain in raw_domains:
            domain = _normalize_cert_domain(raw_domain)
            if not domain:
                continue
            task_domain_set.add(domain)
            ip_domain_map.setdefault(ip, []).append(domain)

    for ip, domains in list(ip_domain_map.items()):
        ordered = []
        seen = set()
        for domain in domains:
            if domain in seen:
                continue
            ordered.append(domain)
            seen.add(domain)
        ip_domain_map[ip] = ordered

    return ip_domain_map, task_domain_set


def _resolve_cert_domain(item, cert_obj, ip_domain_map=None, task_domain_set=None):
    """
    导出域名优先级：
    1) cert 记录中的 sni_domain/domain/domains
    2) 任务内 IP 关联域名
    3) SAN（优先命中任务域名）
    4) CN
    """
    ip_domain_map = ip_domain_map if isinstance(ip_domain_map, dict) else {}
    task_domain_set = task_domain_set if isinstance(task_domain_set, set) else set()

    sni_domain = _normalize_cert_domain(item.get("sni_domain", ""))
    if sni_domain:
        return sni_domain

    item_domain = _normalize_cert_domain(item.get("domain", ""))
    if item_domain:
        return item_domain

    scan_mode = sanitize_excel_value(item.get("scan_mode", "")).strip().lower()
    if scan_mode == "sni":
        item_domains = item.get("domains", [])
        if isinstance(item_domains, str):
            item_domains = [item_domains]
        if isinstance(item_domains, list):
            for domain in item_domains:
                normalized = _normalize_cert_domain(domain)
                if normalized:
                    return normalized

    ip = sanitize_excel_value(item.get("ip", "")).strip()
    mapped_domains = ip_domain_map.get(ip, [])
    if isinstance(mapped_domains, list):
        for domain in mapped_domains:
            if domain:
                return domain

    san_domains = _extract_cert_san_domains(cert_obj)
    if task_domain_set:
        for domain in san_domains:
            if domain in task_domain_set:
                return domain
    if san_domains:
        return san_domains[0]

    subject = cert_obj.get("subject", {}) if isinstance(cert_obj, dict) else {}
    if isinstance(subject, dict):
        common_name = _normalize_cert_domain(subject.get("common_name", ""))
        if common_name:
            return common_name

    return ""


def _extract_protocol_names(ssl_security):
    """
    从证书安全字段中提取协议名称列表。
    """
    if not isinstance(ssl_security, dict):
        return []

    names = []
    protocols = ssl_security.get("protocols", [])
    if isinstance(protocols, list):
        for item in protocols:
            if isinstance(item, dict):
                name = sanitize_excel_value(item.get("name", "")).strip()
            else:
                name = sanitize_excel_value(item).strip()
            if name:
                names.append(name)

    if not names and isinstance(ssl_security.get("protocol_names"), list):
        for name in ssl_security.get("protocol_names", []):
            text = sanitize_excel_value(name).strip()
            if text:
                names.append(text)

    return sorted(list(set(names)))


def _extract_cipher_suite_lines(ssl_security, max_items=50):
    """
    组装加密套件文本（协议 + 套件 + 强度），并限制导出长度。
    """
    if not isinstance(ssl_security, dict):
        return []

    lines = []
    cipher_suites = ssl_security.get("cipher_suites", [])
    if not isinstance(cipher_suites, list):
        return lines

    for item in cipher_suites:
        if not isinstance(item, dict):
            continue
        protocol = sanitize_excel_value(item.get("protocol", "")).strip()
        cipher_name = sanitize_excel_value(item.get("name", "")).strip()
        strength = sanitize_excel_value(item.get("strength", "")).strip().upper()
        if not cipher_name:
            continue
        line = cipher_name
        if protocol:
            line = "[{}] {}".format(protocol, line)
        if strength:
            line = "{} ({})".format(line, strength)
        lines.append(line)

    if len(lines) > max_items:
        hidden = len(lines) - max_items
        lines = lines[:max_items]
        lines.append("... 其余 {} 条省略".format(hidden))

    return lines


def _extract_cert_rows(task_ids):
    """
    汇总 SSL 证书导出行（支持协议/套件/强度与 TLS 合规信息）。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    ai_lookup = _build_ai_denoise_lookup(task_id_list, "cert")
    rows = []
    now_dt = datetime.utcnow()

    for task_id in task_id_list:
        task_id = str(task_id or "").strip()
        if not task_id:
            continue

        ip_domain_map, task_domain_set = _build_cert_domain_context(task_id)
        cert_items = list(get_cert_data(task_id))
        for item in _select_preferred_cert_items(cert_items):
            cert_obj = item.get("cert", {}) if isinstance(item.get("cert"), dict) else {}
            validity = cert_obj.get("validity", {}) if isinstance(cert_obj.get("validity"), dict) else {}
            ssl_security = cert_obj.get("ssl_security", {}) if isinstance(cert_obj.get("ssl_security"), dict) else {}

            ip = sanitize_excel_value(item.get("ip", "")).strip()
            port = sanitize_excel_value(item.get("port", "")).strip()
            host = sanitize_excel_value(item.get("host", "")).strip()
            if not host:
                host = "{}:{}".format(ip, port) if ip and port else ip or port

            domain = _resolve_cert_domain(
                item=item,
                cert_obj=cert_obj,
                ip_domain_map=ip_domain_map,
                task_domain_set=task_domain_set,
            )
            if not domain:
                domain = "-"

            validity_start = sanitize_excel_value(validity.get("start", "")).strip()
            validity_end = sanitize_excel_value(validity.get("end", "")).strip()

            remain_days = ""
            end_dt = _parse_datetime_safe(validity_end)
            if end_dt:
                remain_days = (end_dt - now_dt).days

            protocol_names = _extract_protocol_names(ssl_security)
            protocol_text = "、".join(protocol_names)

            least_strength = sanitize_excel_value(ssl_security.get("least_strength", "")).strip().upper()
            ecdhe_count = ssl_security.get("ecdhe_count", "")
            try:
                ecdhe_count = int(ecdhe_count)
            except Exception:
                ecdhe_count = ""

            cipher_lines = _extract_cipher_suite_lines(ssl_security)
            cipher_text = " \r\n".join(cipher_lines)
            compliance = get_ssl_security_compliance(ssl_security)
            non_compliant_text = ""
            remediation_text = ""
            if isinstance(compliance, dict) and compliance.get("has_issue"):
                non_compliant_text = sanitize_excel_value(compliance.get("non_compliant_text", "")).strip()
                remediation_text = sanitize_excel_value(compliance.get("remediation_text", "")).strip()

            sha256 = ""
            fingerprint = cert_obj.get("fingerprint", {})
            if isinstance(fingerprint, dict):
                sha256 = sanitize_excel_value(fingerprint.get("sha256", "")).strip()

            san = ""
            extensions = cert_obj.get("extensions", {})
            if isinstance(extensions, dict):
                san = sanitize_excel_value(extensions.get("subjectAltName", "")).strip()

            rows.append(
                [
                    sanitize_excel_value(domain),
                    sanitize_excel_value(host),
                    sanitize_excel_value(cert_obj.get("subject_dn", "")),
                    sanitize_excel_value(cert_obj.get("issuer_dn", "")),
                    sanitize_excel_value(validity_start),
                    sanitize_excel_value(validity_end),
                    sanitize_excel_value(remain_days),
                    sanitize_excel_value(protocol_text),
                    sanitize_excel_value(least_strength),
                    sanitize_excel_value(ecdhe_count),
                    sanitize_excel_value(cipher_text),
                    sanitize_excel_value(non_compliant_text),
                    sanitize_excel_value(remediation_text),
                    sanitize_excel_value(sha256),
                    sanitize_excel_value(san),
                    sanitize_excel_value(
                        _resolve_ai_lookup_result(
                            ai_lookup,
                            data_id=_normalize_ai_lookup_key(item.get("_id", "")),
                            row_key=_normalize_ai_lookup_key(item.get("_id", "")),
                        ).get("text", "未分析")
                    ),
                ]
            )

    return rows


def _build_cert_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增 SSL 证书工作表，并补充 TLS 合规审计列。
    """
    ws = wb.create_sheet(title="SSL证书")
    ws.column_dimensions['A'].width = 32.0
    ws.column_dimensions['B'].width = 26.0
    ws.column_dimensions['C'].width = 40.0
    ws.column_dimensions['D'].width = 40.0
    ws.column_dimensions['E'].width = 21.0
    ws.column_dimensions['F'].width = 21.0
    ws.column_dimensions['G'].width = 12.0
    ws.column_dimensions['H'].width = 24.0
    ws.column_dimensions['I'].width = 12.0
    ws.column_dimensions['J'].width = 14.0
    ws.column_dimensions['K'].width = 68.0
    ws.column_dimensions['L'].width = 72.0
    ws.column_dimensions['M'].width = 78.0
    ws.column_dimensions['N'].width = 42.0
    ws.column_dimensions['O'].width = 60.0
    ws.column_dimensions['P'].width = 24.0

    ws.append(
        [
            "域名",
            "HOST",
            "主题名称",
            "签发者名称",
            "生效时间",
            "失效时间",
            "剩余天数",
            "支持协议",
            "最弱强度",
            "ECDHE套件数",
            "加密套件",
            "不合规项（协议/套件）",
            "修复建议",
            "SHA-256",
            "使用者备用名称",
            "AI分析",
        ]
    )

    for row in _extract_cert_rows(task_ids):
        ws.append(row)

    if apply_style:
        beautify_cert_sheet(ws)


def _extract_vuln_rows(task_ids):
    """
    汇总漏洞明细（合并 vuln 与 nuclei_result），并按关键字段去重
    """
    def _safe_load_json_dict(raw_text):
        text = sanitize_excel_value(raw_text).strip()
        if not text:
            return {}
        try:
            payload = json.loads(text)
        except Exception:
            return {}
        return payload if isinstance(payload, dict) else {}

    def _extract_afrog_rule_id(plugin_name):
        plugin_text = sanitize_excel_value(plugin_name).strip()
        if plugin_text.lower().startswith("afrog:"):
            return sanitize_excel_value(plugin_text.split(":", 1)[1]).strip()
        return ""

    def _build_afrog_export_detail(item):
        verify_payload = _safe_load_json_dict(item.get("verify_data", ""))
        plugin_name = sanitize_excel_value(item.get("plg_name", "")).strip()
        vuln_name = sanitize_excel_value(item.get("vul_name", "")).strip()
        severity = sanitize_excel_value(item.get("severity", "")).strip().lower()
        target = sanitize_excel_value(item.get("target", "")).strip()
        poc_id = (
            sanitize_excel_value(verify_payload.get("id", "")).strip()
            or _extract_afrog_rule_id(plugin_name)
        )

        references = verify_payload.get("reference", [])
        if isinstance(references, str):
            references = [references]
        if not isinstance(references, list):
            references = []
        references = [
            _truncate_report_text(item_text, 140)
            for item_text in references
            if _truncate_report_text(item_text, 140)
        ][:2]

        parts = ["source=afrog", "poc_id={}".format(poc_id or "-")]
        if vuln_name:
            parts.append("name={}".format(_truncate_report_text(vuln_name, 120)))
        if severity:
            parts.append("severity={}".format(_truncate_report_text(severity, 24)))
        if target:
            parts.append("target={}".format(_truncate_report_text(target, 180)))
        if references:
            parts.append("reference={}".format(" ; ".join(references)))

        request_text = _truncate_report_text(verify_payload.get("request"), 180)
        response_text = _truncate_report_text(verify_payload.get("response"), 180)
        if request_text:
            parts.append("request={}".format(request_text))
        if response_text:
            parts.append("response={}".format(response_text))

        return _truncate_report_text(" | ".join(parts), 900)

    def _resolve_vuln_detail_text(item):
        description = sanitize_excel_value(item.get("description", "")).strip()
        detail = sanitize_excel_value(item.get("detail", "")).strip()
        verify_data = sanitize_excel_value(item.get("verify_data", "")).strip()
        vuln_type = sanitize_excel_value(item.get("plg_type", "")).strip().lower()
        plugin_name = sanitize_excel_value(item.get("plg_name", "")).strip().lower()

        if vuln_type == "afrog" or plugin_name.startswith("afrog"):
            detail_lower = detail.lower()
            detail_is_placeholder = (
                not detail
                or detail_lower in ("source=afrog", "source=afrog poc_id=-", "source=afrog poc_id=")
                or (
                    detail_lower.startswith("source=afrog")
                    and "poc_id=-" in detail_lower
                    and len(detail_lower) <= 40
                )
            )
            if detail_is_placeholder:
                rebuilt = _build_afrog_export_detail(item)
                if rebuilt:
                    return rebuilt

            return _truncate_report_text(detail or description or verify_data, 900)

        return _truncate_report_text(description or detail or verify_data, 900)

    task_id_list = _normalize_task_id_list(task_ids)
    vuln_ai_lookup = _build_ai_denoise_lookup(task_id_list, "vuln")
    nuclei_ai_lookup = _build_ai_denoise_lookup(task_id_list, "nuclei_result")
    rows = []
    dedup_keys = set()

    for task_id in task_id_list:
        task_id = str(task_id or "").strip()
        if not task_id:
            continue

        for item in get_vuln_data(task_id):
            vuln_name = sanitize_excel_value(item.get("vul_name", ""))
            severity = sanitize_excel_value(item.get("severity", ""))
            target = sanitize_excel_value(item.get("target", ""))
            vuln_url = target if str(target).startswith("http") else ""
            plugin = sanitize_excel_value(item.get("plg_name", ""))
            vuln_type = sanitize_excel_value(item.get("plg_type", ""))
            detail = _resolve_vuln_detail_text(item)

            dedup_key = (
                task_id, "npoc", vuln_name, severity, target, vuln_url, plugin, vuln_type
            )
            if dedup_key in dedup_keys:
                continue
            dedup_keys.add(dedup_key)
            item_id = _normalize_ai_lookup_key(item.get("_id", ""))
            ai_result = _resolve_ai_lookup_result(vuln_ai_lookup, data_id=item_id, row_key=item_id)
            rows.append(
                [
                    "npoc",
                    vuln_name,
                    severity,
                    target,
                    vuln_url,
                    plugin,
                    vuln_type,
                    detail,
                    sanitize_excel_value(ai_result.get("text", "未分析")),
                ]
            )

        for item in get_nuclei_result_data(task_id):
            vuln_name = sanitize_excel_value(item.get("vuln_name", ""))
            severity = sanitize_excel_value(item.get("vuln_severity", ""))
            target = sanitize_excel_value(item.get("target", ""))
            vuln_url = sanitize_excel_value(item.get("vuln_url", ""))
            template_id = sanitize_excel_value(item.get("template_id", ""))
            template_url = sanitize_excel_value(item.get("template_url", ""))

            dedup_key = (
                task_id, "nuclei", vuln_name, severity, target, vuln_url, template_id
            )
            if dedup_key in dedup_keys:
                continue
            dedup_keys.add(dedup_key)
            item_id = _normalize_ai_lookup_key(item.get("_id", ""))
            ai_result = _resolve_ai_lookup_result(nuclei_ai_lookup, data_id=item_id, row_key=item_id)
            rows.append(
                [
                    "nuclei",
                    vuln_name,
                    severity,
                    target,
                    vuln_url,
                    template_id,
                    "nuclei",
                    template_url,
                    sanitize_excel_value(ai_result.get("text", "未分析")),
                ]
            )

    return rows


def build_task_export_summary(task_ids):
    """
    基于导出口径生成任务汇总，保证通知/知识库概览与实际报告一致。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    valid_task_ids = []
    seen_valid_task_ids = set()
    task_states = {}

    for raw_task_id in task_id_list:
        task_data = get_task_data(raw_task_id)
        if not task_data:
            continue

        task_id = sanitize_excel_value(task_data.get("_id", raw_task_id)).strip()
        if not task_id or task_id in seen_valid_task_ids:
            continue

        seen_valid_task_ids.add(task_id)
        valid_task_ids.append(task_id)
        task_states[task_id] = {
            "site_keys": set(),
            "domain_keys": set(),
            "ip_cnt": 0,
            "url_keys": set(),
            "vuln_keys": set(),
        }

    merged_site_keys = set()
    merged_domain_keys = set()
    merged_ip_cnt = 0
    merged_url_keys = set()
    merged_vuln_keys = set()

    for task_id in valid_task_ids:
        state = task_states[task_id]

        for site_item in get_site_data(task_id):
            site = sanitize_excel_value(site_item.get("site") or site_item.get("url") or "").strip()
            if not site:
                continue
            state["site_keys"].add(site)
            merged_site_keys.add(site)

        for domain_item in get_domain_data(task_id):
            domain = sanitize_excel_value(domain_item.get("domain", "")).strip()
            if not domain:
                continue
            state["domain_keys"].add(domain)
            merged_domain_keys.add(domain)

        for ip_item in get_ip_data(task_id):
            ip = sanitize_excel_value(ip_item.get("ip", "")).strip()
            if not ip:
                continue
            state["ip_cnt"] += 1
            merged_ip_cnt += 1

        for item in get_url_data(task_id):
            row = (
                sanitize_excel_value(item.get("url", "")),
                sanitize_excel_value(item.get("site", "")),
                sanitize_excel_value(item.get("title", "")),
                sanitize_excel_value(item.get("status_code", "")),
                sanitize_excel_value(item.get("content_length", "")),
                sanitize_excel_value(item.get("source", "")),
            )
            state["url_keys"].add(row)
            merged_url_keys.add(row)

        for item in get_vuln_data(task_id):
            vuln_name = sanitize_excel_value(item.get("vul_name", ""))
            severity = sanitize_excel_value(item.get("severity", ""))
            target = sanitize_excel_value(item.get("target", ""))
            vuln_url = target if str(target).startswith("http") else ""
            plugin = sanitize_excel_value(item.get("plg_name", ""))
            vuln_type = sanitize_excel_value(item.get("plg_type", ""))
            dedup_key = (
                task_id,
                "npoc",
                vuln_name,
                severity,
                target,
                vuln_url,
                plugin,
                vuln_type,
            )
            state["vuln_keys"].add(dedup_key)
            merged_vuln_keys.add(dedup_key)

        for item in get_nuclei_result_data(task_id):
            vuln_name = sanitize_excel_value(item.get("vuln_name", ""))
            severity = sanitize_excel_value(item.get("vuln_severity", ""))
            target = sanitize_excel_value(item.get("target", ""))
            vuln_url = sanitize_excel_value(item.get("vuln_url", ""))
            template_id = sanitize_excel_value(item.get("template_id", ""))
            dedup_key = (
                task_id,
                "nuclei",
                vuln_name,
                severity,
                target,
                vuln_url,
                template_id,
            )
            state["vuln_keys"].add(dedup_key)
            merged_vuln_keys.add(dedup_key)

    task_summaries = {}
    for task_id in valid_task_ids:
        state = task_states.get(task_id, {})
        task_summaries[task_id] = {
            "site_cnt": len(state.get("site_keys", set())),
            "domain_cnt": len(state.get("domain_keys", set())),
            "ip_cnt": int(state.get("ip_cnt", 0) or 0),
            "url_cnt": len(state.get("url_keys", set())),
            "vuln_cnt": len(state.get("vuln_keys", set())),
        }

    return {
        "task_ids": valid_task_ids,
        "site_cnt": len(merged_site_keys),
        "domain_cnt": len(merged_domain_keys),
        "ip_cnt": merged_ip_cnt,
        "url_cnt": len(merged_url_keys),
        "vuln_cnt": len(merged_vuln_keys),
        "task_summaries": task_summaries,
    }


def _severity_rank(severity_text):
    """
    风险等级排序权重（值越大越高）。
    """
    level = sanitize_excel_value(severity_text).strip().lower()
    return {
        "critical": 5,
        "high": 4,
        "medium": 3,
        "low": 2,
        "info": 1,
    }.get(level, 0)


def _build_vuln_severity_distribution(vuln_rows):
    """
    统计风险等级分布（critical/high/medium/low/info）。
    """
    stat = {
        "critical": 0,
        "high": 0,
        "medium": 0,
        "low": 0,
        "info": 0,
    }
    for row in vuln_rows:
        if not isinstance(row, list) or len(row) < 3:
            continue
        raw = sanitize_excel_value(row[2]).strip().lower()
        if raw in ("critical", "严重"):
            stat["critical"] += 1
        elif raw in ("high", "高"):
            stat["high"] += 1
        elif raw in ("medium", "中"):
            stat["medium"] += 1
        elif raw in ("low", "低"):
            stat["low"] += 1
        elif raw in ("info", "informational", "信息"):
            stat["info"] += 1
    return stat


def _split_scan_target_items(raw_target):
    """
    解析任务目标文本，兼容换行/逗号/分号分隔。
    """
    text = sanitize_excel_value(raw_target).replace("\r", "\n").strip()
    if not text:
        return []

    items = []
    seen = set()
    for part in re.split(r"[\n,，;；]+", text):
        target = sanitize_excel_value(part).strip()
        if not target or target in seen:
            continue
        seen.add(target)
        items.append(target)
    return items


def _build_scan_target_rows(task_items, per_task_limit=60, total_limit=300):
    """
    构建“扫描目标清单”表格数据。
    """
    rows = []
    total_count = 0
    for item in task_items:
        task_name = sanitize_excel_value(item.get("name", "")).strip() or "-"
        targets = _split_scan_target_items(item.get("target", ""))
        if not targets:
            fallback_target = sanitize_excel_value(item.get("target", "")).strip()
            targets = [fallback_target] if fallback_target else ["-"]

        for idx, target in enumerate(targets):
            if idx >= per_task_limit or total_count >= total_limit:
                break
            rows.append(
                {
                    "task_name": task_name,
                    "target": _truncate_report_text(target, 220),
                }
            )
            total_count += 1

        if total_count >= total_limit:
            break

    return rows


def _build_task_asset_rows(task_items, summary):
    """
    构建按任务维度的资产产出统计行。
    """
    rows = []
    summary_map = summary.get("task_summaries", {}) if isinstance(summary, dict) else {}
    for item in task_items:
        task_id = sanitize_excel_value(item.get("_id", "")).strip()
        task_name = sanitize_excel_value(item.get("name", "")).strip() or "-"
        task_target = _truncate_report_text(item.get("target", ""), 120) or "-"
        task_stat = summary_map.get(task_id, {})
        rows.append(
            {
                "task_name": task_name,
                "task_target": task_target,
                "site_cnt": int(task_stat.get("site_cnt", 0) or 0),
                "domain_cnt": int(task_stat.get("domain_cnt", 0) or 0),
                "ip_cnt": int(task_stat.get("ip_cnt", 0) or 0),
                "url_cnt": int(task_stat.get("url_cnt", 0) or 0),
                "vuln_cnt": int(task_stat.get("vuln_cnt", 0) or 0),
            }
        )
    return rows


def _collect_asset_samples(task_ids, sample_limit=20):
    """
    汇总资产样本，展示“得到了哪些资产”。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    sample_limit = max(1, int(sample_limit or 20))

    site_seen = set()
    domain_seen = set()
    ip_seen = set()
    url_seen = set()

    site_samples = []
    domain_samples = []
    ip_samples = []
    url_samples = []

    for task_id in task_id_list:
        for item in get_site_data(task_id):
            site = sanitize_excel_value(item.get("site") or item.get("url") or "").strip()
            if not site or site in site_seen:
                continue
            site_seen.add(site)
            if len(site_samples) < sample_limit:
                site_samples.append(_truncate_report_text(site, 180))

        for item in get_domain_data(task_id):
            domain = sanitize_excel_value(item.get("domain", "")).strip()
            if not domain or domain in domain_seen:
                continue
            domain_seen.add(domain)
            if len(domain_samples) < sample_limit:
                domain_samples.append(_truncate_report_text(domain, 180))

        for item in get_ip_data(task_id):
            ip = sanitize_excel_value(item.get("ip", "")).strip()
            if not ip or ip in ip_seen:
                continue
            ip_seen.add(ip)
            if len(ip_samples) < sample_limit:
                ip_samples.append(_truncate_report_text(ip, 180))

        for item in get_url_data(task_id):
            url = sanitize_excel_value(item.get("url", "")).strip()
            if not url or url in url_seen:
                continue
            url_seen.add(url)
            if len(url_samples) < sample_limit:
                url_samples.append(_truncate_report_text(url, 180))

    return {
        "site_total": len(site_seen),
        "domain_total": len(domain_seen),
        "ip_total": len(ip_seen),
        "url_total": len(url_seen),
        "site_samples": site_samples,
        "domain_samples": domain_samples,
        "ip_samples": ip_samples,
        "url_samples": url_samples,
    }


def _build_risk_cluster_rows(vuln_rows, limit=15):
    """
    将风险明细聚合为“风险名称 + 来源 + 最高等级 + 数量”。
    """
    counter = {}
    for row in vuln_rows:
        if not isinstance(row, list) or len(row) < 3:
            continue
        source = sanitize_excel_value(row[0]).strip() or "unknown"
        vuln_name = sanitize_excel_value(row[1]).strip() or "未知风险"
        severity = sanitize_excel_value(row[2]).strip().lower()
        key = (source, vuln_name)
        item = counter.get(key, {"count": 0, "severity": severity})
        item["count"] += 1
        if _severity_rank(severity) > _severity_rank(item.get("severity", "")):
            item["severity"] = severity
        counter[key] = item

    cluster_rows = []
    for (source, vuln_name), item in counter.items():
        cluster_rows.append(
            {
                "source": source,
                "vuln_name": vuln_name,
                "severity": item.get("severity", ""),
                "count": int(item.get("count", 0) or 0),
            }
        )

    cluster_rows.sort(
        key=lambda x: (
            -_severity_rank(x.get("severity", "")),
            -int(x.get("count", 0) or 0),
            x.get("vuln_name", ""),
        )
    )
    return cluster_rows[:limit]


def _build_suspected_fp_rows(vuln_rows, limit=12):
    """
    规则化筛选“疑似误报”候选，辅助人工复核。
    """
    fp_keywords = [
        "响应长度差异",
        "响应内容结构变化",
        "疑似",
        "可能",
        "结构变化",
    ]

    suspects = []
    seen = set()
    for row in vuln_rows:
        if not isinstance(row, list) or len(row) < 8:
            continue
        severity = sanitize_excel_value(row[2]).strip().lower()
        detail = sanitize_excel_value(row[7]).strip()
        if not detail:
            continue

        low_confidence = _severity_rank(severity) <= _severity_rank("low")
        weak_signal = any(keyword in detail for keyword in fp_keywords)
        if not (low_confidence and weak_signal):
            continue

        source = sanitize_excel_value(row[0]).strip() or "unknown"
        vuln_name = sanitize_excel_value(row[1]).strip() or "未知风险"
        target = sanitize_excel_value(row[3]).strip()
        key = (source, vuln_name, target, detail[:120])
        if key in seen:
            continue
        seen.add(key)
        suspects.append(
            {
                "source": source,
                "vuln_name": vuln_name,
                "target": target,
                "reason": detail[:220],
            }
        )
        if len(suspects) >= limit:
            break

    return suspects


def _build_ai_markdown_report(task_ids, ai_settings):
    """
    生成 AI 报告导出用 Markdown 模板（基于现有扫描数据）。
    """
    task_id_list = _normalize_task_id_list(task_ids)
    if not task_id_list:
        raise ValueError("未找到可导出的任务数据")

    task_items = []
    for task_id in task_id_list:
        item = get_task_data(task_id)
        if isinstance(item, dict):
            task_items.append(item)

    summary = build_task_export_summary(task_id_list)
    vuln_rows = _extract_vuln_rows(task_id_list)
    poc_rows = _extract_nuclei_rows(task_id_list)
    waf_rows = _extract_waf_rows(task_id_list)
    wih_rows = _extract_wih_rows(task_id_list)
    ai_denoise_rows = _extract_ai_denoise_rows(task_id_list)

    severity_stat = _build_vuln_severity_distribution(vuln_rows)
    risk_clusters = _build_risk_cluster_rows(vuln_rows, limit=15)
    suspected_fp_rows = _build_suspected_fp_rows(vuln_rows, limit=12)
    ai_overview = _build_ai_denoise_overview(ai_denoise_rows)
    ai_high_value_rows = _build_ai_high_value_rows(ai_denoise_rows, limit=30)
    ai_suspected_fp_rows = _build_ai_suspected_fp_rows(ai_denoise_rows, limit=20)
    scan_target_rows = _build_scan_target_rows(task_items, per_task_limit=80, total_limit=400)
    task_asset_rows = _build_task_asset_rows(task_items, summary)
    asset_samples = _collect_asset_samples(task_id_list, sample_limit=25)

    ai_denoise_enabled = True
    raw_ai_denoise_enable = ""
    if isinstance(ai_settings, dict):
        raw_ai_denoise_enable = sanitize_excel_value(ai_settings.get("ai_denoise_enable", "")).strip().lower()
    if raw_ai_denoise_enable in ("0", "false", "off", "no"):
        ai_denoise_enabled = False

    names = []
    targets = []
    start_values = []
    end_values = []
    for item in task_items:
        name = sanitize_excel_value(item.get("name", "")).strip()
        target = sanitize_excel_value(item.get("target", "")).strip()
        start_time = sanitize_excel_value(item.get("start_time", "")).strip()
        end_time = sanitize_excel_value(item.get("end_time", "")).strip()
        if name and name not in names:
            names.append(name)
        if target and target not in targets:
            targets.append(target)
        if start_time:
            start_values.append(start_time)
        if end_time:
            end_values.append(end_time)

    scan_start = min(start_values) if start_values else "-"
    scan_end = max(end_values) if end_values else "-"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    lines = []
    lines.append("# ARL AI分析报告")
    lines.append("")
    lines.append("> 生成时间：`{}`".format(generated_at))
    lines.append("> 扫描开始时间：`{}`".format(scan_start))
    lines.append("> 扫描截止时间：`{}`".format(scan_end))
    lines.append("> 报告类型：`AI分析报告固定模板 V2`")
    lines.append("> 生成方式：`离线结构化汇总（仅读取扫描与AI去噪落库结果，不触发实时模型调用）`")
    lines.append("")
    lines.append("## 任务概览")
    lines.append("")
    lines.append("| 任务名 | 目标 |")
    lines.append("| --- | --- |")
    if names or targets:
        max_len = max(len(names), len(targets))
        for idx in range(max_len):
            name = names[idx] if idx < len(names) else "-"
            target = targets[idx] if idx < len(targets) else "-"
            lines.append("| {} | {} |".format(name, target))
    else:
        lines.append("| - | - |")
    lines.append("")
    lines.append("## 扫描目标清单")
    lines.append("")
    lines.append("| 任务名 | 扫描目标 |")
    lines.append("| --- | --- |")
    if scan_target_rows:
        for item in scan_target_rows:
            lines.append(
                "| {} | {} |".format(
                    sanitize_excel_value(item.get("task_name", "")),
                    sanitize_excel_value(item.get("target", "")),
                )
            )
    else:
        lines.append("| - | - |")
    lines.append("")
    lines.append("## 任务资产产出统计")
    lines.append("")
    lines.append("| 任务名 | 任务目标 | 站点 | 子域名 | IP | URL | 风险 |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- |")
    if task_asset_rows:
        for item in task_asset_rows:
            lines.append(
                "| {} | {} | {} | {} | {} | {} | {} |".format(
                    sanitize_excel_value(item.get("task_name", "")),
                    sanitize_excel_value(item.get("task_target", "")),
                    int(item.get("site_cnt", 0) or 0),
                    int(item.get("domain_cnt", 0) or 0),
                    int(item.get("ip_cnt", 0) or 0),
                    int(item.get("url_cnt", 0) or 0),
                    int(item.get("vuln_cnt", 0) or 0),
                )
            )
    else:
        lines.append("| - | - | 0 | 0 | 0 | 0 | 0 |")
    lines.append("")
    lines.append("## 资产样本（节选）")
    lines.append("")
    lines.append("- 站点资产（总数 `{}`，展示 `{}`）".format(
        int(asset_samples.get("site_total", 0) or 0),
        len(asset_samples.get("site_samples", [])),
    ))
    if asset_samples.get("site_samples"):
        for idx, value in enumerate(asset_samples.get("site_samples", []), start=1):
            lines.append("{}. {}".format(idx, sanitize_excel_value(value)))
    else:
        lines.append("- 无")
    lines.append("")
    lines.append("- 子域名资产（总数 `{}`，展示 `{}`）".format(
        int(asset_samples.get("domain_total", 0) or 0),
        len(asset_samples.get("domain_samples", [])),
    ))
    if asset_samples.get("domain_samples"):
        for idx, value in enumerate(asset_samples.get("domain_samples", []), start=1):
            lines.append("{}. {}".format(idx, sanitize_excel_value(value)))
    else:
        lines.append("- 无")
    lines.append("")
    lines.append("- IP资产（总数 `{}`，展示 `{}`）".format(
        int(asset_samples.get("ip_total", 0) or 0),
        len(asset_samples.get("ip_samples", [])),
    ))
    if asset_samples.get("ip_samples"):
        for idx, value in enumerate(asset_samples.get("ip_samples", []), start=1):
            lines.append("{}. {}".format(idx, sanitize_excel_value(value)))
    else:
        lines.append("- 无")
    lines.append("")
    lines.append("- URL资产（总数 `{}`，展示 `{}`）".format(
        int(asset_samples.get("url_total", 0) or 0),
        len(asset_samples.get("url_samples", [])),
    ))
    if asset_samples.get("url_samples"):
        for idx, value in enumerate(asset_samples.get("url_samples", []), start=1):
            lines.append("{}. {}".format(idx, sanitize_excel_value(value)))
    else:
        lines.append("- 无")
    lines.append("")
    lines.append("## 执行摘要（固定模板）")
    lines.append("")
    lines.append("- 整体风险态势：`待人工研判`")
    lines.append("- 高优先级处置方向：`优先处理高危漏洞、弱口令与认证缺失入口`")
    lines.append("- 误报复核建议：`优先复核低危且弱信号风险项`")
    lines.append("")
    lines.append("## 关键资产")
    lines.append("")
    lines.append("- 站点：`{}`".format(summary.get("site_cnt", 0)))
    lines.append("- 子域名：`{}`".format(summary.get("domain_cnt", 0)))
    lines.append("- IP：`{}`".format(summary.get("ip_cnt", 0)))
    lines.append("- URL信息：`{}`".format(summary.get("url_cnt", 0)))
    lines.append("- 风险总数：`{}`".format(summary.get("vuln_cnt", 0)))
    lines.append("- PoC风险：`{}`".format(len(poc_rows)))
    lines.append("- WAF识别：`{}`".format(len(waf_rows)))
    lines.append("- WIH记录：`{}`".format(len(wih_rows)))
    lines.append("")
    lines.append("## 风险等级分布")
    lines.append("")
    lines.append("| 严重级别 | 数量 |")
    lines.append("| --- | --- |")
    lines.append("| 严重 | {} |".format(int(severity_stat.get("critical", 0) or 0)))
    lines.append("| 高危 | {} |".format(int(severity_stat.get("high", 0) or 0)))
    lines.append("| 中危 | {} |".format(int(severity_stat.get("medium", 0) or 0)))
    lines.append("| 低危 | {} |".format(int(severity_stat.get("low", 0) or 0)))
    lines.append("| 信息 | {} |".format(int(severity_stat.get("info", 0) or 0)))
    lines.append("")
    lines.append("## AI去噪概览")
    lines.append("")
    lines.append("- AI去噪配置开关：`{}`".format("开启" if ai_denoise_enabled else "关闭"))
    lines.append("- AI去噪落库记录：`{}`".format(ai_overview.get("total", 0)))
    lines.append("- 已完成分析（AI/规则）：`{}`".format(ai_overview.get("analyzed", 0)))
    lines.append("- 高价值目标（危险/可疑）：`{}`".format(ai_overview.get("high_value", 0)))
    lines.append("- 疑似误报候选：`{}`".format(ai_overview.get("suspected_fp", 0)))
    lines.append(
        "- 分析来源分布：`AI {}` / `规则 {}` / `未分析 {}`".format(
            ai_overview.get("source_stat", {}).get("ai", 0),
            ai_overview.get("source_stat", {}).get("rule", 0),
            ai_overview.get("source_stat", {}).get("disabled", 0),
        )
    )
    lines.append(
        "- 结果级别分布：`危险 {}` / `可疑 {}` / `正常 {}` / `未分析 {}`".format(
            ai_overview.get("result_stat", {}).get("danger", 0),
            ai_overview.get("result_stat", {}).get("suspicious", 0),
            ai_overview.get("result_stat", {}).get("safe", 0),
            ai_overview.get("result_stat", {}).get("disabled", 0),
        )
    )
    lines.append("")
    lines.append("| 模块 | 记录数 | 危险 | 可疑 | 正常 | AI模型 | 规则 | 未分析 | 疑似误报 |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- | --- | --- |")
    if ai_overview.get("module_rows"):
        for item in ai_overview.get("module_rows", []):
            lines.append(
                "| {} | {} | {} | {} | {} | {} | {} | {} | {} |".format(
                    sanitize_excel_value(item.get("module_label", "")),
                    int(item.get("total", 0) or 0),
                    int(item.get("danger", 0) or 0),
                    int(item.get("suspicious", 0) or 0),
                    int(item.get("safe", 0) or 0),
                    int(item.get("ai", 0) or 0),
                    int(item.get("rule", 0) or 0),
                    int(item.get("disabled", 0) or 0),
                    int(item.get("suspected_fp", 0) or 0),
                )
            )
    else:
        lines.append("| - | 0 | 0 | 0 | 0 | 0 | 0 | 0 | 0 |")
    lines.append("")
    lines.append("## AI高价值目标（危险/可疑）")
    lines.append("")
    lines.append("| 模块 | 目标 | 结论 | 来源 | 风险等级 | 可信度 | 分析时间 | 摘要 |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- | --- |")
    if ai_high_value_rows:
        for item in ai_high_value_rows:
            lines.append(
                "| {} | {} | {} | {} | {} | {} | {} | {} |".format(
                    sanitize_excel_value(item.get("module_label", "")),
                    sanitize_excel_value(item.get("target", "")),
                    sanitize_excel_value(item.get("result_label", "")),
                    sanitize_excel_value(item.get("source_label", "")),
                    sanitize_excel_value(item.get("risk_level", "")),
                    sanitize_excel_value(item.get("trust", "")),
                    sanitize_excel_value(item.get("analyzed_at", "")),
                    sanitize_excel_value(
                        _truncate_report_text(item.get("summary") or item.get("display_text"), 140)
                    ),
                )
            )
    else:
        lines.append("| - | - | - | - | - | - | - | - |")
    lines.append("")
    lines.append("## AI疑似误报候选")
    lines.append("")
    lines.append("| 模块 | 目标 | 结论 | 来源 | 风险等级 | 分析时间 | 依据摘要 |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- |")
    if ai_suspected_fp_rows:
        for item in ai_suspected_fp_rows:
            evidence_text = ""
            if isinstance(item.get("evidence"), list) and item.get("evidence"):
                evidence_text = "；".join(item.get("evidence", []))
            lines.append(
                "| {} | {} | {} | {} | {} | {} | {} |".format(
                    sanitize_excel_value(item.get("module_label", "")),
                    sanitize_excel_value(item.get("target", "")),
                    sanitize_excel_value(item.get("result_label", "")),
                    sanitize_excel_value(item.get("source_label", "")),
                    sanitize_excel_value(item.get("risk_level", "")),
                    sanitize_excel_value(item.get("analyzed_at", "")),
                    sanitize_excel_value(_truncate_report_text(evidence_text or item.get("summary"), 140)),
                )
            )
    else:
        lines.append("| - | - | - | - | - | - | - |")
    lines.append("")
    lines.append("## 风险聚类")
    lines.append("")
    lines.append("| 来源 | 风险名称 | 最高等级 | 数量 |")
    lines.append("| --- | --- | --- | --- |")
    if risk_clusters:
        for item in risk_clusters:
            lines.append(
                "| {} | {} | {} | {} |".format(
                    sanitize_excel_value(item.get("source", "")),
                    sanitize_excel_value(item.get("vuln_name", "")),
                    sanitize_excel_value(item.get("severity", "")),
                    int(item.get("count", 0) or 0),
                )
            )
    else:
        lines.append("| - | - | - | 0 |")
    lines.append("")
    lines.append("## 规则误报疑似项")
    lines.append("")
    if suspected_fp_rows:
        for idx, item in enumerate(suspected_fp_rows, start=1):
            lines.append(
                "{}. [{}] {} | 目标：`{}` | 依据：{}".format(
                    idx,
                    sanitize_excel_value(item.get("source", "")),
                    sanitize_excel_value(item.get("vuln_name", "")),
                    sanitize_excel_value(item.get("target", "")),
                    sanitize_excel_value(item.get("reason", "")),
                )
            )
    else:
        lines.append("- 未发现明显的低置信弱信号误报候选。")
    lines.append("")
    lines.append("## 优先修复建议")
    lines.append("")
    lines.append("1. 优先处理 `critical/high` 风险项，先修复可被外网直接访问且存在认证缺失的入口。")
    lines.append("2. 对同一来源的重复风险按“单漏洞多目标”方式集中修复，降低修复切换成本。")
    lines.append("3. 若命中 `WAF识别` 且厂商为 `unknown`，建议补做人工指纹确认后再决定绕过策略。")
    lines.append("")
    lines.append("## 复测建议")
    lines.append("")
    lines.append("- 修复后建议对高风险目标执行一次完整复测，并对误报疑似项执行最小化人工验证。")
    lines.append("- 对云凭证/令牌类问题建议同步轮换密钥并追踪近 30 天相关调用日志。")
    lines.append("- 对 `PoC风险` 建议保留 `verify_data/curl` 作为复测脚本，避免人工复现偏差。")
    lines.append("")
    lines.append("## 说明")
    lines.append("")
    lines.append("- 本报告为 AI分析报告固定模板，内容基于当前任务扫描结果自动结构化生成。")
    lines.append("- 若需更强语义总结，可在后续版本接入在线模型推理后扩展。")
    lines.append("")

    return "\n".join(lines).encode("utf-8")


def _build_vuln_sheet(wb, task_ids, apply_style=True):
    """
    在导出工作簿中新增风险明细工作表
    """
    ws = wb.create_sheet(title="风险")
    ws.column_dimensions['A'].width = 12.0
    ws.column_dimensions['B'].width = 36.0
    ws.column_dimensions['C'].width = 14.0
    ws.column_dimensions['D'].width = 36.0
    ws.column_dimensions['E'].width = 60.0
    ws.column_dimensions['F'].width = 28.0
    ws.column_dimensions['G'].width = 20.0
    ws.column_dimensions['H'].width = 80.0
    ws.column_dimensions['I'].width = 24.0

    ws.append(["来源", "风险名称", "严重级别", "目标", "风险URL", "模板/插件", "风险类型", "详情", "AI分析"])
    for row in _extract_vuln_rows(task_ids):
        ws.append(row)

    if apply_style:
        set_sheet_style(ws)


def port_service_product_statist(task_id):
    """
    端口和服务统计分析
    
    参数：
        task_id: 任务ID
    
    返回：
        tuple: (端口Top20列表, 服务Top20列表)
    
    说明：
    - 统计开放端口的分布情况
    - 统计识别的服务类型分布
    - 返回Top20排行榜
    """
    ip_data = get_ip_data(task_id)
    total = 0
    port_info_list = []
    
    # 收集所有端口信息
    for item in ip_data:
        if not item["port_info"]:
            continue
        port_info_list.extend(item["port_info"])
        total += len(item["port_info"])

    # 统计端口分布Top20
    counter = Counter([info["port_id"] for info in port_info_list])
    top_20 = counter.most_common(20)
    port_percent_list = []
    for port_info in top_20:
        port_id, amount = port_info
        item = {
            "port_id" : port_id,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / total)
        }
        port_percent_list.append(item)

    # 统计服务类型分布
    service_name_list = []
    for info in port_info_list:
        if  not  info.get("product"):
            continue
        if info["product"] or info["version"]:
            service_name = info["service_name"]
            if service_name == "https-alt":
                service_name = "https"

            service_name_list.append(service_name)

    service_top_20 = Counter(service_name_list).most_common(20)

    service_percent_list = []
    for port_info in service_top_20:
        service_name, amount = port_info
        item = {
            "service_name" : service_name,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / len(service_name_list))
        }
        service_percent_list.append(item)



    product_name_list = []
    for info in port_info_list:
        if not info.get("product"):
            continue
        product = info["product"]
        if product and "**" not in product:
            product = product.strip()
            product_name_list.append(product)

    product_top_20 = Counter(product_name_list).most_common(20)
    product_percent_list = []
    for info in product_top_20:
        product, amount = info
        item = {
            "product" : product,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / len(product_name_list))
        }
        product_percent_list.append(item)

    statist = {
        "port_total": total, #端口开放总数
        "port_percent_list": port_percent_list, #端口开放 top 20比例详情
        "service_total": len(service_name_list),  #系统服务类别总数
        "service_percent_list": service_percent_list, #系统服务类别 top 20比例详情
        "product_total": len(product_name_list), #产品种类总数
        "product_percent_list": product_percent_list ##产品种类总数 top 20比例详情
    }
    return statist



class SaveTask(object):
    """docstring for ClassName"""

    def __init__(self, task_id, apply_style=True):
        self.task_id = task_id
        self.wb = Workbook()
        self.is_ip_task = False
        self.apply_style = bool(apply_style)

    def set_style(self, ws):
        if self.apply_style:
            set_sheet_style(ws)

    def build_service_xl(self):
        ws = self.wb.create_sheet(title="系统服务")
        ws.column_dimensions['A'].width = 22.0
        ws.column_dimensions['B'].width = 10.0
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 40.0

        column_tilte = ["IP", "端口","服务", "产品", "版本"]
        ws.append(column_tilte)
        fallback_ip_items = list(get_ip_data(self.task_id))
        for row in _build_service_rows([self.task_id], fallback_ip_items=fallback_ip_items):
            ws.append([
                sanitize_excel_value(row[0]),
                sanitize_excel_value(row[1]),
                sanitize_excel_value(row[2]),
                sanitize_excel_value(row[3]),
                sanitize_excel_value(row[4]),
            ])

        self.set_style(ws)

    def build_ip_xl(self):
        ws = self.wb.create_sheet(title="IP")
        ws.column_dimensions['A'].width = 22.0
        ws.column_dimensions['B'].width = 50.0
        ws.column_dimensions['C'].width = 10.0
        ws.column_dimensions['D'].width = 25.0
        ws.column_dimensions['E'].width = 55.0
        if self.is_ip_task:
            ws.column_dimensions['F'].width = 55.0
            column_tilte = ["IP", "端口信息", "开放端口数目", "geo", "as 编号", "操作系统"]
            ws.append(column_tilte)
            for item in get_ip_data(self.task_id):
                row = []
                row.append(item["ip"])

                port_ids = [str(x["port_id"]) for x in item["port_info"]]
                row.append(" \r\n".join(port_ids))
                row.append(len(item["port_info"]))
                if "country_name" in item["geo_city"]:
                    row.append("{}/{}".format(item["geo_city"]["country_name"],
                                              item["geo_city"]["region_name"]))
                    row.append(item["geo_asn"].get("organization", ""))
                else:
                    row.append("")
                    row.append("")

                osname = ""
                if item.get("os_info"):
                    osname = item["os_info"]["name"]
                row.append(osname)
                ws.append(row)
        else:
            ws.column_dimensions['F'].width = 60.0
            ws.column_dimensions['G'].width = 40.0
            ws.column_dimensions['H'].width = 40.0
            ws.column_dimensions['I'].width = 20.0
            column_tilte = ["IP", "端口信息", "开放端口数目", "geo", "as 编号"]
            column_tilte.append("domain")
            column_tilte.append("操作系统")
            column_tilte.append("CDN")
            column_tilte.append("类别")
            ws.append(column_tilte)
            for item in get_ip_data(self.task_id):
                row = []
                row.append(item["ip"])

                port_ids = [str(x["port_id"]) for x in item["port_info"]]
                row.append(" \r\n".join(port_ids))

                row.append(len(item["port_info"]))
                if "country_name" in item["geo_city"]:
                    row.append("{}/{}".format(item["geo_city"]["country_name"],
                                              item["geo_city"]["region_name"]))
                    row.append(item["geo_asn"].get("organization", ""))
                else:
                    row.append("")
                    row.append("")

                row.append(" \r\n".join(item.get("domain", [])))

                osname = ""
                if item.get("os_info"):
                    osname = item["os_info"]["name"]
                row.append(osname)
                row.append(item.get("cdn_name", ""))
                row.append(item.get("ip_type", ""))
                ws.append(row)

        self.set_style(ws)

    def ignore_illegal(self, content):
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        content = ILLEGAL_CHARACTERS_RE.sub(r'', content)
        return content

    def build_site_xl(self):
        ws = self.wb.active
        ws.column_dimensions['A'].width = 35.0
        ws.column_dimensions['B'].width = 40.0
        ws.column_dimensions['C'].width = 56.0
        ws.column_dimensions['D'].width = 60.0
        ws.column_dimensions['E'].width = 20.0
        ws.column_dimensions['F'].width = 30.0
        ws.column_dimensions['G'].width = 56.0
        ws.column_dimensions['H'].width = 24.0
        ws.title = "站点"
        column_tilte = ["site", "title", "headers", "指纹", "状态码", "favicon hash", "截图", "AI分析"]
        ws.append(column_tilte)
        ai_lookup = _build_ai_denoise_lookup([self.task_id], "site")
        for item in get_site_data(self.task_id):
            item_id = _normalize_ai_lookup_key(item.get("_id", ""))
            ai_result = _resolve_ai_lookup_result(ai_lookup, data_id=item_id, row_key=item_id)
            row = []
            row.append(self.ignore_illegal(item["site"]))
            row.append(self.ignore_illegal(item["title"]))
            row.append(self.ignore_illegal(sanitize_excel_value(item.get("headers", ""))))
            row.append(" \r\n".join([self.ignore_illegal(x["name"]) for x in item["finger"]]))
            row.append(item["status"])
            row.append(item["favicon"].get("hash", ""))
            row.append(self.ignore_illegal(sanitize_excel_value(item.get("screenshot", ""))))
            row.append(sanitize_excel_value(ai_result.get("text", "未分析")))
            ws.append(row)

        self.set_style(ws)

    def build_domain_xl(self):
        ws = self.wb.create_sheet(title="域名")
        ws.column_dimensions['A'].width = 30.0
        ws.column_dimensions['B'].width = 20.0
        ws.column_dimensions['C'].width = 50.0
        ws.column_dimensions['D'].width = 50.0
        ws.column_dimensions['E'].width = 24.0

        column_tilte = ["域名", "解析类型", "记录值", "关联ip", "来源"]

        ws.append(column_tilte)
        for item in get_domain_data(self.task_id):
            row = []
            row.append(item["domain"])
            row.append(item["type"])
            row.append(" \r\n".join(item["record"]))
            row.append(" \r\n".join(item["ips"]))
            row.append(_format_domain_source_text(item.get("source", "")))
            ws.append(row)

        self.set_style(ws)

    def build_url_xl(self):
        """
        构建 URL 信息工作表。
        """
        _build_url_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_fileleak_xl(self):
        """
        构建目录扫描工作表。
        """
        _build_fileleak_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_wih_xl(self):
        """
        构建 WIH 工作表。
        """
        _build_wih_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_wih_endpoint_xl(self):
        """
        构建 WIH 接口提取工作表。
        """
        _build_wih_endpoint_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_waf_xl(self):
        """
        构建 WAF 识别工作表。
        """
        _build_waf_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_nuclei_xl(self):
        """
        构建 PoC 风险工作表。
        """
        _build_nuclei_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_ai_pen_xl(self):
        """
        构建 AI 渗透测试工作表。
        """
        _build_ai_pen_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_stat_finger_xl(self):
        """
        构建指纹统计工作表。
        """
        _build_stat_finger_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_cert_xl(self):
        """
        生成 SSL 证书工作表（协议/套件/强度）。
        """
        _build_cert_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_statist(self):
        statist = port_service_product_statist(self.task_id)
        ws = self.wb.create_sheet(title="资产统计")
        ws.column_dimensions['A'].width = 20.0
        ws.column_dimensions['F'].width = 20.0
        ws.column_dimensions['K'].width = 40.0
        ws["A1"] = "端口信息统计"
        ws["F1"] = "系统服务信息统计"
        ws["K1"] = "软件产品信息统计"

        ports = ["端口", "数量", "占比"]
        port_percent_list = statist["port_percent_list"]
        port_total = statist["port_total"]
        for port_info in port_percent_list:
            ports.append(port_info["port_id"])
            ports.append(port_info["amount"])
            ports.append(port_info["percent"])

        cnt = 0
        for row in range(5, 27):
            for col in range(1, 4):
                if cnt >= len(ports):
                    continue
                ws.cell(column=col, row=row, value=ports[cnt])
                cnt += 1

        ws["A27"] = "端口开放总数"
        ws["A28"] = port_total

        services = ["系统服务", "数量", "占比"]
        service_percent_list = statist["service_percent_list"]
        if len(service_percent_list) >= 0:
            service_total = statist["service_total"]
            for port_info in service_percent_list:
                services.append(port_info["service_name"])
                services.append(port_info["amount"])
                services.append(port_info["percent"])
            cnt = 0
            for row in range(5, 27):
                for col in range(6, 9):
                    if cnt >= len(services):
                        continue
                    ws.cell(column=col, row=row, value=services[cnt])
                    cnt += 1
            ws["F27"] = "系统服务类别总数"
            ws["F28"] = service_total

        product = ["产品", "数量", "占比"]
        product_percent_list = statist["product_percent_list"]
        if len(product_percent_list) >= 0:
            product_total = statist["product_total"]
            for port_info in product_percent_list:
                product.append(port_info["product"])
                product.append(port_info["amount"])
                product.append(port_info["percent"])
            cnt = 0
            for row in range(5, 27):
                for col in range(11, 14):
                    if cnt >= len(product):
                        continue
                    ws.cell(column=col, row=row, value=product[cnt])
                    cnt += 1
            ws["K27"] = "产品类别总数"
            ws["K28"] = product_total

        self.set_style(ws)

    def build_vuln_xl(self):
        _build_vuln_sheet(self.wb, [self.task_id], apply_style=self.apply_style)

    def build_workbook(self):
        task_data = get_task_data(self.task_id)
        if not task_data:
            print("not found {}".format(self.task_id))
            return None

        domain = task_data["target"].replace("/", "_")[:20]

        if re.findall(r"\b\d+\.\d+\.\d+\.\d+", domain):
            self.is_ip_task = True
        else:
            if task_data.get("type", "") == "ip":
                self.is_ip_task = True

        self.build_site_xl()
        self.build_ip_xl()
        self.build_service_xl()
        self.build_cert_xl()
        self.build_domain_xl()
        self.build_url_xl()
        self.build_fileleak_xl()
        self.build_wih_xl()
        self.build_wih_endpoint_xl()
        self.build_waf_xl()
        self.build_vuln_xl()
        self.build_nuclei_xl()
        self.build_ai_pen_xl()
        self.build_stat_finger_xl()
        self.build_statist()

        return self.wb

    def run(self):
        workbook = self.build_workbook()
        if not workbook:
            return None

        return save_virtual_workbook(workbook)


def build_single_task_workbook(task_id, apply_style=True):
    """
    构建单任务导出工作簿，供 Excel/HTML 两种格式复用。
    """
    task_id = task_id.strip()
    save = SaveTask(task_id, apply_style=apply_style)
    return save.build_workbook()


def export_arl(task_id):
    workbook = build_single_task_workbook(task_id, apply_style=True)
    if not workbook:
        return None
    return save_virtual_workbook(workbook)


def export_arl_html(task_id):
    """
    导出单任务 HTML 报告。
    """
    workbook = build_single_task_workbook(task_id, apply_style=False)
    if not workbook:
        return None
    task_data = get_task_data(task_id.strip()) or {}
    target = sanitize_excel_value(task_data.get("target", "")).strip() or sanitize_excel_value(task_id)
    title = "ARL资产导出报告 - {}".format(target)
    metadata = build_html_report_metadata([task_data])
    return render_workbook_html(workbook, title, metadata=metadata)


def export_arl_ai_markdown(task_id):
    """
    导出单任务 AI 报告（Markdown 模板）。
    """
    ai_settings = _get_ai_export_settings()
    return _build_ai_markdown_report([task_id], ai_settings)


def build_merge_tasks_workbook(task_id_list, apply_style=True):
    """
    整合多个任务并构建统一工作簿，供 Excel/HTML 两种格式复用。
    
    参数：
        task_id_list: 任务ID列表
    
    返回：
        Workbook 对象
    
    说明：
    - 合并多个任务的所有扫描数据
    - 按照单个任务的导出格式生成报告
    - 保留任务原始IP/服务明细（不做跨任务折叠），保证与页面口径一致
    - 域名、站点仍按值合并去重，避免重复噪音
    """
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    valid_tasks = []
    valid_task_ids = []
    for task_id in task_id_list:
        if not task_id:
            continue
        task_data = get_task_data(task_id)
        if task_data:
            valid_tasks.append(task_data)
            valid_task_ids.append(str(task_data.get("_id", "")))

    if not valid_tasks:
        raise ValueError("未找到可导出的任务数据")

    # 与单任务保持一致：仅当全部任务都是 IP 类型时，按 IP 任务列导出；否则按通用任务列导出
    is_ip_task = True
    for task_data in valid_tasks:
        target = sanitize_excel_value(task_data.get("target", ""))
        if not (re.findall(r"\b\d+\.\d+\.\d+\.\d+", target) or task_data.get("type", "") == "ip"):
            is_ip_task = False
            break

    merged_ip_items = []  # 保留原始 ip 文档（不跨任务合并）
    merged_domains = {}   # key: domain
    merged_sites = {}     # key: site
    site_ai_lookup = _build_ai_denoise_lookup(valid_task_ids, "site")

    for task_data in valid_tasks:
        task_id = str(task_data.get("_id"))

        for ip_item in get_ip_data(task_id):
            ip = sanitize_excel_value(ip_item.get("ip", "")).strip()
            if not ip:
                continue

            port_info_list = []
            for port_info in as_list(ip_item.get("port_info", [])):
                if isinstance(port_info, dict):
                    port_info_list.append(port_info)

            domain_list = []
            domain_seen = set()
            for domain in as_list(ip_item.get("domain", [])):
                domain_text = sanitize_excel_value(domain).strip()
                if not domain_text or domain_text in domain_seen:
                    continue
                domain_seen.add(domain_text)
                domain_list.append(domain_text)

            merged_ip_items.append({
                "task_id": task_id,
                "ip": ip,
                "port_info": port_info_list,
                "geo_city": ip_item.get("geo_city", {}) if isinstance(ip_item.get("geo_city", {}), dict) else {},
                "geo_asn": ip_item.get("geo_asn", {}) if isinstance(ip_item.get("geo_asn", {}), dict) else {},
                "domain": domain_list,
                "os_info": ip_item.get("os_info", {}) if isinstance(ip_item.get("os_info", {}), dict) else {},
                "cdn_name": ip_item.get("cdn_name", ""),
                "ip_type": ip_item.get("ip_type", ""),
            })

        for domain_item in get_domain_data(task_id):
            domain = domain_item.get("domain")
            if not domain:
                continue
            if domain not in merged_domains:
                merged_domains[domain] = {
                    "domain": domain,
                    "type": domain_item.get("type", ""),
                    "record": as_list(domain_item.get("record", [])),
                    "ips": as_list(domain_item.get("ips", [])),
                    "sources": _extract_domain_source_list(domain_item.get("source", "")),
                }
            else:
                merged = merged_domains[domain]
                if not merged.get("type") and domain_item.get("type"):
                    merged["type"] = domain_item.get("type")
                merged["record"] = sorted(list(set(merged.get("record", []) + as_list(domain_item.get("record", [])))))
                merged["ips"] = sorted(list(set(merged.get("ips", []) + as_list(domain_item.get("ips", [])))))
                merged["sources"] = sorted(
                    list(
                        set(
                            as_list(merged.get("sources", []))
                            + _extract_domain_source_list(domain_item.get("source", ""))
                        )
                    )
                )

        for site_item in get_site_data(task_id):
            site = site_item.get("site") or site_item.get("url")
            if not site:
                continue
            site_item_id = _normalize_ai_lookup_key(site_item.get("_id", ""))
            site_ai_result = _resolve_ai_lookup_result(site_ai_lookup, data_id=site_item_id, row_key=site_item_id)
            if site not in merged_sites:
                merged_sites[site] = {
                    "site": site,
                    "title": site_item.get("title", ""),
                    "headers": site_item.get("headers", ""),
                    "finger": as_list(site_item.get("finger", [])),
                    "screenshot": site_item.get("screenshot", ""),
                    "status": site_item.get("status", ""),
                    "favicon": site_item.get("favicon", {}),
                    "ai_result": site_ai_result,
                }
            else:
                merged = merged_sites[site]
                if not merged.get("title") and site_item.get("title"):
                    merged["title"] = site_item.get("title", "")
                if not merged.get("headers") and site_item.get("headers"):
                    merged["headers"] = site_item.get("headers", "")
                if not merged.get("screenshot") and site_item.get("screenshot"):
                    merged["screenshot"] = site_item.get("screenshot", "")
                if not merged.get("status") and site_item.get("status"):
                    merged["status"] = site_item.get("status", "")
                if (not isinstance(merged.get("favicon"), dict) or not merged.get("favicon", {}).get("hash")) and \
                        isinstance(site_item.get("favicon"), dict):
                    merged["favicon"] = site_item.get("favicon", {})
                current_ai_result = merged.get("ai_result")
                if _is_ai_lookup_result_better(site_ai_result, current_ai_result):
                    merged["ai_result"] = site_ai_result

                # 按指纹名称去重
                name_set = set()
                new_fingers = []
                for finger in as_list(merged.get("finger", [])) + as_list(site_item.get("finger", [])):
                    if isinstance(finger, dict):
                        name = sanitize_excel_value(finger.get("name", ""))
                        key = ("dict", name)
                    else:
                        name = sanitize_excel_value(finger)
                        key = ("str", name)
                    if key in name_set:
                        continue
                    name_set.add(key)
                    new_fingers.append(finger)
                merged["finger"] = new_fingers

    if not merged_ip_items and not merged_domains and not merged_sites:
        raise ValueError("未找到可导出的任务数据")

    # 站点（与单任务导出同结构）
    ws = wb.create_sheet(title="站点")
    ws.column_dimensions['A'].width = 35.0
    ws.column_dimensions['B'].width = 40.0
    ws.column_dimensions['C'].width = 56.0
    ws.column_dimensions['D'].width = 60.0
    ws.column_dimensions['E'].width = 20.0
    ws.column_dimensions['F'].width = 30.0
    ws.column_dimensions['G'].width = 56.0
    ws.column_dimensions['H'].width = 24.0
    ws.append(["site", "title", "headers", "指纹", "状态码", "favicon hash", "截图", "AI分析"])
    for site in sorted(merged_sites.keys()):
        item = merged_sites[site]
        ws.append([
            sanitize_excel_value(item.get("site", "")),
            sanitize_excel_value(item.get("title", "")),
            sanitize_excel_value(item.get("headers", "")),
            sanitize_excel_value(extract_finger_names(item.get("finger", []))).replace(",", " \r\n"),
            sanitize_excel_value(item.get("status", "")),
            sanitize_excel_value((item.get("favicon", {}) or {}).get("hash", "")),
            sanitize_excel_value(item.get("screenshot", "")),
            sanitize_excel_value((item.get("ai_result") or {}).get("text", "未分析")),
        ])
    if apply_style:
        set_sheet_style(ws)

    # IP（与单任务导出同结构）
    ws = wb.create_sheet(title="IP")
    ws.column_dimensions['A'].width = 22.0
    ws.column_dimensions['B'].width = 50.0
    ws.column_dimensions['C'].width = 10.0
    ws.column_dimensions['D'].width = 25.0
    ws.column_dimensions['E'].width = 55.0

    if is_ip_task:
        ws.column_dimensions['F'].width = 55.0
        ws.append(["IP", "端口信息", "开放端口数目", "geo", "as 编号", "操作系统"])
        for item in merged_ip_items:
            port_ids = [str(x.get("port_id")) for x in item.get("port_info", []) if x.get("port_id") is not None]
            geo_city = item.get("geo_city", {}) if isinstance(item.get("geo_city", {}), dict) else {}
            geo_asn = item.get("geo_asn", {}) if isinstance(item.get("geo_asn", {}), dict) else {}
            geo_text = ""
            as_text = ""
            if "country_name" in geo_city:
                geo_text = "{}/{}".format(geo_city.get("country_name", ""), geo_city.get("region_name", ""))
                as_text = geo_asn.get("organization", "")
            osname = ""
            if isinstance(item.get("os_info", {}), dict):
                osname = item.get("os_info", {}).get("name", "")
            ws.append([
                sanitize_excel_value(item.get("ip", "")),
                sanitize_excel_value(" \r\n".join(port_ids)),
                len(item.get("port_info", [])),
                sanitize_excel_value(geo_text),
                sanitize_excel_value(as_text),
                sanitize_excel_value(osname),
            ])
    else:
        ws.column_dimensions['F'].width = 60.0
        ws.column_dimensions['G'].width = 40.0
        ws.column_dimensions['H'].width = 40.0
        ws.column_dimensions['I'].width = 20.0
        ws.append(["IP", "端口信息", "开放端口数目", "geo", "as 编号", "domain", "操作系统", "CDN", "类别"])
        for item in merged_ip_items:
            port_ids = [str(x.get("port_id")) for x in item.get("port_info", []) if x.get("port_id") is not None]
            geo_city = item.get("geo_city", {}) if isinstance(item.get("geo_city", {}), dict) else {}
            geo_asn = item.get("geo_asn", {}) if isinstance(item.get("geo_asn", {}), dict) else {}
            geo_text = ""
            as_text = ""
            if "country_name" in geo_city:
                geo_text = "{}/{}".format(geo_city.get("country_name", ""), geo_city.get("region_name", ""))
                as_text = geo_asn.get("organization", "")
            osname = ""
            if isinstance(item.get("os_info", {}), dict):
                osname = item.get("os_info", {}).get("name", "")
            ws.append([
                sanitize_excel_value(item.get("ip", "")),
                sanitize_excel_value(" \r\n".join(port_ids)),
                len(item.get("port_info", [])),
                sanitize_excel_value(geo_text),
                sanitize_excel_value(as_text),
                sanitize_excel_value(" \r\n".join(as_list(item.get("domain", [])))),
                sanitize_excel_value(osname),
                sanitize_excel_value(item.get("cdn_name", "")),
                sanitize_excel_value(item.get("ip_type", "")),
            ])
    if apply_style:
        set_sheet_style(ws)

    # 系统服务（与单任务导出同结构）
    ws = wb.create_sheet(title="系统服务")
    ws.column_dimensions['A'].width = 22.0
    ws.column_dimensions['B'].width = 10.0
    ws.column_dimensions['C'].width = 20.0
    ws.column_dimensions['D'].width = 40.0
    ws.append(["IP", "端口", "服务", "产品", "版本"])
    for row in _build_service_rows(valid_task_ids, fallback_ip_items=merged_ip_items):
        ws.append([
            sanitize_excel_value(row[0]),
            sanitize_excel_value(row[1]),
            sanitize_excel_value(row[2]),
            sanitize_excel_value(row[3]),
            sanitize_excel_value(row[4]),
        ])
    if apply_style:
        set_sheet_style(ws)

    _build_cert_sheet(wb, valid_task_ids, apply_style=apply_style)

    # 域名（统一保留，IP任务为空时仅输出表头）
    ws = wb.create_sheet(title="域名")
    ws.column_dimensions['A'].width = 30.0
    ws.column_dimensions['B'].width = 20.0
    ws.column_dimensions['C'].width = 50.0
    ws.column_dimensions['D'].width = 50.0
    ws.column_dimensions['E'].width = 24.0
    ws.append(["域名", "解析类型", "记录值", "关联ip", "来源"])
    for domain in sorted(merged_domains.keys()):
        item = merged_domains[domain]
        ws.append([
            sanitize_excel_value(item.get("domain", "")),
            sanitize_excel_value(item.get("type", "")),
            sanitize_excel_value(" \r\n".join(as_list(item.get("record", [])))),
            sanitize_excel_value(" \r\n".join(as_list(item.get("ips", [])))),
            sanitize_excel_value(_format_domain_source_text(item.get("sources", []))),
        ])
    if apply_style:
        set_sheet_style(ws)

    # URL信息 / 目录扫描 / WIH / WIH接口提取 / 风险（与单任务导出顺序保持一致）
    _build_url_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_fileleak_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_wih_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_wih_endpoint_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_waf_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_vuln_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_nuclei_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_ai_pen_sheet(wb, valid_task_ids, apply_style=apply_style)
    _build_stat_finger_sheet(wb, valid_task_ids, apply_style=apply_style)

    # 资产统计（与单任务导出同结构）
    statist = calc_port_service_product_statist_from_ip_items(merged_ip_items)
    ws = wb.create_sheet(title="资产统计")
    ws.column_dimensions['A'].width = 20.0
    ws.column_dimensions['F'].width = 20.0
    ws.column_dimensions['K'].width = 40.0
    ws["A1"] = "端口信息统计"
    ws["F1"] = "系统服务信息统计"
    ws["K1"] = "软件产品信息统计"

    ports = ["端口", "数量", "占比"]
    for port_info in statist["port_percent_list"]:
        ports.extend([port_info["port_id"], port_info["amount"], port_info["percent"]])
    cnt = 0
    for row in range(5, 27):
        for col in range(1, 4):
            if cnt >= len(ports):
                continue
            ws.cell(column=col, row=row, value=ports[cnt])
            cnt += 1
    ws["A27"] = "端口开放总数"
    ws["A28"] = statist["port_total"]

    services = ["系统服务", "数量", "占比"]
    for service_info in statist["service_percent_list"]:
        services.extend([service_info["service_name"], service_info["amount"], service_info["percent"]])
    cnt = 0
    for row in range(5, 27):
        for col in range(6, 9):
            if cnt >= len(services):
                continue
            ws.cell(column=col, row=row, value=services[cnt])
            cnt += 1
    ws["F27"] = "系统服务类别总数"
    ws["F28"] = statist["service_total"]

    product = ["产品", "数量", "占比"]
    for product_info in statist["product_percent_list"]:
        product.extend([product_info["product"], product_info["amount"], product_info["percent"]])
    cnt = 0
    for row in range(5, 27):
        for col in range(11, 14):
            if cnt >= len(product):
                continue
            ws.cell(column=col, row=row, value=product[cnt])
            cnt += 1
    ws["K27"] = "产品类别总数"
    ws["K28"] = statist["product_total"]
    if apply_style:
        set_sheet_style(ws)

    return wb


def export_merge_tasks(task_id_list):
    workbook = build_merge_tasks_workbook(task_id_list, apply_style=True)
    return save_virtual_workbook(workbook)


def export_merge_tasks_html(task_id_list):
    """
    导出批量任务 HTML 报告。
    """
    workbook = build_merge_tasks_workbook(task_id_list, apply_style=False)
    task_items = []
    task_names = []
    for task_id in _normalize_task_id_list(task_id_list):
        task_data = get_task_data(task_id)
        if not task_data:
            continue
        task_items.append(task_data)
        task_name = sanitize_excel_value(task_data.get("name", "")).strip()
        if task_name and task_name not in task_names:
            task_names.append(task_name)

    title = "ARL批量导出报告"
    if task_names:
        title = "{} - {}".format(title, " / ".join(task_names[:3]))
        if len(task_names) > 3:
            title = "{} 等{}个任务".format(title, len(task_names))

    metadata = build_html_report_metadata(task_items)
    return render_workbook_html(workbook, title, metadata=metadata)


def export_merge_tasks_ai_markdown(task_id_list):
    """
    导出批量任务 AI 报告（Markdown 模板）。
    """
    ai_settings = _get_ai_export_settings()
    return _build_ai_markdown_report(task_id_list, ai_settings)
