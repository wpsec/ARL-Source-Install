"""
钉钉开放平台（知识库）接口工具

说明：
- 不依赖钉钉 SDK，直接调用 HTTP API
- 支持 access_token 获取与缓存
- 支持创建钉钉知识库表格（WORKBOOK）
"""
import time
import urllib.parse
import threading
import os
import re
from pathlib import Path
from io import BytesIO
import yaml
from app import utils
from app.config import Config
try:
    from bson import ObjectId
except Exception:
    ObjectId = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


logger = utils.get_logger()

_TOKEN_CACHE = {
    "access_token": "",
    "expires_at": 0,
    "signature": "",
}
_TOKEN_CACHE_LOCK = threading.Lock()
_RUNTIME_CONFIG_LOCK = threading.Lock()
_RUNTIME_CONFIG_STATE = {
    "path": "",
    "mtime_ns": -1,
}
_DEFAULT_CREATE_DOC_PATH = "/v1.0/doc/workspaces/{workspace_id}/docs"
_LEGACY_MARKDOWN_PATH = "/v2.0/wiki/nodes"


def _safe_bool(value, default_value=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("1", "true", "yes", "y", "on")
    return bool(default_value)


def _safe_int(value, default_value, min_value=1):
    try:
        parsed = int(value)
    except Exception:
        return int(default_value)
    if parsed < min_value:
        return int(default_value)
    return parsed


def _resolve_config_path():
    """
    解析配置文件路径，优先容器挂载路径，兼容本地源码运行。
    """
    custom_path = str(os.environ.get("ARL_CONFIG_EDIT_PATH", "") or "").strip()
    candidates = [
        Path(custom_path) if custom_path else None,
        Path("/code/app/config.yaml"),
        Path(__file__).resolve().parents[2] / "docker" / "config-docker.yaml",
    ]
    for item in candidates:
        if not item:
            continue
        if item.exists() and item.is_file():
            return item
    return Path(__file__).resolve().parents[2] / "docker" / "config-docker.yaml"


def _load_config_from_file(config_path):
    if not config_path.exists():
        return {}
    with config_path.open("r", encoding="utf-8") as file_obj:
        loaded = yaml.safe_load(file_obj) or {}
    if not isinstance(loaded, dict):
        raise ValueError("配置文件根节点必须为对象")
    return loaded


def _extract_dingtalk_runtime_config(config_obj):
    dingding_conf = config_obj.get("DINGDING", {})
    if not isinstance(dingding_conf, dict):
        dingding_conf = {}

    dingtalk_api_conf = config_obj.get("DINGTALK_API", {})
    if not isinstance(dingtalk_api_conf, dict):
        dingtalk_api_conf = {}

    return {
        "dingding_access_token": str(dingding_conf.get("ACCESS_TOKEN", Config.DINGDING_ACCESS_TOKEN or "")),
        "dingding_secret": str(dingding_conf.get("SECRET", Config.DINGDING_SECRET or "")),
        "kb_enable": _safe_bool(dingtalk_api_conf.get("ENABLE"), Config.DINGTALK_KB_ENABLE),
        "base_url": str(dingtalk_api_conf.get("BASE_URL", Config.DINGTALK_API_BASE_URL or "https://api.dingtalk.com")),
        "corp_id": str(dingtalk_api_conf.get("CORP_ID", Config.DINGTALK_CORP_ID or "")),
        "app_key": str(dingtalk_api_conf.get("APP_KEY", Config.DINGTALK_APP_KEY or "")),
        "app_secret": str(dingtalk_api_conf.get("APP_SECRET", Config.DINGTALK_APP_SECRET or "")),
        "operator_id": str(dingtalk_api_conf.get("OPERATOR_ID", Config.DINGTALK_OPERATOR_ID or "")),
        "workspace_id": str(dingtalk_api_conf.get("WORKSPACE_ID", Config.DINGTALK_WORKSPACE_ID or "")),
        "parent_node_id": str(dingtalk_api_conf.get("PARENT_NODE_ID", Config.DINGTALK_PARENT_NODE_ID or "")),
        "create_node_path": str(
            dingtalk_api_conf.get("CREATE_NODE_PATH", Config.DINGTALK_KB_CREATE_NODE_PATH or "")
        ),
        "kb_timeout": _safe_int(dingtalk_api_conf.get("KB_TIMEOUT"), Config.DINGTALK_KB_TIMEOUT),
        "title_prefix": str(dingtalk_api_conf.get("TITLE_PREFIX", Config.DINGTALK_KB_TITLE_PREFIX or "")),
        "dry_run": _safe_bool(dingtalk_api_conf.get("DRY_RUN"), Config.DINGTALK_KB_DRY_RUN),
        "report_base_url": str(dingtalk_api_conf.get("REPORT_BASE_URL", Config.DINGTALK_REPORT_BASE_URL or "")),
        "ssl_cert_notify_enable": _safe_bool(
            dingtalk_api_conf.get("SSL_CERT_NOTIFY_ENABLE"), Config.DINGTALK_SSL_CERT_NOTIFY_ENABLE
        ),
        "ssl_cert_notify_days": _safe_int(
            dingtalk_api_conf.get("SSL_CERT_NOTIFY_DAYS"), Config.DINGTALK_SSL_CERT_NOTIFY_DAYS, min_value=1
        ),
    }


def _apply_runtime_dingtalk_config(dingtalk_config):
    Config.DINGDING_ACCESS_TOKEN = str(dingtalk_config.get("dingding_access_token", "")).strip()
    Config.DINGDING_SECRET = str(dingtalk_config.get("dingding_secret", "")).strip()
    Config.DINGTALK_KB_ENABLE = _safe_bool(dingtalk_config.get("kb_enable"), False)
    Config.DINGTALK_API_BASE_URL = str(dingtalk_config.get("base_url", "")).strip() or "https://api.dingtalk.com"
    Config.DINGTALK_CORP_ID = str(dingtalk_config.get("corp_id", "")).strip()
    Config.DINGTALK_APP_KEY = str(dingtalk_config.get("app_key", "")).strip()
    Config.DINGTALK_APP_SECRET = str(dingtalk_config.get("app_secret", "")).strip()
    Config.DINGTALK_OPERATOR_ID = str(dingtalk_config.get("operator_id", "")).strip()
    Config.DINGTALK_WORKSPACE_ID = str(dingtalk_config.get("workspace_id", "")).strip()
    Config.DINGTALK_PARENT_NODE_ID = str(dingtalk_config.get("parent_node_id", "")).strip()
    Config.DINGTALK_KB_CREATE_NODE_PATH = (
        str(dingtalk_config.get("create_node_path", "")).strip() or "/v1.0/doc/workspaces/{workspace_id}/docs"
    )
    Config.DINGTALK_KB_TIMEOUT = _safe_int(dingtalk_config.get("kb_timeout"), 20)
    Config.DINGTALK_KB_TITLE_PREFIX = str(dingtalk_config.get("title_prefix", "")).strip()
    Config.DINGTALK_KB_DRY_RUN = _safe_bool(dingtalk_config.get("dry_run"), False)
    Config.DINGTALK_REPORT_BASE_URL = str(dingtalk_config.get("report_base_url", "")).strip()
    Config.DINGTALK_SSL_CERT_NOTIFY_ENABLE = _safe_bool(
        dingtalk_config.get("ssl_cert_notify_enable"), False
    )
    Config.DINGTALK_SSL_CERT_NOTIFY_DAYS = _safe_int(
        dingtalk_config.get("ssl_cert_notify_days"), 30, min_value=1
    )


def refresh_runtime_dingtalk_config_best_effort(force=False):
    """
    最佳努力从配置文件同步钉钉相关配置到当前进程内存。

    说明：
    - 用于 web/worker/scheduler 多进程场景下，避免配置更新后必须重启进程才生效。
    - 当开放平台核心配置变化时自动清理 access_token 缓存。
    """
    config_path = _resolve_config_path()
    path_text = str(config_path)
    mtime_ns = -1
    try:
        mtime_ns = int(config_path.stat().st_mtime_ns)
    except Exception:
        pass

    old_signature = _runtime_signature()
    try:
        with _RUNTIME_CONFIG_LOCK:
            if (
                not force
                and _RUNTIME_CONFIG_STATE.get("path") == path_text
                and int(_RUNTIME_CONFIG_STATE.get("mtime_ns", -1)) == mtime_ns
            ):
                return False

            config_obj = _load_config_from_file(config_path)
            dingtalk_config = _extract_dingtalk_runtime_config(config_obj)
            _apply_runtime_dingtalk_config(dingtalk_config)
            _RUNTIME_CONFIG_STATE["path"] = path_text
            _RUNTIME_CONFIG_STATE["mtime_ns"] = mtime_ns
    except Exception as exc:
        logger.warning("refresh dingtalk runtime config failed: %s", exc)
        return False

    new_signature = _runtime_signature()
    if old_signature != new_signature:
        reset_access_token_cache()
    return True


def _is_config_ready(require_enable=True, require_workspace=True, require_parent_node=True):
    """
    检查钉钉开放平台配置是否完整
    """
    if require_enable and not Config.DINGTALK_KB_ENABLE:
        return False

    required_values = [
        Config.DINGTALK_CORP_ID,
        Config.DINGTALK_APP_KEY,
        Config.DINGTALK_APP_SECRET,
        Config.DINGTALK_OPERATOR_ID,
    ]
    if require_workspace:
        required_values.append(Config.DINGTALK_WORKSPACE_ID)
    if require_parent_node:
        required_values.append(Config.DINGTALK_PARENT_NODE_ID)

    return all(required_values)


def _missing_required_fields(require_workspace=False, require_parent_node=False):
    """
    返回缺失的配置字段名列表
    """
    required_fields = {
        "CORP_ID": Config.DINGTALK_CORP_ID,
        "APP_KEY": Config.DINGTALK_APP_KEY,
        "APP_SECRET": Config.DINGTALK_APP_SECRET,
        "OPERATOR_ID": Config.DINGTALK_OPERATOR_ID,
    }
    if require_workspace:
        required_fields["WORKSPACE_ID"] = Config.DINGTALK_WORKSPACE_ID
    if require_parent_node:
        required_fields["PARENT_NODE_ID"] = Config.DINGTALK_PARENT_NODE_ID

    missing_fields = []
    for key, value in required_fields.items():
        if not str(value or "").strip():
            missing_fields.append(key)

    return missing_fields


def _parse_response(conn):
    """
    统一解析 HTTP 响应
    """
    try:
        data = conn.json()
    except Exception:
        data = {"raw": conn.text}
    return data


def _is_success(data, status_code):
    """
    判断钉钉接口调用是否成功
    """
    if status_code >= 400:
        return False

    if not isinstance(data, dict):
        return True

    err_code = data.get("errcode")
    if err_code not in [None, 0, "0"]:
        return False

    # 兼容部分接口返回 code/message
    code_value = data.get("code")
    if code_value is None:
        return True

    code_text = str(code_value).lower()
    if code_text in ["0", "ok", "success", "200"]:
        return True

    return False


def _build_url(path):
    base_url = (Config.DINGTALK_API_BASE_URL or "").rstrip("/")
    if not path.startswith("/"):
        path = "/" + path
    return base_url + path


def _runtime_signature():
    """
    钉钉开放平台配置签名，用于识别配置变更后触发 token 缓存失效。
    """
    return "|".join(
        [
            str(Config.DINGTALK_API_BASE_URL or "").strip(),
            str(Config.DINGTALK_CORP_ID or "").strip(),
            str(Config.DINGTALK_APP_KEY or "").strip(),
            str(Config.DINGTALK_APP_SECRET or "").strip(),
        ]
    )


def reset_access_token_cache():
    """
    清理 access_token 缓存（配置变更或鉴权失败后调用）。
    """
    with _TOKEN_CACHE_LOCK:
        _TOKEN_CACHE["access_token"] = ""
        _TOKEN_CACHE["expires_at"] = 0
        _TOKEN_CACHE["signature"] = _runtime_signature()


def _extract_response_error_text(data):
    if not isinstance(data, dict):
        return str(data or "")

    for key in ["error", "message", "msg", "errMsg", "errmsg", "error_message"]:
        value = data.get(key)
        if value:
            return str(value)

    nested = data.get("data")
    if isinstance(nested, dict):
        for key in ["error", "message", "msg", "errMsg", "errmsg", "error_message"]:
            value = nested.get(key)
            if value:
                return str(value)

    return str(data)


def _is_auth_failure(status_code, data):
    if int(status_code or 0) in [401, 403]:
        return True

    err_code = ""
    if isinstance(data, dict):
        err_code = str(data.get("errcode") or data.get("code") or "").strip().lower()
    err_text = _extract_response_error_text(data).lower()

    auth_code_hits = {
        "invalidauthentication",
        "invalid_authentication",
        "invalidtoken",
        "invalid_token",
        "invalidaccesstoken",
        "invalid_access_token",
        "accesstokeninvalid",
        "access_token_invalid",
        "accesstokenexpired",
        "access_token_expired",
    }
    if err_code and err_code in auth_code_hits:
        return True

    auth_keywords = [
        "access token",
        "access_token",
        "token",
        "unauthorized",
        "invalid authentication",
        "invalid token",
        "token expired",
        "signature not match",
    ]
    return any(keyword in err_text for keyword in auth_keywords)


def _is_transient_failure(status_code, data, error_text=""):
    if int(status_code or 0) in [408, 409, 425, 429, 500, 502, 503, 504]:
        return True

    probe_text = " ".join(
        [
            _extract_response_error_text(data).lower(),
            str(error_text or "").lower(),
        ]
    )
    transient_keywords = [
        "timeout",
        "timed out",
        "temporarily unavailable",
        "too many requests",
        "rate limit",
        "connection reset",
        "connection aborted",
        "gateway",
        "service unavailable",
    ]
    return any(keyword in probe_text for keyword in transient_keywords)


def get_access_token(force_refresh=False, require_enable=True):
    """
    获取 access_token（含简单内存缓存）
    """
    # 每次请求前按 mtime 轻量检查配置文件，保证常驻进程可感知配置变更。
    refresh_runtime_dingtalk_config_best_effort()

    if not _is_config_ready(require_enable=require_enable, require_workspace=False, require_parent_node=False):
        return ""

    now_ts = int(time.time())
    current_signature = _runtime_signature()
    with _TOKEN_CACHE_LOCK:
        if _TOKEN_CACHE.get("signature") != current_signature:
            _TOKEN_CACHE["access_token"] = ""
            _TOKEN_CACHE["expires_at"] = 0
            _TOKEN_CACHE["signature"] = current_signature

        if (
            not force_refresh
            and _TOKEN_CACHE.get("access_token")
            and _TOKEN_CACHE.get("expires_at", 0) > now_ts + 60
        ):
            return _TOKEN_CACHE["access_token"]

    token_url = _build_url("/v1.0/oauth2/{}/token".format(Config.DINGTALK_CORP_ID))
    payload = {
        "client_id": Config.DINGTALK_APP_KEY,
        "client_secret": Config.DINGTALK_APP_SECRET,
        "grant_type": "client_credentials",
    }
    headers = {
        "Content-Type": "application/json",
    }

    try:
        conn = utils.http_req(token_url, method="post", json=payload, headers=headers, timeout=(8, Config.DINGTALK_KB_TIMEOUT))
        data = _parse_response(conn)
        if conn.status_code >= 400:
            logger.warning("dingtalk token status:{} body:{}".format(conn.status_code, data))
            return ""

        token = data.get("access_token", "")
        expires_in = int(data.get("expires_in", 7200))
        if not token:
            logger.warning("dingtalk token missing, body:{}".format(data))
            return ""

        with _TOKEN_CACHE_LOCK:
            _TOKEN_CACHE["access_token"] = token
            _TOKEN_CACHE["expires_at"] = now_ts + max(expires_in - 120, 60)
            _TOKEN_CACHE["signature"] = current_signature
        return token
    except Exception as e:
        logger.warning("get dingtalk token error {}".format(e))
        return ""


def request_openapi(
    method,
    path,
    params=None,
    json_data=None,
    require_enable=True,
    force_refresh_token=False,
    retry_on_transient=False,
    retry_max=3,
):
    """
    调用钉钉开放平台接口
    """
    method_text = str(method or "get").strip().lower()
    max_attempts = max(1, int(retry_max or 1))
    if not retry_on_transient:
        # 非瞬时重试模式下仍保留一次“鉴权失败后刷新 token”重试机会。
        max_attempts = max(2, min(max_attempts, 2))

    force_refresh = bool(force_refresh_token)
    auth_retry_used = False
    last_result = {"error": "request not executed"}
    url = _build_url(path)

    for attempt in range(1, max_attempts + 1):
        token = get_access_token(force_refresh=force_refresh, require_enable=require_enable)
        force_refresh = False
        if not token:
            last_result = {
                "error": "access_token is empty",
                "missing_fields": _missing_required_fields(),
                "attempt": attempt,
            }
            if attempt < max_attempts:
                time.sleep(min(0.45 * attempt, 1.5))
                continue
            return False, last_result

        headers = {
            "Content-Type": "application/json",
            "x-acs-dingtalk-access-token": token,
        }

        try:
            conn = utils.http_req(
                url,
                method=method_text,
                headers=headers,
                params=params,
                json=json_data,
                timeout=(8, Config.DINGTALK_KB_TIMEOUT),
            )
            data = _parse_response(conn)
            success = _is_success(data, conn.status_code)
            result = {
                "status_code": conn.status_code,
                "data": data,
                "attempt": attempt,
            }
            if success:
                return True, result

            if not auth_retry_used and _is_auth_failure(conn.status_code, data):
                auth_retry_used = True
                reset_access_token_cache()
                force_refresh = True
                if attempt < max_attempts:
                    continue

            if retry_on_transient and attempt < max_attempts and _is_transient_failure(conn.status_code, data):
                time.sleep(min(0.45 * attempt, 1.5))
                last_result = result
                continue

            return False, result
        except Exception as e:
            error_text = str(e)
            last_result = {
                "error": error_text,
                "attempt": attempt,
            }
            if retry_on_transient and attempt < max_attempts and _is_transient_failure(0, {}, error_text=error_text):
                time.sleep(min(0.45 * attempt, 1.5))
                continue
            return False, last_result

    return False, last_result


def _extract_node_meta(data):
    """
    解析节点信息（兼容不同响应结构）
    """
    node_id = ""
    node_url = ""

    if isinstance(data, dict):
        node_id = data.get("nodeId", "") or data.get("id", "")
        node_url = data.get("url", "")

        if not node_id and isinstance(data.get("node"), dict):
            node_id = data["node"].get("nodeId", "") or data["node"].get("id", "")
            node_url = node_url or data["node"].get("url", "")

        if not node_id and isinstance(data.get("result"), dict):
            node_id = data["result"].get("nodeId", "") or data["result"].get("id", "")
            node_url = node_url or data["result"].get("url", "")

    return {
        "node_id": node_id,
        "node_url": node_url,
    }


def _extract_doc_meta(data):
    """
    解析创建知识库文档/表格响应信息
    """
    node_meta = _extract_node_meta(data)
    doc_key = ""
    dentry_uuid = ""
    workspace_id = ""

    candidates = []
    if isinstance(data, dict):
        candidates.append(data)
        if isinstance(data.get("result"), dict):
            candidates.append(data.get("result"))
        if isinstance(data.get("node"), dict):
            candidates.append(data.get("node"))

    for item in candidates:
        if not doc_key:
            doc_key = item.get("docKey", "")
        if not dentry_uuid:
            dentry_uuid = item.get("dentryUuid", "")
        if not workspace_id:
            workspace_id = item.get("workspaceId", "")

    return {
        "node_id": node_meta.get("node_id", ""),
        "node_url": node_meta.get("node_url", ""),
        "doc_key": doc_key,
        "dentry_uuid": dentry_uuid,
        "workspace_id": workspace_id,
    }


def _normalize_title(title):
    """
    标题清洗：
    - 移除换行，避免接口报错
    - 限制长度，避免超长标题
    """
    clean_title = str(title or "").replace("\r", " ").replace("\n", " ").strip()
    if not clean_title:
        prefix = (Config.DINGTALK_KB_TITLE_PREFIX or "互联网资产自动化收集").strip()
        now_text = utils.curr_date().replace(" ", "_").replace(":", "-")
        clean_title = "{}-{}".format(prefix, now_text)
    return clean_title[:120]


def _normalize_workbook_id(workbook_id):
    """
    规范化 workbook_id（对应 dentry_uuid）
    """
    return str(workbook_id or "").strip().strip("{}")


def _normalize_sheet_name(sheet_name):
    """
    规范化工作表名称
    """
    name = str(sheet_name or "").strip()
    if not name:
        name = "Sheet1"
    return name[:100]


def _normalize_sheet_name_key(sheet_name):
    """
    归一化工作表名称用于排序匹配
    """
    return str(sheet_name or "").strip().lower()


def _compact_sheet_cell_text(value):
    """
    归一化单元格文本（用于表头匹配）。
    """
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", "", text)


def _parse_int_from_cell(value):
    """
    从单元格值解析整数，失败返回 None。
    """
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)

    text = str(value or "").strip()
    if not text:
        return None

    try:
        return int(float(text))
    except Exception:
        match = re.search(r"-?\d+", text)
        if not match:
            return None
        try:
            return int(match.group(0))
        except Exception:
            return None


def _format_validity_days_text(value):
    """
    将剩余天数字段格式化为“证书有效期”文本。
    """
    parsed = _parse_int_from_cell(value)
    if parsed is None:
        return str(value or "-").strip() or "-"
    if parsed < 0:
        return "已过期 {} 天".format(abs(parsed))
    if parsed == 0:
        return "今日到期"
    return "剩余 {} 天".format(parsed)


def _strip_task_id_column(values):
    """
    若首列为任务ID，移除该列（仅用于钉钉知识库展示）。
    """
    if not isinstance(values, list) or not values:
        return []

    normalized_rows = []
    for row in values:
        if isinstance(row, list):
            normalized_rows.append(list(row))
        else:
            normalized_rows.append([row])

    header = normalized_rows[0] if normalized_rows else []
    if not isinstance(header, list) or not header:
        return normalized_rows

    first_col = _compact_sheet_cell_text(header[0])
    if first_col not in ["任务id", "taskid"]:
        return normalized_rows

    output = []
    for row in normalized_rows:
        if isinstance(row, list) and len(row) > 0:
            output.append(row[1:])
        else:
            output.append([])
    return output


def _build_expired_ssl_sheet_item(ssl_sheet_item):
    """
    从 SSL证书 工作表构建“过期证书”工作表（仅过期项）。
    """
    if not isinstance(ssl_sheet_item, dict):
        return None

    values = ssl_sheet_item.get("values", [])
    if not isinstance(values, list) or len(values) == 0:
        return None

    rows = []
    for row in values:
        if isinstance(row, list):
            rows.append(list(row))
        else:
            rows.append([row])

    header = rows[0] if rows else []
    if not isinstance(header, list) or not header:
        return None

    remain_idx = -1
    for idx, cell in enumerate(header):
        if _compact_sheet_cell_text(cell) in ["剩余天数", "证书有效期", "有效期"]:
            remain_idx = idx
            break
    if remain_idx < 0:
        return None

    expired_rows = []
    for row in rows[1:]:
        if not isinstance(row, list):
            continue
        if remain_idx >= len(row):
            continue
        remaining_days = _parse_int_from_cell(row[remain_idx])
        if remaining_days is None or remaining_days >= 0:
            continue
        row_cp = list(row)
        row_cp[remain_idx] = _format_validity_days_text(remaining_days)
        expired_rows.append(row_cp)

    output_header = list(header)
    output_header[remain_idx] = "证书有效期"

    output_values = [output_header]
    if expired_rows:
        output_values.extend(expired_rows)
    else:
        empty_row = ["-"] * max(len(output_header), 1)
        empty_row[0] = "无过期证书"
        output_values.append(empty_row)

    return {
        "sheet_name": "过期证书",
        "values": output_values,
    }


def _prepare_task_export_sheet_items(raw_sheet_items):
    """
    预处理任务导出工作表：
    - SSL证书工作表去掉任务ID列
    - 风险工作表去掉任务ID列（兼容历史导出结构）
    - 追加“过期证书”工作表
    """
    if not isinstance(raw_sheet_items, list):
        return []

    prepared_items = []
    ssl_sheet_item = None

    for item in raw_sheet_items:
        if not isinstance(item, dict):
            continue

        sheet_name = str(item.get("sheet_name", "")).strip()
        values = item.get("values", [])
        current_item = {
            "sheet_name": sheet_name,
            "values": values if isinstance(values, list) else [],
        }

        normalized_key = _normalize_sheet_name_key(sheet_name)
        if normalized_key in [
            _normalize_sheet_name_key("SSL证书"),
            _normalize_sheet_name_key("风险"),
            _normalize_sheet_name_key("漏洞"),
        ]:
            current_item["values"] = _strip_task_id_column(current_item.get("values", []))
        if normalized_key == _normalize_sheet_name_key("SSL证书"):
            ssl_sheet_item = current_item

        prepared_items.append(current_item)

    expired_sheet_item = _build_expired_ssl_sheet_item(ssl_sheet_item)
    if expired_sheet_item:
        prepared_items.append(expired_sheet_item)

    return prepared_items


def _build_ordered_export_sheet_items(raw_sheet_items):
    """
    按固定顺序重排导出工作表
    期望顺序：域名、IP、系统服务、SSL证书、过期证书、站点、风险、资产统计
    """
    preferred_order = ["域名", "IP", "系统服务", "SSL证书", "过期证书", "站点", "风险", "资产统计"]
    preferred_keys = [_normalize_sheet_name_key(name) for name in preferred_order]
    risk_key = _normalize_sheet_name_key("风险")
    legacy_vuln_key = _normalize_sheet_name_key("漏洞")
    sheet_map = {}
    ignored_sheet_names = []

    for item in raw_sheet_items:
        if not isinstance(item, dict):
            continue
        sheet_name = str(item.get("sheet_name", "")).strip()
        if not sheet_name:
            continue
        normalized_key = _normalize_sheet_name_key(sheet_name)
        if normalized_key == legacy_vuln_key:
            normalized_key = risk_key
        if normalized_key == "sheet1":
            continue
        if normalized_key in sheet_map:
            continue
        sheet_map[normalized_key] = {
            "sheet_name": sheet_name,
            "values": item.get("values", []),
        }

    ordered_items = []
    for idx, key in enumerate(preferred_keys):
        if key in sheet_map:
            selected = dict(sheet_map[key])
            selected["sheet_name"] = preferred_order[idx]
            ordered_items.append(selected)

    for key, item in sheet_map.items():
        if key not in preferred_keys:
            ignored_sheet_names.append(item.get("sheet_name", ""))

    return ordered_items, ignored_sheet_names


def _normalize_task_ids(task_ids):
    """
    规范化任务 ID 列表并去重
    """
    if not isinstance(task_ids, list):
        return []

    output = []
    for item in task_ids:
        task_id = str(item or "").strip()
        if not task_id:
            continue
        if task_id in output:
            continue
        output.append(task_id)
    return output


def _column_index_to_name(index):
    """
    列号转 A1 列名（1 -> A, 27 -> AA）
    """
    index = max(int(index), 1)
    result = ""
    while index:
        index, mod = divmod(index - 1, 26)
        result = chr(65 + mod) + result
    return result


def _build_a1_range(row_count, col_count=1):
    """
    构建 A1 范围
    """
    row_count = max(int(row_count), 1)
    col_count = max(int(col_count), 1)
    end_col = _column_index_to_name(col_count)
    return "A1:{}{}".format(end_col, row_count)


def _markdown_to_values(markdown_content, max_rows=500, max_cell_len=1800):
    """
    将 markdown 文本转为表格 values（按行写入）
    """
    text = str(markdown_content or "")
    lines = text.splitlines()
    if not lines:
        lines = ["(empty report)"]

    values = []
    for line in lines[:max_rows]:
        values.append([str(line)[:max_cell_len]])

    return {
        "values": values,
        "row_count": len(values),
        "total_lines": len(lines),
        "truncated": len(lines) > max_rows,
    }


def _normalize_cell_value(value, max_cell_len=1800):
    """
    规范化单元格值，确保可序列化并限制长度
    """
    if value is None:
        return ""

    text = str(value)
    # 统一换行符，去掉每行首尾空格，避免展示出现大量“空行+缩进”
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    if "\n" in text:
        compact_lines = []
        prev_blank = False
        for line in text.split("\n"):
            clean_line = line.strip()
            if not clean_line:
                if prev_blank:
                    continue
                compact_lines.append("")
                prev_blank = True
                continue
            compact_lines.append(clean_line)
            prev_blank = False
        text = "\n".join(compact_lines).strip()

    if len(text) > max_cell_len:
        return text[:max_cell_len]
    return text


def _normalize_sheet_values(values, max_rows=2000, max_cols=26, max_cell_len=1800):
    """
    将二维数据标准化为可写入工作表的矩形 values
    """
    rows = values if isinstance(values, list) else []
    output_rows = []
    for row in rows[:max_rows]:
        if isinstance(row, (list, tuple)):
            row_values = list(row)
        else:
            row_values = [row]

        row_values = row_values[:max_cols]
        normalized_row = [_normalize_cell_value(item, max_cell_len=max_cell_len) for item in row_values]

        while normalized_row and normalized_row[-1] == "":
            normalized_row.pop()
        if not normalized_row:
            normalized_row = [""]

        output_rows.append(normalized_row)

    def _row_is_empty(row):
        return all(str(cell or "").strip() == "" for cell in row)

    # 去掉尾部纯空白行，避免导出模板样式导致的超长空行写入
    while output_rows and all(str(cell or "").strip() == "" for cell in output_rows[-1]):
        output_rows.pop()

    # 仅保留单个空行，压缩视觉噪音
    compact_rows = []
    prev_blank = False
    for row in output_rows:
        is_blank = _row_is_empty(row)
        if is_blank:
            if prev_blank:
                continue
            compact_rows.append([""])
            prev_blank = True
            continue
        compact_rows.append(row)
        prev_blank = False
    output_rows = compact_rows

    # 去掉首尾空行
    while output_rows and _row_is_empty(output_rows[0]):
        output_rows.pop(0)
    while output_rows and _row_is_empty(output_rows[-1]):
        output_rows.pop()

    if not output_rows:
        output_rows = [["(empty sheet)"]]

    col_count = max(len(row) for row in output_rows)
    col_count = max(1, min(col_count, max_cols))
    for row in output_rows:
        if len(row) < col_count:
            row.extend([""] * (col_count - len(row)))
        elif len(row) > col_count:
            del row[col_count:]

    return {
        "values": output_rows,
        "row_count": len(output_rows),
        "col_count": col_count,
        "truncated_rows": len(rows) > max_rows,
    }


def _build_task_overview_sheet_values(title, task_ids, overview_meta=None):
    """
    构建默认 Sheet1 的执行概览
    """
    task_id_list = _normalize_task_ids(task_ids)
    meta = overview_meta if isinstance(overview_meta, dict) else {}
    rows = [
        ["互联网资产自动化收集执行报告", ""],
        ["报告类型", "执行概览"],
        ["报告标题", str(title or "")],
        ["生成时间", utils.curr_date()],
        ["任务数量", str(len(task_id_list))],
        ["任务ID列表", "、".join(task_id_list)],
    ]

    schedule_name = str(meta.get("schedule_name", "") or "").strip()
    if schedule_name:
        rows.append(["计划任务名称", schedule_name])

    run_number = str(meta.get("run_number", "") or "").strip()
    if run_number:
        rows.append(["执行轮次", run_number])

    status_map = {"finished": "已完成", "error": "执行异常", "running": "运行中"}
    raw_status = str(meta.get("status", "") or "").strip().lower()
    if raw_status:
        rows.append(["执行状态", status_map.get(raw_status, raw_status)])

    start_date = str(meta.get("start_date", "") or "").strip()
    end_date = str(meta.get("end_date", "") or "").strip()
    if start_date:
        rows.append(["开始时间", start_date])
    if end_date:
        rows.append(["结束时间", end_date])

    compare_summary = meta.get("compare_summary", {})
    if isinstance(compare_summary, dict):
        if compare_summary.get("has_baseline", False):
            rows.append(["对比基线时间", str(compare_summary.get("baseline_end_date", "") or "-")])
            metric_order = ["site_cnt", "domain_cnt", "ip_cnt", "url_cnt", "vuln_cnt"]
            for metric_key in metric_order:
                metric_item = compare_summary.get("metrics", {}).get(metric_key, {})
                if not isinstance(metric_item, dict):
                    continue
                label = str(metric_item.get("label", metric_key))
                if "漏洞" in label:
                    label = label.replace("漏洞", "风险")
                current_val = int(metric_item.get("current", 0) or 0)
                previous_val = int(metric_item.get("previous", 0) or 0)
                delta_val = int(metric_item.get("delta", 0) or 0)
                ratio_text = str(metric_item.get("ratio_text", "0.00%"))
                rows.append(
                    [
                        "{}变化".format(label),
                        "当前 {} / 较上次 {:+d}（上次 {}，变化 {}）".format(
                            current_val, delta_val, previous_val, ratio_text
                        ),
                    ]
                )
        else:
            rows.append(["对比基线", "首次执行（无历史基线）"])

    rows.extend(
        [
            [""],
            [
                "任务ID",
                "任务名称",
                "任务类型",
                "执行状态",
                "开始时间",
                "结束时间",
                "目标",
                "站点",
                "域名",
                "IP",
                "URL",
                "风险",
            ],
        ]
    )

    if not task_id_list:
        rows.append(["-", "-", "-", "-", "-", "-", "-", "0", "0", "0", "0", "0"])
        return rows

    task_map = {}
    if ObjectId is not None:
        object_ids = []
        for task_id in task_id_list:
            try:
                object_ids.append(ObjectId(task_id))
            except Exception:
                continue

        if object_ids:
            query = {"_id": {"$in": object_ids}}
            projection = {
                "name": 1,
                "type": 1,
                "status": 1,
                "start_time": 1,
                "end_time": 1,
                "target": 1,
                "statistic": 1,
            }
            items = list(utils.conn_db("task").find(query, projection))
            for item in items:
                task_map[str(item.get("_id", ""))] = item

    total_site = 0
    total_domain = 0
    total_ip = 0
    total_url = 0
    total_vuln = 0

    for task_id in task_id_list:
        item = task_map.get(task_id, {})
        statistic = item.get("statistic", {})
        if not isinstance(statistic, dict):
            statistic = {}

        site_cnt = int(statistic.get("site_cnt", 0) or 0)
        domain_cnt = int(statistic.get("domain_cnt", 0) or 0)
        ip_cnt = int(statistic.get("ip_cnt", 0) or 0)
        url_cnt = int(statistic.get("url_cnt", 0) or 0)
        vuln_cnt = int(statistic.get("vuln_cnt", 0) or 0)
        nuclei_vuln_cnt = int(statistic.get("nuclei_result_cnt", 0) or 0)
        total_vuln_cnt = vuln_cnt + nuclei_vuln_cnt

        total_site += site_cnt
        total_domain += domain_cnt
        total_ip += ip_cnt
        total_url += url_cnt
        total_vuln += total_vuln_cnt

        rows.append(
            [
                task_id,
                str(item.get("name", "")),
                str(item.get("type", "")),
                str(item.get("status", "")),
                str(item.get("start_time", "")),
                str(item.get("end_time", "")),
                str(item.get("target", "")),
                str(site_cnt),
                str(domain_cnt),
                str(ip_cnt),
                str(url_cnt),
                str(total_vuln_cnt),
            ]
        )

    rows.append([""])
    rows.append(
        [
            "汇总",
            "",
            "",
            "",
            "",
            "",
            "",
            str(total_site),
            str(total_domain),
            str(total_ip),
            str(total_url),
            str(total_vuln),
        ]
    )
    return rows


def _build_github_overview_sheet_values(title, keyword, result_items, overview_meta=None):
    """
    构建 GitHub 监控知识库执行概览（结构化表格）
    """
    meta = overview_meta if isinstance(overview_meta, dict) else {}
    keyword = str(keyword or "")
    raw_items = result_items if isinstance(result_items, list) else []

    rows = [
        ["互联网资产自动化收集执行报告", ""],
        ["报告类型", "GitHub监控执行概览"],
        ["报告标题", str(title or "")],
        ["生成时间", utils.curr_date()],
        ["关键词", keyword],
    ]

    scheduler_id = str(meta.get("source_id", "") or meta.get("scheduler_id", "") or "")
    if scheduler_id:
        rows.append(["调度ID", scheduler_id])

    unique_repo_set = set()
    normalized_rows = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        repo_name = str(item.get("repo_full_name", "") or "")
        file_path = str(item.get("path", "") or "")
        commit_date = str(item.get("commit_date", "") or "")
        html_url = str(item.get("html_url", "") or "")

        if not (repo_name or file_path or html_url):
            continue

        if repo_name:
            unique_repo_set.add(repo_name)
        normalized_rows.append([repo_name, file_path, commit_date, html_url])

    rows.append(["新增结果数", str(len(normalized_rows))])
    rows.append(["涉及仓库数", str(len(unique_repo_set))])
    rows.append([""])
    rows.append(["序号", "仓库", "文件路径", "最后提交时间", "链接"])

    if not normalized_rows:
        rows.append(["-", "-", "-", "-", "-"])
        return rows

    max_rows = 1000
    for idx, item in enumerate(normalized_rows[:max_rows], 1):
        rows.append([str(idx)] + item)

    if len(normalized_rows) > max_rows:
        rows.append(["", "", "", "", "已截断，原始结果总数：{}".format(len(normalized_rows))])

    return rows


def _load_workbook_sheet_items(excel_bytes, max_sheets=10, max_rows=2000, max_cols=26):
    """
    解析 Excel 二进制内容，提取工作表二维数据
    """
    if load_workbook is None:
        return False, {"error": "openpyxl unavailable"}

    if not excel_bytes:
        return False, {"error": "excel bytes is empty"}

    try:
        workbook = load_workbook(filename=BytesIO(excel_bytes), read_only=True, data_only=True)
    except Exception as e:
        return False, {"error": "load workbook failed", "detail": str(e)}

    items = []
    truncated_sheets = len(workbook.worksheets) > max_sheets
    try:
        for ws in workbook.worksheets[:max_sheets]:
            raw_rows = []
            for row in ws.iter_rows(values_only=True):
                raw_rows.append(list(row))

            normalized = _normalize_sheet_values(
                raw_rows, max_rows=max_rows, max_cols=max_cols, max_cell_len=1800
            )
            items.append(
                {
                    "sheet_name": _normalize_sheet_name(ws.title),
                    "values": normalized["values"],
                    "row_count": normalized["row_count"],
                    "col_count": normalized["col_count"],
                    "truncated_rows": normalized["truncated_rows"],
                }
            )
    finally:
        try:
            workbook.close()
        except Exception:
            pass

    if not items:
        items = [
            {
                "sheet_name": "Sheet1",
                "values": [["(empty workbook)"]],
                "row_count": 1,
                "col_count": 1,
                "truncated_rows": False,
            }
        ]

    return True, {"items": items, "sheet_count": len(items), "truncated_sheets": truncated_sheets}


def _extract_error_text(detail):
    """
    从接口返回中提取可读错误信息
    """
    if not isinstance(detail, dict):
        return str(detail or "")

    for key in ["error", "error_message", "message"]:
        val = detail.get(key)
        if val:
            return str(val)

    data = detail.get("data")
    if isinstance(data, dict):
        for key in ["error", "message", "msg", "errMsg", "errmsg", "error_message"]:
            val = data.get(key)
            if val:
                return str(val)

    return str(detail)


def _resolve_create_doc_path(workspace_id=""):
    """
    解析创建文档路径：
    - 兼容旧配置 /v2.0/wiki/nodes（自动切换为表格接口）
    - 支持 {workspace_id}/{workspaceId} 占位符
    """
    path = str(Config.DINGTALK_KB_CREATE_NODE_PATH or "").strip()
    if not path or path == _LEGACY_MARKDOWN_PATH:
        path = _DEFAULT_CREATE_DOC_PATH

    resolved_workspace = str(workspace_id or Config.DINGTALK_WORKSPACE_ID or "").strip()
    path = path.replace("{workspace_id}", resolved_workspace).replace("{workspaceId}", resolved_workspace)
    return path


def _build_create_workbook_payload(title, operator_id="", parent_node_id=""):
    """
    构建创建知识库表格（WORKBOOK）的请求体
    """
    return {
        "name": _normalize_title(title),
        "docType": "WORKBOOK",
        "operatorId": operator_id or Config.DINGTALK_OPERATOR_ID,
        "parentNodeId": parent_node_id or Config.DINGTALK_PARENT_NODE_ID,
    }


def get_runtime_status():
    """
    获取当前钉钉开放平台配置状态（脱敏）
    """
    return {
        "enable": bool(Config.DINGTALK_KB_ENABLE),
        "base_url": Config.DINGTALK_API_BASE_URL,
        "corp_id": Config.DINGTALK_CORP_ID,
        "app_key": Config.DINGTALK_APP_KEY,
        "app_secret_set": bool(str(Config.DINGTALK_APP_SECRET or "").strip()),
        "operator_id": Config.DINGTALK_OPERATOR_ID,
        "workspace_id": Config.DINGTALK_WORKSPACE_ID,
        "parent_node_id": Config.DINGTALK_PARENT_NODE_ID,
        "create_node_path": Config.DINGTALK_KB_CREATE_NODE_PATH,
        "title_prefix": Config.DINGTALK_KB_TITLE_PREFIX,
        "ssl_cert_notify_enable": bool(Config.DINGTALK_SSL_CERT_NOTIFY_ENABLE),
        "ssl_cert_notify_days": int(Config.DINGTALK_SSL_CERT_NOTIFY_DAYS or 30),
        "missing_basic_fields": _missing_required_fields(),
        "missing_publish_fields": _missing_required_fields(require_workspace=True, require_parent_node=True),
    }


def test_connection(force_refresh_token=False):
    """
    测试钉钉开放平台连通性（基于当前配置）
    """
    missing_fields = _missing_required_fields()
    if missing_fields:
        return False, {
            "error": "dingtalk api config missing",
            "missing_fields": missing_fields,
        }

    token = get_access_token(force_refresh=force_refresh_token, require_enable=False)
    if not token:
        return False, {"error": "get_access_token failed", "missing_fields": missing_fields}

    success, result = request_openapi(
        method="get",
        path="/v2.0/wiki/workspaces",
        params={"operatorId": Config.DINGTALK_OPERATOR_ID},
        require_enable=False,
        force_refresh_token=force_refresh_token,
        retry_on_transient=True,
        retry_max=3,
    )
    result["access_token_ok"] = True
    if success:
        data = result.get("data", {})
        if isinstance(data, dict):
            result["workspace_count"] = len(data.get("workspaces", []) or [])
    return success, result


def list_workspaces(operator_id=""):
    """
    获取知识库空间列表
    """
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()
    missing_fields = _missing_required_fields()
    if not op_id:
        missing_fields.append("OPERATOR_ID")
    if missing_fields:
        return False, {"error": "dingtalk api config missing", "missing_fields": sorted(set(missing_fields))}

    success, result = request_openapi(
        method="get",
        path="/v2.0/wiki/workspaces",
        params={"operatorId": op_id},
        require_enable=False,
        retry_on_transient=True,
        retry_max=3,
    )

    data = result.get("data", {})
    workspaces = []
    if isinstance(data, dict):
        for item in data.get("workspaces", []) or []:
            if not isinstance(item, dict):
                continue
            workspaces.append(
                {
                    "workspace_id": item.get("workspaceId", ""),
                    "name": item.get("name", ""),
                    "root_node_id": item.get("rootNodeId", ""),
                    "url": item.get("url", ""),
                    "type": item.get("type", ""),
                }
            )
    result["items"] = workspaces
    result["operator_id"] = op_id
    return success, result


def list_nodes(parent_node_id="", operator_id=""):
    """
    获取指定父节点下的知识库目录节点列表
    """
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()
    parent_id = str(parent_node_id or Config.DINGTALK_PARENT_NODE_ID or "").strip()

    missing_fields = _missing_required_fields()
    if not op_id:
        missing_fields.append("OPERATOR_ID")
    if not parent_id:
        missing_fields.append("PARENT_NODE_ID")
    if missing_fields:
        return False, {"error": "dingtalk api config missing", "missing_fields": sorted(set(missing_fields))}

    success, result = request_openapi(
        method="get",
        path="/v2.0/wiki/nodes",
        params={
            "parentNodeId": parent_id,
            "operatorId": op_id,
        },
        require_enable=False,
        retry_on_transient=True,
        retry_max=3,
    )

    data = result.get("data", {})
    nodes = []
    if isinstance(data, dict):
        for item in data.get("nodes", []) or []:
            if not isinstance(item, dict):
                continue
            nodes.append(
                {
                    "node_id": item.get("nodeId", ""),
                    "name": item.get("name", ""),
                    "type": item.get("type", ""),
                    "has_children": bool(item.get("hasChildren", False)),
                    "workspace_id": item.get("workspaceId", ""),
                    "url": item.get("url", ""),
                }
            )

    result["items"] = nodes
    result["parent_node_id"] = parent_id
    result["operator_id"] = op_id
    return success, result


def list_workbook_sheets(workbook_id, operator_id="", require_enable=False):
    """
    获取 workbook 下所有工作表
    """
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()
    wb_id = _normalize_workbook_id(workbook_id)

    missing_fields = _missing_required_fields()
    if not op_id:
        missing_fields.append("OPERATOR_ID")
    if not wb_id:
        missing_fields.append("WORKBOOK_ID")
    if missing_fields:
        return False, {"error": "dingtalk api config missing", "missing_fields": sorted(set(missing_fields))}

    path = "/v1.0/doc/workbooks/{}/sheets".format(urllib.parse.quote(wb_id, safe=""))
    success, result = request_openapi(
        method="get",
        path=path,
        params={"operatorId": op_id},
        require_enable=require_enable,
    )

    data = result.get("data", {})
    sheet_items = []
    if isinstance(data, dict):
        for item in data.get("value", []) or []:
            if not isinstance(item, dict):
                continue
            sheet_items.append(
                {
                    "sheet_id": item.get("id", ""),
                    "name": item.get("name", ""),
                    "visibility": item.get("visibility", ""),
                }
            )

    result["items"] = sheet_items
    result["workbook_id"] = wb_id
    result["operator_id"] = op_id
    result["request_path"] = path
    return success, result


def create_workbook_sheet(workbook_id, sheet_name, operator_id="", require_enable=False):
    """
    创建工作表
    """
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()
    wb_id = _normalize_workbook_id(workbook_id)
    sheet_name = _normalize_sheet_name(sheet_name)

    missing_fields = _missing_required_fields()
    if not op_id:
        missing_fields.append("OPERATOR_ID")
    if not wb_id:
        missing_fields.append("WORKBOOK_ID")
    if missing_fields:
        return False, {"error": "dingtalk api config missing", "missing_fields": sorted(set(missing_fields))}

    path = "/v1.0/doc/workbooks/{}/sheets".format(urllib.parse.quote(wb_id, safe=""))
    payload = {"name": sheet_name}
    success, result = request_openapi(
        method="post",
        path=path,
        params={"operatorId": op_id},
        json_data=payload,
        require_enable=require_enable,
    )
    result["workbook_id"] = wb_id
    result["operator_id"] = op_id
    result["request_path"] = path
    result["request_payload"] = payload
    if success and isinstance(result.get("data"), dict):
        result["sheet_id"] = result["data"].get("id", "")
        result["sheet_name"] = result["data"].get("name", sheet_name)
    return success, result


def rename_workbook_sheet(workbook_id, sheet_ref, new_sheet_name, operator_id="", require_enable=False):
    """
    重命名工作表（尽力而为，不影响主流程）
    """
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()
    wb_id = _normalize_workbook_id(workbook_id)
    ref = str(sheet_ref or "").strip()
    target_name = _normalize_sheet_name(new_sheet_name)

    missing_fields = _missing_required_fields()
    if not op_id:
        missing_fields.append("OPERATOR_ID")
    if not wb_id:
        missing_fields.append("WORKBOOK_ID")
    if not ref:
        missing_fields.append("SHEET_REF")
    if not target_name:
        missing_fields.append("SHEET_NAME")
    if missing_fields:
        return False, {"error": "dingtalk api config missing", "missing_fields": sorted(set(missing_fields))}

    path = "/v1.0/doc/workbooks/{}/sheets/{}".format(
        urllib.parse.quote(wb_id, safe=""),
        urllib.parse.quote(ref, safe=""),
    )
    payload = {"name": target_name}
    params = {"operatorId": op_id}

    success, result = request_openapi(
        method="patch",
        path=path,
        params=params,
        json_data=payload,
        require_enable=require_enable,
    )
    if not success:
        success, result = request_openapi(
            method="put",
            path=path,
            params=params,
            json_data=payload,
            require_enable=require_enable,
        )

    result["workbook_id"] = wb_id
    result["sheet_ref"] = ref
    result["sheet_name"] = target_name
    result["operator_id"] = op_id
    result["request_path"] = path
    result["request_payload"] = payload
    return success, result


def update_workbook_range(workbook_id, sheet_name, range_a1, values, operator_id="", require_enable=False):
    """
    更新工作表单元格区域
    """
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()
    wb_id = _normalize_workbook_id(workbook_id)
    sheet_name = _normalize_sheet_name(sheet_name)
    range_a1 = str(range_a1 or "").strip()

    missing_fields = _missing_required_fields()
    if not op_id:
        missing_fields.append("OPERATOR_ID")
    if not wb_id:
        missing_fields.append("WORKBOOK_ID")
    if not range_a1:
        missing_fields.append("RANGE")
    if missing_fields:
        return False, {"error": "dingtalk api config missing", "missing_fields": sorted(set(missing_fields))}

    if not isinstance(values, list) or not values:
        return False, {"error": "values is empty"}

    safe_workbook_id = urllib.parse.quote(wb_id, safe="")
    safe_sheet_name = urllib.parse.quote(sheet_name, safe="")
    safe_range = urllib.parse.quote(range_a1, safe=":")
    path = "/v1.0/doc/workbooks/{}/sheets/{}/ranges/{}".format(
        safe_workbook_id, safe_sheet_name, safe_range
    )
    # 再做一次请求前规范化，避免数值类型触发 MissingString
    row_limit = max(len(values), 1) if isinstance(values, list) else 1
    col_limit = 1
    if isinstance(values, list):
        for row in values:
            if isinstance(row, (list, tuple)):
                col_limit = max(col_limit, len(row))
            else:
                col_limit = max(col_limit, 1)
    normalized = _normalize_sheet_values(values, max_rows=row_limit, max_cols=max(col_limit, 1), max_cell_len=1800)
    request_values = normalized.get("values", [[""]])

    # 仅提交 values，避免 wordWrap 等可选样式字段在不同版本接口上校验不一致
    payload = {
        "values": request_values,
    }
    success, result = request_openapi(
        method="put",
        path=path,
        params={"operatorId": op_id},
        json_data=payload,
        require_enable=require_enable,
    )
    result["workbook_id"] = wb_id
    result["sheet_name"] = sheet_name
    result["range"] = range_a1
    result["request_path"] = path
    result["request_payload"] = payload
    return success, result


def write_sheet_values_to_workbook(
    workbook_id,
    values,
    operator_id="",
    sheet_name="",
    require_enable=False,
    fallback_workbook_ids=None,
):
    """
    将二维 values 写入 workbook 的指定 sheet
    """
    wb_id = _normalize_workbook_id(workbook_id)
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()
    requested_sheet_name = _normalize_sheet_name(sheet_name or "Sheet1")
    sheet_list_retry_max = 8
    sheet_create_retry_max = 5
    write_retry_max = 5

    workbook_candidates = []
    for item in [wb_id] + list(fallback_workbook_ids or []):
        normalized = _normalize_workbook_id(item)
        if normalized and normalized not in workbook_candidates:
            workbook_candidates.append(normalized)

    if not workbook_candidates:
        return False, {"error": "workbook_id is empty"}

    values_meta = _normalize_sheet_values(values)
    safe_values = values_meta.get("values", [])
    row_count = int(values_meta.get("row_count", len(safe_values)) or 1)
    col_count = int(values_meta.get("col_count", 1) or 1)
    range_a1 = _build_a1_range(row_count=row_count, col_count=col_count)

    last_error = {}
    for candidate_id in workbook_candidates:
        target_sheet_name = requested_sheet_name
        target_sheet_id = ""

        # 列表接口偶发延迟，做短重试
        sheet_success, sheet_result = False, {}
        sheet_items = []
        for retry_idx in range(sheet_list_retry_max):
            sheet_success, sheet_result = list_workbook_sheets(
                workbook_id=candidate_id,
                operator_id=op_id,
                require_enable=require_enable,
            )
            if not sheet_success:
                if retry_idx < sheet_list_retry_max - 1:
                    time.sleep(0.6 * (retry_idx + 1))
                continue

            current_items = sheet_result.get("items", []) or []
            if isinstance(current_items, list):
                sheet_items = current_items

            # 创建文档后，sheet 列表有概率短暂为空，稍等后重试一次
            if sheet_items or retry_idx == sheet_list_retry_max - 1:
                break
            time.sleep(0.6 * (retry_idx + 1))

        if not sheet_success:
            last_error = {
                "error": "list workbook sheets failed",
                "workbook_id": candidate_id,
                "workbook_candidates": workbook_candidates,
                "sheet_result": sheet_result,
                "error_text": _extract_error_text(sheet_result),
            }
            continue

        sheet_name_map = {}
        for item in sheet_items:
            if isinstance(item, dict):
                item_name = str(item.get("name", "")).strip()
                if item_name:
                    sheet_name_map[item_name] = item

        if target_sheet_name not in sheet_name_map:
            should_fallback_first_sheet = (
                bool(sheet_items) and requested_sheet_name in ["", "Sheet1"]
            )
            if should_fallback_first_sheet:
                first_sheet = sheet_items[0] if isinstance(sheet_items[0], dict) else {}
                target_sheet_name = str(first_sheet.get("name", "") or target_sheet_name)
                target_sheet_id = str(first_sheet.get("sheet_id", "") or "")
            elif requested_sheet_name in ["", "Sheet1"]:
                # 默认页优先写入，不主动创建，避免生成额外 SheetN
                target_sheet_name = "Sheet1"
                target_sheet_id = ""
            else:
                create_success, create_result = False, {}
                for retry_idx in range(sheet_create_retry_max):
                    create_success, create_result = create_workbook_sheet(
                        workbook_id=candidate_id,
                        sheet_name=requested_sheet_name,
                        operator_id=op_id,
                        require_enable=require_enable,
                    )
                    if create_success:
                        break
                    if retry_idx < sheet_create_retry_max - 1:
                        time.sleep(0.6 * (retry_idx + 1))

                if not create_success:
                    # 兼容“创建实际上成功但响应失败”的场景，再查一次列表确认
                    relist_success, relist_result = list_workbook_sheets(
                        workbook_id=candidate_id,
                        operator_id=op_id,
                        require_enable=require_enable,
                    )
                    if relist_success:
                        relist_items = relist_result.get("items", []) or []
                        for relist_item in relist_items:
                            if not isinstance(relist_item, dict):
                                continue
                            relist_name = str(relist_item.get("name", "")).strip()
                            if relist_name:
                                sheet_name_map[relist_name] = relist_item
                        if requested_sheet_name in sheet_name_map:
                            create_success = True
                            create_result = {
                                "sheet_name": requested_sheet_name,
                                "sheet_id": sheet_name_map[requested_sheet_name].get("sheet_id", ""),
                            }

                if not create_success:
                    last_error = {
                        "error": "create workbook sheet failed",
                        "workbook_id": candidate_id,
                        "workbook_candidates": workbook_candidates,
                        "sheet_result": sheet_result,
                        "create_result": create_result,
                        "error_text": _extract_error_text(create_result),
                    }
                    continue
                target_sheet_name = str(create_result.get("sheet_name", target_sheet_name) or target_sheet_name)
                target_sheet_id = str(create_result.get("sheet_id", "") or "")
        else:
            target_sheet = sheet_name_map.get(target_sheet_name, {})
            target_sheet_id = str(target_sheet.get("sheet_id", "") or "")

        # 写入接口偶发延迟，做短重试
        write_success, write_result = False, {}
        write_sheet_refs = [target_sheet_name]
        if target_sheet_id and target_sheet_id not in write_sheet_refs:
            write_sheet_refs.append(target_sheet_id)

        for sheet_ref in write_sheet_refs:
            for retry_idx in range(write_retry_max):
                write_success, write_result = update_workbook_range(
                    workbook_id=candidate_id,
                    sheet_name=sheet_ref,
                    range_a1=range_a1,
                    values=safe_values,
                    operator_id=op_id,
                    require_enable=require_enable,
                )
                if write_success:
                    break
                if retry_idx < write_retry_max - 1:
                    time.sleep(0.6 * (retry_idx + 1))
            if write_success:
                break

        output = {
            "workbook_id": candidate_id,
            "workbook_candidates": workbook_candidates,
            "sheet_name": target_sheet_name,
            "sheet_id": target_sheet_id,
            "range": range_a1,
            "row_count": row_count,
            "col_count": col_count,
            "truncated_rows": bool(values_meta.get("truncated_rows", False)),
            "sheet_result": sheet_result,
            "write_result": write_result,
        }
        if write_success:
            return True, output

        last_error = {
            "error": "write workbook range failed",
            "workbook_id": candidate_id,
            "workbook_candidates": workbook_candidates,
            "sheet_name": target_sheet_name,
            "sheet_id": target_sheet_id,
            "range": range_a1,
            "sheet_result": sheet_result,
            "write_result": write_result,
            "error_text": _extract_error_text(write_result),
        }

    return False, last_error or {"error": "write workbook failed"}


def write_sheet_items_to_workbook(
    workbook_id,
    sheet_items,
    operator_id="",
    require_enable=False,
    fallback_workbook_ids=None,
):
    """
    将多个工作表内容写入 workbook
    """
    if not isinstance(sheet_items, list) or not sheet_items:
        return False, {"error": "sheet_items is empty"}

    output_items = []
    success_count = 0
    failed_count = 0
    last_error = {}
    active_workbook_id = _normalize_workbook_id(workbook_id)
    active_fallback_ids = list(fallback_workbook_ids or [])

    for idx, item in enumerate(sheet_items, 1):
        if not isinstance(item, dict):
            item = {}
        sheet_name = _normalize_sheet_name(item.get("sheet_name", "Sheet{}".format(idx)))
        values = item.get("values", [])

        current_fallback_ids = active_fallback_ids if idx == 1 else []
        sheet_success, sheet_result = write_sheet_values_to_workbook(
            workbook_id=active_workbook_id,
            values=values,
            operator_id=operator_id,
            sheet_name=sheet_name,
            require_enable=require_enable,
            fallback_workbook_ids=current_fallback_ids,
        )
        output_items.append(
            {
                "index": idx,
                "sheet_name": sheet_name,
                "success": sheet_success,
                "result": sheet_result,
            }
        )
        if sheet_success:
            success_count += 1
            used_workbook_id = _normalize_workbook_id(sheet_result.get("workbook_id", ""))
            if used_workbook_id:
                active_workbook_id = used_workbook_id
            # 锁定命中的 workbook_id，后续不再切换候选，避免跨 id 产生额外 sheet
            active_fallback_ids = []
        else:
            failed_count += 1
            last_error = sheet_result

    output = {
        "sheet_count": len(output_items),
        "sheet_success_count": success_count,
        "sheet_failed_count": failed_count,
        "items": output_items,
        "workbook_id": active_workbook_id,
    }
    if failed_count == 0:
        return True, output

    output["error"] = "write workbook sheets failed"
    output["last_error"] = last_error
    return False, output


def write_markdown_to_workbook(
    workbook_id,
    markdown_content,
    operator_id="",
    sheet_name="",
    require_enable=False,
    fallback_workbook_ids=None,
):
    """
    将 markdown 文本写入 workbook
    """
    markdown_meta = _markdown_to_values(markdown_content)
    write_success, write_result = write_sheet_values_to_workbook(
        workbook_id=workbook_id,
        values=markdown_meta.get("values", []),
        operator_id=operator_id,
        sheet_name=sheet_name,
        require_enable=require_enable,
        fallback_workbook_ids=fallback_workbook_ids,
    )
    write_result["total_lines"] = markdown_meta.get("total_lines", 0)
    write_result["truncated"] = bool(markdown_meta.get("truncated", False))
    return write_success, write_result


def create_workbook(title, workspace_id="", parent_node_id="", operator_id="", require_enable=False):
    """
    在钉钉知识库中创建 WORKBOOK 文档
    """
    ws_id = str(workspace_id or Config.DINGTALK_WORKSPACE_ID or "").strip()
    parent_id = str(parent_node_id or Config.DINGTALK_PARENT_NODE_ID or "").strip()
    op_id = str(operator_id or Config.DINGTALK_OPERATOR_ID or "").strip()

    missing_fields = _missing_required_fields()
    if not ws_id:
        missing_fields.append("WORKSPACE_ID")
    if not parent_id:
        missing_fields.append("PARENT_NODE_ID")
    if not op_id:
        missing_fields.append("OPERATOR_ID")
    if missing_fields:
        return False, {"error": "dingtalk api config missing", "missing_fields": sorted(set(missing_fields))}

    path = _resolve_create_doc_path(workspace_id=ws_id)
    payload = _build_create_workbook_payload(title, operator_id=op_id, parent_node_id=parent_id)
    success, result = request_openapi(
        method="post",
        path=path,
        json_data=payload,
        require_enable=require_enable,
    )
    result["request_path"] = path
    result["request_payload"] = payload
    result["workspace_id"] = ws_id
    if success:
        result.update(_extract_doc_meta(result.get("data", {})))
        workbook_candidates = []
        dentry_uuid = result.get("dentry_uuid", "")
        doc_key = result.get("doc_key", "")
        if dentry_uuid:
            workbook_candidates.append(dentry_uuid)
        if doc_key and doc_key not in workbook_candidates:
            workbook_candidates.append(doc_key)
        result["workbook_candidates"] = workbook_candidates

    return success, result


def publish_markdown_to_kb(title, markdown_content):
    """
    创建钉钉知识库表格并写入 Markdown 行内容

    返回：
        (success: bool, result: dict)
    """
    if not _is_config_ready(require_enable=True, require_workspace=True, require_parent_node=True):
        return False, {
            "error": "dingtalk kb config incomplete or disabled",
            "missing_fields": _missing_required_fields(require_workspace=True, require_parent_node=True),
        }

    if Config.DINGTALK_KB_DRY_RUN:
        return True, {
            "dry_run": True,
            "title": title,
            "node_id": "",
            "node_url": "",
            "create_path": _resolve_create_doc_path(workspace_id=Config.DINGTALK_WORKSPACE_ID),
        }

    create_success, create_result = create_workbook(
        title=title,
        workspace_id=Config.DINGTALK_WORKSPACE_ID,
        parent_node_id=Config.DINGTALK_PARENT_NODE_ID,
        operator_id=Config.DINGTALK_OPERATOR_ID,
        require_enable=True,
    )
    if not create_success:
        return False, create_result

    workbook_id = _normalize_workbook_id(create_result.get("dentry_uuid", ""))
    create_result["workbook_id"] = workbook_id

    if not workbook_id:
        create_result["error"] = "workbook_id missing from create response"
        return False, create_result

    write_success, write_result = write_markdown_to_workbook(
        workbook_id=workbook_id,
        markdown_content=markdown_content,
        operator_id=Config.DINGTALK_OPERATOR_ID,
        sheet_name="Sheet1",
        require_enable=True,
        fallback_workbook_ids=[create_result.get("doc_key", "")],
    )
    create_result["write_result"] = write_result
    if not write_success:
        create_result["error"] = "write workbook failed"
        return False, create_result

    return True, create_result


def publish_github_monitor_to_kb(title, keyword, result_items, overview_context=None):
    """
    创建钉钉知识库表格并写入 GitHub 监控结构化结果
    """
    if not _is_config_ready(require_enable=True, require_workspace=True, require_parent_node=True):
        return False, {
            "error": "dingtalk kb config incomplete or disabled",
            "missing_fields": _missing_required_fields(require_workspace=True, require_parent_node=True),
        }

    normalized_keyword = str(keyword or "")
    raw_items = result_items if isinstance(result_items, list) else []
    overview_meta = overview_context if isinstance(overview_context, dict) else {}

    if Config.DINGTALK_KB_DRY_RUN:
        return True, {
            "dry_run": True,
            "title": title,
            "keyword": normalized_keyword,
            "result_count": len(raw_items),
            "node_id": "",
            "node_url": "",
            "create_path": _resolve_create_doc_path(workspace_id=Config.DINGTALK_WORKSPACE_ID),
        }

    create_success, create_result = create_workbook(
        title=title,
        workspace_id=Config.DINGTALK_WORKSPACE_ID,
        parent_node_id=Config.DINGTALK_PARENT_NODE_ID,
        operator_id=Config.DINGTALK_OPERATOR_ID,
        require_enable=True,
    )
    if not create_success:
        return False, create_result

    workbook_id = _normalize_workbook_id(create_result.get("dentry_uuid", ""))
    create_result["workbook_id"] = workbook_id
    create_result["keyword"] = normalized_keyword
    create_result["result_count"] = len(raw_items)

    if not workbook_id:
        create_result["error"] = "workbook_id missing from create response"
        return False, create_result

    overview_values = _build_github_overview_sheet_values(
        title=title,
        keyword=normalized_keyword,
        result_items=raw_items,
        overview_meta=overview_meta,
    )

    write_success, write_result = write_sheet_values_to_workbook(
        workbook_id=workbook_id,
        values=overview_values,
        operator_id=Config.DINGTALK_OPERATOR_ID,
        sheet_name="Sheet1",
        require_enable=True,
        fallback_workbook_ids=[create_result.get("doc_key", "")],
    )
    create_result["write_result"] = write_result
    create_result["sheet_items"] = [{"sheet_name": "执行概览", "values": overview_values}]
    create_result["sheet_count"] = 1
    create_result["sheet_write_result"] = {
        "sheet_count": 1,
        "sheet_success_count": 1 if write_success else 0,
        "sheet_failed_count": 0 if write_success else 1,
        "items": [
            {
                "index": 1,
                "sheet_name": "执行概览",
                "success": write_success,
                "result": write_result,
            }
        ],
    }
    if not write_success:
        create_result["sheet_write_result"]["error"] = "write workbook sheets failed"
        create_result["sheet_write_result"]["last_error"] = write_result
        create_result["error"] = "write workbook failed"
        return False, create_result

    resolved_workbook_id = _normalize_workbook_id(write_result.get("workbook_id", "")) or workbook_id
    rename_success, rename_result = rename_workbook_sheet(
        workbook_id=resolved_workbook_id,
        sheet_ref=write_result.get("sheet_id", "") or write_result.get("sheet_name", "") or "Sheet1",
        new_sheet_name="执行概览",
        operator_id=Config.DINGTALK_OPERATOR_ID,
        require_enable=True,
    )
    create_result["overview_rename"] = {
        "success": bool(rename_success),
        "result": rename_result,
    }

    return True, create_result


def publish_task_export_to_kb(title, task_ids, overview_context=None):
    """
    创建钉钉知识库表格并写入任务批量导出内容（多工作表）
    """
    if not _is_config_ready(require_enable=True, require_workspace=True, require_parent_node=True):
        return False, {
            "error": "dingtalk kb config incomplete or disabled",
            "missing_fields": _missing_required_fields(require_workspace=True, require_parent_node=True),
        }

    normalized_task_ids = _normalize_task_ids(task_ids)
    if not normalized_task_ids:
        return False, {"error": "task_ids is empty"}

    if Config.DINGTALK_KB_DRY_RUN:
        return True, {
            "dry_run": True,
            "title": title,
            "task_count": len(normalized_task_ids),
            "task_ids": normalized_task_ids,
            "node_id": "",
            "node_url": "",
            "create_path": _resolve_create_doc_path(workspace_id=Config.DINGTALK_WORKSPACE_ID),
        }

    try:
        from app.routes.export import export_merge_tasks

        excel_bytes = export_merge_tasks(normalized_task_ids)
    except Exception as e:
        return False, {"error": "export merge tasks failed", "detail": str(e)}

    parse_success, parse_result = _load_workbook_sheet_items(excel_bytes)
    if not parse_success:
        return False, {"error": "parse export workbook failed", "detail": parse_result}

    create_success, create_result = create_workbook(
        title=title,
        workspace_id=Config.DINGTALK_WORKSPACE_ID,
        parent_node_id=Config.DINGTALK_PARENT_NODE_ID,
        operator_id=Config.DINGTALK_OPERATOR_ID,
        require_enable=True,
    )
    if not create_success:
        return False, create_result

    workbook_id = _normalize_workbook_id(create_result.get("dentry_uuid", ""))
    create_result["workbook_id"] = workbook_id

    if not workbook_id:
        create_result["error"] = "workbook_id missing from create response"
        return False, create_result

    raw_sheet_items = parse_result.get("items", [])
    prepared_sheet_items = _prepare_task_export_sheet_items(raw_sheet_items)
    ordered_sheet_items, ignored_sheet_names = _build_ordered_export_sheet_items(prepared_sheet_items)

    overview_values = _build_task_overview_sheet_values(
        title=title,
        task_ids=normalized_task_ids,
        overview_meta=overview_context,
    )
    overview_success, overview_result = write_sheet_values_to_workbook(
        workbook_id=workbook_id,
        values=overview_values,
        operator_id=Config.DINGTALK_OPERATOR_ID,
        sheet_name="Sheet1",
        require_enable=True,
        fallback_workbook_ids=[create_result.get("doc_key", "")],
    )
    if not overview_success:
        create_result["task_ids"] = normalized_task_ids
        create_result["sheet_items"] = [{"sheet_name": "执行概览", "values": overview_values}] + ordered_sheet_items
        create_result["sheet_write_result"] = {
            "sheet_count": len(create_result["sheet_items"]),
            "sheet_success_count": 0,
            "sheet_failed_count": 1,
            "items": [
                {
                    "index": 1,
                    "sheet_name": "Sheet1",
                    "success": False,
                    "result": overview_result,
                }
            ],
            "error": "write workbook sheets failed",
            "last_error": overview_result,
        }
        create_result["sheet_count"] = len(create_result["sheet_items"])
        create_result["truncated_sheets"] = bool(parse_result.get("truncated_sheets", False))
        create_result["ignored_sheet_names"] = ignored_sheet_names
        create_result["error"] = "write workbook sheet items failed"
        return False, create_result

    resolved_workbook_id = _normalize_workbook_id(overview_result.get("workbook_id", "")) or workbook_id
    overview_sheet_name = "执行概览"
    rename_success, rename_result = rename_workbook_sheet(
        workbook_id=resolved_workbook_id,
        sheet_ref=overview_result.get("sheet_id", "") or overview_result.get("sheet_name", "") or "Sheet1",
        new_sheet_name=overview_sheet_name,
        operator_id=Config.DINGTALK_OPERATOR_ID,
        require_enable=True,
    )
    if rename_success and isinstance(rename_result, dict):
        overview_sheet_name = str(rename_result.get("sheet_name", overview_sheet_name) or overview_sheet_name)
    else:
        # 重命名失败时保留原默认名称，不影响报告写入
        overview_sheet_name = str(overview_result.get("sheet_name", "Sheet1") or "Sheet1")

    detail_success = True
    detail_result = {
        "sheet_count": 0,
        "sheet_success_count": 0,
        "sheet_failed_count": 0,
        "items": [],
        "workbook_id": resolved_workbook_id,
    }
    # 钉钉当前行为：新建工作表会插入到前面，按反向创建可保证最终展示顺序正确
    detail_write_items = list(reversed(ordered_sheet_items))
    if ordered_sheet_items:
        detail_success, detail_result = write_sheet_items_to_workbook(
            workbook_id=resolved_workbook_id,
            sheet_items=detail_write_items,
            operator_id=Config.DINGTALK_OPERATOR_ID,
            require_enable=True,
            fallback_workbook_ids=[],
        )

    write_items = [
        {
            "index": 1,
            "sheet_name": overview_sheet_name,
            "success": True,
            "result": overview_result,
        }
    ]
    for idx, item in enumerate(detail_result.get("items", []), 2):
        if not isinstance(item, dict):
            continue
        output_item = dict(item)
        output_item["index"] = idx
        write_items.append(output_item)

    sheet_success_count = 1 + int(detail_result.get("sheet_success_count", 0) or 0)
    sheet_failed_count = int(detail_result.get("sheet_failed_count", 0) or 0)
    write_success = detail_success and sheet_failed_count == 0
    write_result = {
        "sheet_count": len(write_items),
        "sheet_success_count": sheet_success_count,
        "sheet_failed_count": sheet_failed_count,
        "items": write_items,
        "workbook_id": _normalize_workbook_id(detail_result.get("workbook_id", "")) or resolved_workbook_id,
        "ignored_sheet_names": ignored_sheet_names,
        "overview_rename": {
            "success": bool(rename_success),
            "result": rename_result,
        },
    }
    if not write_success:
        write_result["error"] = "write workbook sheets failed"
        write_result["last_error"] = detail_result.get("last_error", detail_result)

    sheet_items = [{"sheet_name": overview_sheet_name, "values": overview_values}] + ordered_sheet_items
    create_result["task_ids"] = normalized_task_ids
    create_result["sheet_items"] = sheet_items
    create_result["sheet_write_result"] = write_result
    create_result["sheet_count"] = len(sheet_items)
    create_result["truncated_sheets"] = bool(parse_result.get("truncated_sheets", False))
    create_result["ignored_sheet_names"] = ignored_sheet_names
    if not write_success:
        create_result["error"] = "write workbook sheet items failed"
        return False, create_result

    return True, create_result
